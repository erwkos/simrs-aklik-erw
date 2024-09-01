import ast
import hashlib
import io
import json
import random
import re
import time
import urllib
import uuid
import calendar
import msoffcrypto
import numpy as np
import pandas as pd
import requests
import xlsxwriter
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction, IntegrityError
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.shortcuts import render, redirect
import datetime
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from msoffcrypto.exceptions import InvalidKeyError, DecryptionError
from xlrd import XLRDError
from django.db.models import Max, Sum

from user.models import User
from verifikator.storages import TemporaryStorage
from vpkaak.cekgroupingfix import grouping, pindahrs
from vpkaak.filters import SamplingDataKlaimCBGFilter, RegisterPostKlaimFilter, CBGKoreksiBoaFilter
from vpkaak.forms import RegisterPostKlaimForm, ImportSamplingDataKlaimForm, SamplingDataKlaimCBGForm, \
    FinalisasiRegisterPostKlaimForm, InputNomorBAForm, RegisterPostKlaimSupervisorKPForm
from vpkaak.models import RegisterPostKlaim, SamplingDataKlaimCBG
from vpkaak.choices import StatusChoices, StatusReviewChoices, JenisAuditChoices
from user.decorators import check_device, permissions
from openpyxl import Workbook
from urllib.parse import urlparse, parse_qs


# Create your views here.
@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
def register_post_klaim(request):
    queryset = RegisterPostKlaim.objects.filter(
        user__kantorcabang=request.user.kantorcabang_set.all().first(), is_kp=False).order_by('-created_at')

    # filter
    myFilter = RegisterPostKlaimFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    # pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    if request.POST.get('add_register_post_klaim'):
        form = RegisterPostKlaimForm(data=request.POST)
        if form.is_valid():
            jenis_audit = form.cleaned_data['jenis_audit']

            # validasi field
            if jenis_audit == 'AAK-FKRTL':
                required_fields = ['inisiasi', 'periode_awal', 'periode_akhir', 'surat_tugas']
            elif jenis_audit == 'VPK-FKRTL':
                required_fields = ['bulan_beban']
            else:
                messages.warning(request, 'Jenis audit tidak valid')
                return redirect(request.headers.get('Referer'))

            for field_name in required_fields:
                if not form.cleaned_data.get(field_name):
                    messages.warning(request, f'Field {field_name} harus diisi')
                    return redirect(request.headers.get('Referer'))

            # deteksi VPK dengan buban ganda
            if jenis_audit == JenisAuditChoices.VPK:
                cek_vpk = RegisterPostKlaim.objects.filter(
                    user__kantorcabang=request.user.kantorcabang_set.all().first(),
                    bulan_beban=form.cleaned_data['bulan_beban'], is_kp=False)
                if cek_vpk:
                    messages.warning(request,
                                     f'VPK dengan Bulan beban '
                                     f'{calendar.month_name[cek_vpk[0].bulan_beban.month]}-{cek_vpk[0].bulan_beban.year} '
                                     f'sudah ada')
                    return redirect(request.headers.get('Referer'))

            # Jika validasi berhasil, bentuk nomor register klaim, simpan data dan atur status menjadi Register
            if jenis_audit == 'AAK-FKRTL':
                jenis_audit_prefix = 'AAK-FKRTL'
            elif jenis_audit == 'VPK-FKRTL':
                jenis_audit_prefix = 'VPK-FKRTL'
            else:
                jenis_audit_prefix = 'OTHER'
            bulan_tahun = timezone.now().strftime("%m%y")
            tahun = timezone.now().strftime("%y")
            max_nomor_urut = RegisterPostKlaim.objects.filter(
                nomor_register__contains=f"{jenis_audit_prefix}/{request.user.kantorcabang_set.all().first().kode_cabang}",
                nomor_register__endswith=tahun).aggregate(Max('nomor_register'))
            if max_nomor_urut['nomor_register__max']:
                nomor_urut = int(max_nomor_urut['nomor_register__max'].split('/')[0])
                nomor_urut += 1
            else:
                nomor_urut = 1

            register_post_klaim = form.save(commit=False)
            register_post_klaim.status = 'Register'
            register_post_klaim.user = request.user
            register_post_klaim.nomor_register = f"{nomor_urut:03d}/{jenis_audit_prefix}/{request.user.kantorcabang_set.all().first().kode_cabang}/{bulan_tahun}"
            register_post_klaim.save()

            messages.success(request, 'Register berhasil dibuat')
            return redirect(request.headers.get('Referer'))
    else:
        form = RegisterPostKlaimForm()

    context = {
        'form': form,
        'register_list': queryset,
        'myFilter': myFilter,
    }
    return render(request, 'vpkaak/register.html', context)


def update_register_post_klaim(request, pk):
    pass


@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
def import_sampling_data_klaim(request):
    storage = TemporaryStorage()
    import_form = ImportSamplingDataKlaimForm()
    # verifikator = User.objects.filter(kantorcabang__in=request.user.kantorcabang_set.all(),
    #                                   groups__name='verifikator',
    #                                   is_active=True,
    #                                   is_staff=True)
    if request.method == 'POST' and request.POST.get('action') == 'import':
        import_form = ImportSamplingDataKlaimForm(files=request.FILES, data=request.POST)
        if import_form.is_valid():
            list_data_import = ['Nokapst', 'Tgldtgsjp', 'Tglplgsjp', 'Nosjp', 'Tglpelayanan', 'Kdkrlayan', 'Kdkclayan',
                                'Nmkclayan', 'Kddati2Layan', 'Nmdati2Layan', 'Kdppklayan', 'Nmppklayan', 'Nmtkp',
                                'Kdinacbgs', 'Nminacbgs', 'Kddiagprimer', 'Nmdiagprimer', 'Diagsekunder', 'Procedure',
                                'Klsrawat', 'Nmjnspulang', 'kddokter', 'nmdokter', 'Umur', 'kdsa', 'kdsd',
                                'deskripsisd', 'kdsi', 'deskripsisi', 'kdsp', 'deskripsisp', 'kdsr', 'deskripsisr',
                                'Tarifgroup', 'tarifsa', 'tarifsd', 'tarifsi', 'tarifsp', 'tarifsr', 'Biayaverifikasi',
                                'Kodersmenkes', 'Kelasrsmenkes', 'Jkpst', 'redflag']
            nomor_register = import_form.cleaned_data.get('register')
            register = RegisterPostKlaim.objects.get(nomor_register=nomor_register)
            file_name = f'{uuid.uuid4()}-{int(round(time.time() * 1000))}.xlsx'
            storage.save(name=file_name, content=import_form.cleaned_data.get('file'))
            get_password = request.POST.get('password')
            if get_password != '':
                unlocked_file = io.BytesIO()
                try:
                    with open(storage.path(name=file_name), "rb") as file:
                        excel_file = msoffcrypto.OfficeFile(file)
                        excel_file.load_key(password=get_password)
                        excel_file.decrypt(unlocked_file)
                    data_frame = pd.read_excel(unlocked_file, usecols=list_data_import)
                except InvalidKeyError:
                    messages.warning(request, 'Password File Excel yang Anda masukkan salah!')
                    return redirect('/vpkaak/import-sampling-data-klaim')
                except DecryptionError:
                    messages.warning(request, 'File Excel seharusnya tidak memiliki password!')
                    return redirect('/vpkaak/import-sampling-data-klaim')
                except Exception as e:
                    messages.warning(request, f'Terjadi Kesalahan pada saat import File. Keterangan error : {e}')
                    return redirect('/vpkaak/import-sampling-data-klaim')
            else:
                try:
                    data_frame = pd.read_excel(storage.path(name=file_name), usecols=list_data_import)
                except XLRDError:
                    messages.warning(request,
                                     f'File yang Anda import memiliki password. Silahkan masukkan password file!')
                    return redirect('/vpkaak/import-sampling-data-klaim')
                except Exception as e:
                    messages.warning(request, f'Terjadi kesalahan pada saat import File. Keterangan error : {e}')
                    return redirect('/vpkaak/import-sampling-data-klaim')
            data_frame = data_frame.replace(np.nan, None)
            data_frame['register'] = register
            data_frame['Kdkclayan'] = data_frame['Kdkclayan'].astype('str').apply(lambda x: x.zfill(4))
            # data_frame['Kdkclayan'] = data_frame['Kdkclayan'].zfill(4)
            # data_frame['faskes'] = register.faskes
            # data_frame['TGLPULANG'] = pd.to_datetime(data_frame['TGLPULANG'])
            # data_frame['bupel'] = data_frame['TGLPULANG'].dt.to_period('M').dt.to_timestamp()
            # data_frame['bupel'] = pd.to_datetime(data_frame['bupel'])
            register.password_file_excel = get_password
            register.save()

            # ini jadi ga semua retrieval dan penulisan di database tidak semua di kasih transaction.atomic
            # with transaction.atomic():
            try:
                obj_list = []
                for _, row in data_frame.iterrows():
                    data_klaim = SamplingDataKlaimCBG()
                    try:
                        data_klaim = SamplingDataKlaimCBG(**dict(row))
                        data_klaim.full_clean()  # Validate the object
                        obj_list.append(data_klaim)
                        # ini saya hapus mas jadi ini tiap iterasi buat list bulk jadi bikin space di db ke lock for a while kalo kebanyakan
                        # DataKlaimCBG.objects.bulk_create(obj_list)
                    except TypeError as e:
                        messages.warning(request,
                                         f'Terjadi kesalahan pada saat import File. Keterangan error : {e}')
                        return redirect('/vpkaak/import-sampling-data-klaim')
                    except Exception as e:
                        messages.warning(request,
                                         f'Kesalahan terjadi pada No SEP : {data_klaim.Nosjp}. Keterangan error : {e}')
                        return redirect('/vpkaak/import-sampling-data-klaim')

                # ini data penulisan dipisah mas
                with transaction.atomic():
                    SamplingDataKlaimCBG.objects.bulk_create(obj_list)
                    # ini data retrieval
                    valid_data = SamplingDataKlaimCBG.objects.filter(id__in=[obj.id for obj in obj_list])
                    # invalid_data = SamplingDataKlaimCBG.objects.filter(id__in=[obj.id for obj in obj_list
                    #                                                            if
                    #                                                            obj.NOSEP[
                    #                                                            :8] != register.faskes.kode_ppk or
                    #                                                            obj.bupel.month != register.bulan_pelayanan.month or
                    #                                                            obj.bupel.year != register.bulan_pelayanan.year])
                    df_valid = pd.DataFrame.from_records(valid_data.values())
                    total_data_valid = len(df_valid)
                    # df_invalid = pd.DataFrame.from_records(invalid_data.values())
                    # total_data_invalid = len(df_invalid)
                    transaction.set_rollback(True)

            except IntegrityError as e:
                messages.info(request, f'Terjadi kesalahan saat mencoba menyimpan data : {e}')
                return redirect('/vpkaak/import-sampling-data-klaim')
            except Exception as e:
                messages.info(request, f'Kesalahan terjadi pada saat import File. Keterangan error : {e}')
                return redirect('/vpkaak/import-sampling-data-klaim')

            return render(request, 'vpkaak/preview_sampling_data_import_cbg.html',
                          {'preview_data_valid': df_valid,
                           'total_data_valid': total_data_valid,
                           # 'preview_data_invalid': df_invalid,
                           # 'total_data_invalid': total_data_invalid,
                           'file_name': file_name,
                           'register': nomor_register,
                           'password': get_password})

    if request.method == 'POST' and request.POST.get('action') == 'confirm':
        list_data_import = ['Nokapst', 'Tgldtgsjp', 'Tglplgsjp', 'Nosjp', 'Tglpelayanan', 'Kdkrlayan', 'Kdkclayan',
                            'Nmkclayan', 'Kddati2Layan', 'Nmdati2Layan', 'Kdppklayan', 'Nmppklayan', 'Nmtkp',
                            'Kdinacbgs', 'Nminacbgs', 'Kddiagprimer', 'Nmdiagprimer', 'Diagsekunder', 'Procedure',
                            'Klsrawat', 'Nmjnspulang', 'kddokter', 'nmdokter', 'Umur', 'kdsa', 'kdsd',
                            'deskripsisd', 'kdsi', 'deskripsisi', 'kdsp', 'deskripsisp', 'kdsr', 'deskripsisr',
                            'Tarifgroup', 'tarifsa', 'tarifsd', 'tarifsi', 'tarifsp', 'tarifsr', 'Biayaverifikasi',
                            'Kodersmenkes', 'Kelasrsmenkes', 'Jkpst', 'redflag']
        file_name = request.POST.get('file_name')
        nomor_register = request.POST.get('register')
        get_password = request.POST.get('password')
        register = RegisterPostKlaim.objects.get(nomor_register=nomor_register)
        if get_password != '':
            unlocked_file = io.BytesIO()
            try:
                with open(storage.path(name=file_name), "rb") as file:
                    excel_file = msoffcrypto.OfficeFile(file)
                    excel_file.load_key(password=request.POST.get('password'))
                    excel_file.decrypt(unlocked_file)
                data_frame = pd.read_excel(unlocked_file, usecols=list_data_import)
            except InvalidKeyError:
                messages.warning(request, 'Password File Excel yang Anda masukkan salah!')
                return redirect('/vpkaak/import-sampling-data-klaim')
            except DecryptionError:
                messages.warning(request, 'File Excel seharusnya tidak memiliki password!')
                return redirect('/vpkaak/import-sampling-data-klaim')
            except Exception as e:
                messages.warning(request, f'Terjadi Kesalahan pada saat import File. Keterangan error : {e}')
                return redirect('/vpkaak/import-sampling-data-klaim')
        else:
            try:
                data_frame = pd.read_excel(storage.path(name=file_name), usecols=list_data_import)
            except XLRDError:
                messages.warning(request,
                                 f'File yang Anda import memiliki password. Silahkan masukkan password file!')
                return redirect('/vpkaak/import-sampling-data-klaim')
            except Exception as e:
                messages.warning(request, f'Terjadi kesalahan pada saat import File. Keterangan error : {e}')
                return redirect('/vpkaak/import-sampling-data-klaim')

        data_frame = data_frame.replace(np.nan, None)
        data_frame['register'] = register
        data_frame['Kdkclayan'] = data_frame['Kdkclayan'].astype('str').apply(lambda x: x.zfill(4))
        # saya nambahin try and except disini mas jadi ga ngasih 500 tapi ngasih messages error kalo ada kesalahan
        try:
            with transaction.atomic():
                SamplingDataKlaimCBG.objects.bulk_create(
                    [SamplingDataKlaimCBG(**dict(row[1])) for row in data_frame.iterrows()])
            # register jadi verifikasi
            register.status = StatusChoices.Verifikasi
            register.save()
            messages.success(request, "Data Klaim Berhasil Di-import")
        except Exception as e:
            messages.error(request, f"Terjadi kesalahan: {e}")
            return redirect('/vpkaak/import-sampling-data-klaim')

    context = {
        'import_form': import_form,
        # 'verifikator': verifikator,
    }
    return render(request, 'vpkaak/import_data.html', context)


@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
def review(request):
    queryset_sampling_klaim = SamplingDataKlaimCBG.objects.filter(
        register__user__kantorcabang=request.user.kantorcabang_set.all().first(), is_from_kp=False, is_final=False).order_by('created_at')

    # filter
    myFilter = SamplingDataKlaimCBGFilter(request.GET, queryset=queryset_sampling_klaim)
    queryset = myFilter.qs

    # pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    export = request.GET.get('export')
    if export == 'export':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-samplingdataklaimcbg.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = xlsxwriter.Workbook(response, {'in_memory': True})
        worksheet = workbook.add_worksheet('Sampling Data Klaim CBG')

        # # Get active worksheet/tab
        # worksheet = workbook.active
        # worksheet.title = 'Sampling Data Klaim CBG'

        # Define the titles for columns
        columns = [
            'status',
            'register',
            'jenisaudit',
            'Nokapst',
            'Tgldtgsjp',
            'Tglplgsjp',
            'Nosjp',
            'Tglpelayanan',
            'Kdkrlayan',
            'Kdkclayan',
            'Nmkclayan',
            'Kddati2Layan',
            'Nmdati2Layan',
            'Kdppklayan',
            'Nmppklayan',
            'Nmtkp',
            'Kdinacbgs',
            'Nminacbgs',
            'Kddiagprimer',
            'Nmdiagprimer',
            'Diagsekunder',
            'Procedure',
            'Klsrawat',
            'Nmjnspulang',
            'kddokter',
            'nmdokter',
            'Umur',
            'kdsa',
            'kdsd',
            'deskripsisd',
            'kdsi',
            'deskripsisi',
            'kdsp',
            'deskripsisp',
            'kdsr',
            'deskripsisr',
            'Tarifgroup',
            'tarifsa',
            'tarifsd',
            'tarifsi',
            'tarifsp',
            'tarifsr',
            'Biayaverifikasi',
            'verifikator',
            'tglreview',
            'keterangan',
            'redflag',
            'deskripsi_redflag',
        ]

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns):
            worksheet.write(0, col_num, column_title)
            # cell = worksheet.cell(row=row_num, column=col_num)
            # cell.value = column_title

        # Iterate through all
        row_num = 1
        for queryset in queryset_sampling_klaim:
            row = [
                queryset.status,
                queryset.register.nomor_register,
                queryset.register.jenis_audit,
                queryset.Nokapst,
                queryset.Tgldtgsjp,
                queryset.Tglplgsjp,
                queryset.Nosjp,
                queryset.Tglpelayanan,
                queryset.Kdkrlayan,
                queryset.Kdkclayan,
                queryset.Nmkclayan,
                queryset.Kddati2Layan,
                queryset.Nmdati2Layan,
                queryset.Kdppklayan,
                queryset.Nmppklayan,
                queryset.Nmtkp,
                queryset.Kdinacbgs,
                queryset.Nminacbgs,
                queryset.Kddiagprimer,
                queryset.Nmdiagprimer,
                queryset.Diagsekunder,
                queryset.Procedure,
                queryset.Klsrawat,
                queryset.Nmjnspulang,
                queryset.kddokter,
                queryset.nmdokter,
                queryset.Umur,
                queryset.kdsa,
                queryset.kdsd,
                queryset.deskripsisd,
                queryset.kdsi,
                queryset.deskripsisi,
                queryset.kdsp,
                queryset.deskripsisp,
                queryset.kdsr,
                queryset.deskripsisr,
                queryset.Tarifgroup,
                queryset.tarifsa,
                queryset.tarifsd,
                queryset.tarifsi,
                queryset.tarifsp,
                queryset.tarifsr,
                queryset.Biayaverifikasi,
                queryset.redflag,
                queryset.deskripsi_redflag,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row):
                if isinstance(cell_value, datetime.date):
                    worksheet.write_datetime(row_num, col_num, cell_value,
                                             workbook.add_format({'num_format': 'yyyy-mm-dd'}))
                else:
                    worksheet.write(row_num, col_num, cell_value)
                # cell = worksheet.cell(row=row_num, column=col_num)
                # cell.value = cell_value
            row_num += 1

        workbook.close()
        return response
    context = {
        'data_klaim': queryset,
        'myFilter': myFilter,
    }
    return render(request, 'vpkaak/review.html', context)


@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
def reviewkp(request):
    kantor_cabang = request.user.kantorcabang_set.all().first().kode_cabang
    # queryset = SamplingDataKlaimCBG.objects.filter(Kdkclayan=kantor_cabang, is_from_kp=True).order_by('created_at')
    queryset_sampling_klaim = SamplingDataKlaimCBG.objects.filter(Kdkclayan=kantor_cabang, is_from_kp=True, is_final=False).order_by('created_at')

    # filter
    myFilter = SamplingDataKlaimCBGFilter(request.GET, queryset=queryset_sampling_klaim)
    queryset = myFilter.qs

    # pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    export = request.GET.get('export')
    if export == 'export':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-samplingdataklaimcbg.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = xlsxwriter.Workbook(response, {'in_memory': True})
        worksheet = workbook.add_worksheet('Sampling Data Klaim CBG')

        # # Get active worksheet/tab
        # worksheet = workbook.active
        # worksheet.title = 'Sampling Data Klaim CBG'

        # Define the titles for columns
        columns = [
            'status',
            'register',
            'jenisaudit',
            'Nokapst',
            'Tgldtgsjp',
            'Tglplgsjp',
            'Nosjp',
            'Tglpelayanan',
            'Kdkrlayan',
            'Kdkclayan',
            'Nmkclayan',
            'Kddati2Layan',
            'Nmdati2Layan',
            'Kdppklayan',
            'Nmppklayan',
            'Nmtkp',
            'Kdinacbgs',
            'Nminacbgs',
            'Kddiagprimer',
            'Nmdiagprimer',
            'Diagsekunder',
            'Procedure',
            'Klsrawat',
            'Nmjnspulang',
            'kddokter',
            'nmdokter',
            'Umur',
            'kdsa',
            'kdsd',
            'deskripsisd',
            'kdsi',
            'deskripsisi',
            'kdsp',
            'deskripsisp',
            'kdsr',
            'deskripsisr',
            'Tarifgroup',
            'tarifsa',
            'tarifsd',
            'tarifsi',
            'tarifsp',
            'tarifsr',
            'Biayaverifikasi',
            'verifikator',
            'tglreview',
            'keterangan',
            'redflag',
            'deskripsi_redflag',
        ]

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns):
            worksheet.write(0, col_num, column_title)
            # cell = worksheet.cell(row=row_num, column=col_num)
            # cell.value = column_title

        # Iterate through all
        row_num = 1
        for queryset in queryset_sampling_klaim:
            row = [
                queryset.status,
                queryset.register.nomor_register,
                queryset.register.jenis_audit,
                queryset.Nokapst,
                queryset.Tgldtgsjp,
                queryset.Tglplgsjp,
                queryset.Nosjp,
                queryset.Tglpelayanan,
                queryset.Kdkrlayan,
                queryset.Kdkclayan,
                queryset.Nmkclayan,
                queryset.Kddati2Layan,
                queryset.Nmdati2Layan,
                queryset.Kdppklayan,
                queryset.Nmppklayan,
                queryset.Nmtkp,
                queryset.Kdinacbgs,
                queryset.Nminacbgs,
                queryset.Kddiagprimer,
                queryset.Nmdiagprimer,
                queryset.Diagsekunder,
                queryset.Procedure,
                queryset.Klsrawat,
                queryset.Nmjnspulang,
                queryset.kddokter,
                queryset.nmdokter,
                queryset.Umur,
                queryset.kdsa,
                queryset.kdsd,
                queryset.deskripsisd,
                queryset.kdsi,
                queryset.deskripsisi,
                queryset.kdsp,
                queryset.deskripsisp,
                queryset.kdsr,
                queryset.deskripsisr,
                queryset.Tarifgroup,
                queryset.tarifsa,
                queryset.tarifsd,
                queryset.tarifsi,
                queryset.tarifsp,
                queryset.tarifsr,
                queryset.Biayaverifikasi,
                queryset.redflag,
                queryset.deskripsi_redflag,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row):
                if isinstance(cell_value, datetime.date):
                    worksheet.write_datetime(row_num, col_num, cell_value,
                                             workbook.add_format({'num_format': 'yyyy-mm-dd'}))
                else:
                    worksheet.write(row_num, col_num, cell_value)
                # cell = worksheet.cell(row=row_num, column=col_num)
                # cell.value = cell_value
            row_num += 1

        workbook.close()
        return response
    context = {
        'data_klaim': queryset,
        'myFilter': myFilter,
    }
    return render(request, 'vpkaak/reviewkp.html', context)

@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
@csrf_exempt
def is_bayi(request):
    data = request.POST
    queryset = SamplingDataKlaimCBG.objects.filter(
        register__user__kantorcabang=request.user.kantorcabang_set.all().first())
    instance = queryset.get(pk=data.get('id_is_bayi'))
    initial = False
    if data.get('is_bayi') == 1 or data.get('is_bayi') == '1':
        instance.is_bayi = True
    else:
        instance.is_bayi = False
    instance.save()
    # print(instance.is_bayi)
    return HttpResponse('is bayi saved')


@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
@csrf_exempt
def cek_grouping(request):
    data = request.POST
    special_prosedure = data.get('special_prosedure')
    special_prosthesis = data.get('special_prosthesis')
    special_investigation = data.get('special_investigation')
    special_drug = data.get('special_drug')

    tipe_rawat = data['tipe_rawat']
    if tipe_rawat == 'RJTL':
        tipe_rawat = 'outpatient'
    elif tipe_rawat == 'RITL':
        tipe_rawat = 'inpatient'
    # else:
    #     tipe_rawat = 'outpatient'

    kelas_rawat = data['kelas_rawat']
    if kelas_rawat == 'Kelas III':
        kelas_rawat = 'kelas_3'
    elif kelas_rawat == 'Kelas II':
        kelas_rawat = 'kelas_2'
    elif kelas_rawat == 'Kelas I':
        kelas_rawat = 'kelas_1'
    else:
        kelas_rawat = 'kelas_3'

    # Formatted diagnosis
    diagnosis_codes = []
    if 'diagnosis_utama' in data:
        kode = data['diagnosis_utama'].split(' - ')[0]
        if len(kode) > 3 and '.' not in kode:
            kode = kode[:3] + '.' + kode[3:]
        diagnosis_codes.append(kode)

    sekunder_keys = [key for key in data if key.startswith('diagnosis_sekunder')]
    if sekunder_keys:
        for key in sekunder_keys:
            if data[key]:
                kode = data[key].split(' - ')[0]
                if len(kode) > 3 and '.' not in kode:
                    kode = kode[:3] + '.' + kode[3:]
                diagnosis_codes.append(kode)

    formatted_diagnosis_codes = '|'.join(diagnosis_codes)

    # Formatted procedure
    procedure_codes = []
    procedure_keys = [key for key in data if key.startswith('procedure')]
    if procedure_keys:
        for key in procedure_keys:
            if data[key]:
                kode = data[key].split(' - ')[0]
                if len(kode) > 2 and '.' not in kode:
                    kode = kode[:2] + '.' + kode[2:]
                procedure_codes.append(kode)

    formatted_procedure_codes = '|'.join(procedure_codes)

    is_bayi = data.get('is_bayi')
    if is_bayi == 'on':
        is_bayi = True
    else:
        is_bayi = False

    # if 'tanggallahirbayi' in data:
    #     tanggallahirbayi = data['tanggallahirbayi']
    # else:
    #     tanggallahirbayi = '1911-11-11'
    #
    # if 'beratbayi' in data:
    #     birth_weight = data['beratbayi']
    # else:
    #     birth_weight = 0

    tanggallahirbayi = data['tanggallahirbayi']
    Tgldtgsjp = data['Tgldtgsjp']
    Tglplgsjp = data['Tglplgsjp']
    ambil_tarif = data['Kodersmenkes']
    rs_tarif = pindahrs(ambil_tarif)[0]
    # print(list_tarif[0])
    # print(type(rs_tarif))
    #
    # print('Gender: ', data['gender'])
    # print('Umur: ', data['umur'])
    # print('No. SEP: ', data['no_sep'])
    # print('No. Peserta: ', data['no_peserta'])
    # print('Tipe Rawat: ', tipe_rawat)
    # print('Kelas Rawat: ', kelas_rawat)
    # print('Diagnosis: ', formatted_diagnosis_codes)
    # print('Prosedur: ', formatted_procedure_codes)
    # print('Special Prosedure: ', special_prosedure)
    # print('Special Prosthesis: ', special_prosthesis)
    # print('Special Investigation: ', special_investigation)
    # print('Special Drug: ', special_drug)
    # print('Is Bayi: ', is_bayi)
    # print('Birth Weight: ', str(data['beratbayi']))
    # print('Birth Date: ', str(tanggallahirbayi))
    # print('Admission Date: ', str(Tgldtgsjp))
    # print('Discharge Date: ', str(Tglplgsjp))
    # print('Tarif RS: ', rs_tarif)

    # group_code = None
    # nama_list = []
    # data_list = []
    # data_list_match_topup_sp_pattern = []
    # data_list_match_topup_sr_pattern = []
    # data_list_match_topup_si_pattern = []
    # data_list_match_topup_sd_pattern = []
    try:
        (group_code, nama_list, data_list, data_list_match_topup_sp_pattern,
         data_list_match_topup_sr_pattern, data_list_match_topup_si_pattern,
         data_list_match_topup_sd_pattern) = grouping(
            gender=data['gender'],
            umur=data['umur'],
            no_sep=data['no_sep'],
            no_peserta=data['no_peserta'],
            tipe_rawat=tipe_rawat,
            kelas_rawat=kelas_rawat,
            diagnosa=formatted_diagnosis_codes,
            prosedur=formatted_procedure_codes,
            sp=special_prosedure,
            sr=special_prosthesis,
            si=special_investigation,
            sd=special_drug,
            is_bayi=is_bayi,
            birth_weight=str(data['beratbayi']),
            # birth_weight=birth_weight,
            tanggallahirbayi=str(tanggallahirbayi),
            Tgldtgsjp=str(Tgldtgsjp),
            Tglplgsjp=str(Tglplgsjp),
            rs_tarif=rs_tarif,
        )
    except TypeError as e:
        print(f"TypeError occurred: {e}")
        # messages.warning(request, f"TypeError occurred: {e}")
        # Handle the error, log it, or take appropriate action
        # Ensure variables are assigned or reset to default values
        group_code = None
        nama_list = []
        data_list = []
        data_list_match_topup_sp_pattern = []
        data_list_match_topup_sr_pattern = []
        data_list_match_topup_si_pattern = []
        data_list_match_topup_sd_pattern = []
    except Exception as e:
        print(f"An unexpected error occurred: {e}")
        # messages.warning(request, f"TypeError occurred: {e}")
        # Handle other exceptions as needed
        # Ensure variables are assigned or reset to default values
        group_code = None
        nama_list = []
        data_list = []
        data_list_match_topup_sp_pattern = []
        data_list_match_topup_sr_pattern = []
        data_list_match_topup_si_pattern = []
        data_list_match_topup_sd_pattern = []
    # Now you can safely use the unpacked variables

    response_data = {
        'group_code': group_code,
        'nama_list': nama_list,
        'data_list': data_list,
        'data_list_match_topup_sp_pattern': data_list_match_topup_sp_pattern or '<option value="None">None</option>',
        'data_list_match_topup_sr_pattern': data_list_match_topup_sr_pattern or '<option value="None">None</option>',
        'data_list_match_topup_si_pattern': data_list_match_topup_si_pattern or '<option value="None">None</option>',
        'data_list_match_topup_sd_pattern': data_list_match_topup_sd_pattern or '<option value="None">None</option>',
    }
    return JsonResponse(response_data)


@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
def update_review(request, pk):
    data = request.POST
    queryset = SamplingDataKlaimCBG.objects.filter(
        register__user__kantorcabang=request.user.kantorcabang_set.all().first(), is_from_kp=False, is_final=False)
    instance = queryset.get(pk=pk)
    data_klaim_form = SamplingDataKlaimCBGForm(instance=instance)
    if data:
        next = request.POST.get('next', '/')
        if data.get('keterangan_review') and data.get('status') == StatusReviewChoices.Sesuai:
            instance.keterangan_review = data['keterangan_review']
            instance.status = StatusReviewChoices.Sesuai
            instance.verifikator_review = request.user
            instance.tgl_review = datetime.datetime.today()
            instance.save()
            messages.success(request, f'NO SEP {instance.Nosjp} berhasil diupdate.')
            return HttpResponseRedirect(next)
        elif data.get('keterangan_review'):
            try:
                kode_diagprimer_koreksi = data['diagnosis_utama'].split(' - ')[0]
                nama_diagprimer_koreksi = data['diagnosis_utama'].split(' - ')[1]

                # Mengambil semua diagnosis sekunder dari request.POST
                data_payload = {key: value for key, value in request.POST.items() if key.startswith('diagnosis_sekunder')}
                # print(data_payload)

                # Mengubah data payload menjadi format yang diinginkan
                diagnosis_sekunder_list = []
                for key in sorted(data_payload.keys()):
                    if data_payload[key]:
                        code, desc = data_payload[key].split(' - ')
                        diagnosis_sekunder_list.append(f"{code.replace('.', '')} - {desc}")

                # Menggabungkan hasilnya dengan pemisah ';'
                diagsekunder_koreksi = ';'.join(diagnosis_sekunder_list)

                # Mengambil semua diagnosis sekunder dari request.POST
                data_payload_procedure = {key: value for key, value in request.POST.items() if key.startswith('procedure')}
                # print(data_payload_procedure)

                # Mengubah data payload menjadi format yang diinginkan
                procedure_list = []
                for key in sorted(data_payload_procedure.keys()):
                    if data_payload_procedure[key]:
                        code, desc = data_payload_procedure[key].split(' - ')
                        procedure_list.append(f"{code.replace('.', '')} - {desc}")

                # Menggabungkan hasilnya dengan pemisah ';'
                procedure_koreksi = ';'.join(procedure_list)

                if data['inacbg'] == '-':
                    instance.Kdinacbgs_koreksi = None
                    instance.Nminacbgs_koreksi = None
                else:
                    kode_inacbg_koreksi = data['inacbg'].split(' - ')[0]
                    nama_inacbg_koreksi = data['inacbg'].split(' - ')[1]
                    instance.Kdinacbgs_koreksi = kode_inacbg_koreksi
                    instance.Nminacbgs_koreksi = nama_inacbg_koreksi
                instance.Kddiagprimer_koreksi = kode_diagprimer_koreksi
                instance.Nmdiagprimer_koreksi = nama_diagprimer_koreksi
                instance.Diagsekunder_koreksi = diagsekunder_koreksi
                instance.Procedure_koreksi = procedure_koreksi
                instance.biaya_koreksi = int(data.get('tarif').replace(",", ""))
                instance.Klsrawat_koreksi = data.get('kelas_rawat')
                instance.keterangan_review = data.get('keterangan_review')
                instance.status = StatusReviewChoices.TidakSesuai
                instance.is_koreksi = True
                instance.verifikator_review = request.user
                instance.jenis_fraud = data.get('jenis_fraud')
                print(request.POST.get('jenis_fraud'))

                if data.get('tipe_rawat') == 'outpatient':
                    tipe_rawat = "RJTL"
                elif data.get('tipe_rawat') == 'inpatient':
                    tipe_rawat = "RITL"
                else:
                    tipe_rawat = 'Tidak Diketahui'
                instance.Nmtkp_koreksi = tipe_rawat

                is_bayi = data.get('is_bayi')
                # print(is_bayi)
                if is_bayi == 'True':
                    is_bayi = True
                else:
                    is_bayi = False

                instance.is_bayi = is_bayi
                instance.beratbayi = data.get('beratbayi')
                if data.get('tanggallahirbayi') != '':
                    instance.tanggallahirbayi = data.get('tanggallahirbayi')
                # instance.tanggallahirbayi = data.get('tanggallahirbayi')
                instance.tgl_review = datetime.datetime.today()

                instance.save()
                messages.success(request, f'NO SEP {instance.Nosjp} berhasil diupdate.')
                return HttpResponseRedirect(next)
            except Exception as e:
                messages.warning(request, f'Terdapat Kesalahan. Lakukan Cek Grouping sebelum simpan. Silahkan Coba Lagi. : {e}')
                return redirect(request.headers.get('Referer'))
        else:
            messages.warning(request, 'Keterangan review harus diisi!')
    context = {
        'data_klaim': instance,
        'data_klaim_form': data_klaim_form,
    }
    return render(request, 'vpkaak/update_review.html', context)


@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
def finalisasi_register_post_klaim(request):
    queryset = RegisterPostKlaim.objects.filter(
        user__kantorcabang=request.user.kantorcabang_set.all().first(),
        status=StatusChoices.Verifikasi)

    # filter
    myFilter = RegisterPostKlaimFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    # pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'register_list': queryset,
        'myFilter': myFilter,
    }
    return render(request, 'vpkaak/finalisasi_register_post_klaim.html', context)


@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
def update_review_kp(request, pk):
    kantor_cabang = request.user.kantorcabang_set.all().first().kode_cabang
    queryset = SamplingDataKlaimCBG.objects.filter(Kdkclayan=kantor_cabang, is_from_kp=True, is_final=False).order_by('created_at')

    data = request.POST
    instance = queryset.get(pk=pk)
    data_klaim_form = SamplingDataKlaimCBGForm(instance=instance)
    if data:
        next = request.POST.get('next', '/')
        if data.get('keterangan_review') and data.get('status') == StatusReviewChoices.Sesuai:
            instance.keterangan_review = data['keterangan_review']
            instance.status = StatusReviewChoices.Sesuai
            instance.verifikator_review = request.user
            instance.tgl_review = datetime.datetime.today()
            instance.save()
            messages.success(request, f'NO SEP {instance.Nosjp} berhasil diupdate.')
            return HttpResponseRedirect(next)
        elif data.get('keterangan_review'):
            try:
                kode_diagprimer_koreksi = data['diagnosis_utama'].split(' - ')[0]
                nama_diagprimer_koreksi = data['diagnosis_utama'].split(' - ')[1]

                # Mengambil semua diagnosis sekunder dari request.POST
                data_payload = {key: value for key, value in request.POST.items() if key.startswith('diagnosis_sekunder')}
                # print(data_payload)

                # Mengubah data payload menjadi format yang diinginkan
                diagnosis_sekunder_list = []
                for key in sorted(data_payload.keys()):
                    if data_payload[key]:
                        code, desc = data_payload[key].split(' - ')
                        diagnosis_sekunder_list.append(f"{code.replace('.', '')} - {desc}")

                # Menggabungkan hasilnya dengan pemisah ';'
                diagsekunder_koreksi = ';'.join(diagnosis_sekunder_list)

                # Mengambil semua diagnosis sekunder dari request.POST
                data_payload_procedure = {key: value for key, value in request.POST.items() if key.startswith('procedure')}
                # print(data_payload_procedure)

                # Mengubah data payload menjadi format yang diinginkan
                procedure_list = []
                for key in sorted(data_payload_procedure.keys()):
                    if data_payload_procedure[key]:
                        code, desc = data_payload_procedure[key].split(' - ')
                        procedure_list.append(f"{code.replace('.', '')} - {desc}")

                # Menggabungkan hasilnya dengan pemisah ';'
                procedure_koreksi = ';'.join(procedure_list)

                if data['inacbg'] == '-':
                    instance.Kdinacbgs_koreksi = None
                    instance.Nminacbgs_koreksi = None
                else:
                    kode_inacbg_koreksi = data['inacbg'].split(' - ')[0]
                    nama_inacbg_koreksi = data['inacbg'].split(' - ')[1]
                    instance.Kdinacbgs_koreksi = kode_inacbg_koreksi
                    instance.Nminacbgs_koreksi = nama_inacbg_koreksi
                instance.Kddiagprimer_koreksi = kode_diagprimer_koreksi
                instance.Nmdiagprimer_koreksi = nama_diagprimer_koreksi
                instance.Diagsekunder_koreksi = diagsekunder_koreksi
                instance.Procedure_koreksi = procedure_koreksi
                instance.biaya_koreksi = int(data.get('tarif').replace(",", ""))
                instance.Klsrawat_koreksi = data.get('kelas_rawat')
                instance.keterangan_review = data.get('keterangan_review')
                instance.status = StatusReviewChoices.TidakSesuai
                instance.is_koreksi = True
                instance.verifikator_review = request.user
                instance.jenis_fraud = data.get('jenis_fraud')

                if data.get('tipe_rawat') == 'outpatient':
                    tipe_rawat = "RJTL"
                elif data.get('tipe_rawat') == 'inpatient':
                    tipe_rawat = "RITL"
                else:
                    tipe_rawat = 'Tidak Diketahui'
                instance.Nmtkp_koreksi = tipe_rawat

                is_bayi = data.get('is_bayi')

                if is_bayi == 'True':
                    is_bayi = True
                else:
                    is_bayi = False

                instance.is_bayi = is_bayi
                instance.beratbayi = data.get('beratbayi')
                if data.get('tanggallahirbayi') != '':
                    instance.tanggallahirbayi = data.get('tanggallahirbayi')
                # instance.tanggallahirbayi = data.get('tanggallahirbayi')
                instance.tgl_review = datetime.datetime.today()

                instance.save()
                messages.success(request, f'NO SEP {instance.Nosjp} berhasil diupdate.')
                return HttpResponseRedirect(next)
            except Exception as e:
                messages.warning(request, f'Terdapat Kesalahan, Lakukan Cek Grouping sebelum simpan. Silahkan Coba Lagi. : {e}')
                return redirect(request.headers.get('Referer'))
        else:
            messages.warning(request, 'Keterangan review harus diisi!')
    context = {
        'data_klaim': instance,
        'data_klaim_form': data_klaim_form,
    }
    return render(request, 'vpkaak/update_review.html', context)


@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
def finalisasi_register_post_klaim(request):
    queryset = RegisterPostKlaim.objects.filter(
        user__kantorcabang=request.user.kantorcabang_set.all().first(),
        status=StatusChoices.Verifikasi)

    # filter
    myFilter = RegisterPostKlaimFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    # pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'register_list': queryset,
        'myFilter': myFilter,
    }
    return render(request, 'vpkaak/finalisasi_register_post_klaim.html', context)


@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
def update_finalisasi_register_post_klaim(request, pk):
    queryset = RegisterPostKlaim.objects.filter(user__kantorcabang=request.user.kantorcabang_set.all().first())
    instance = queryset.get(pk=pk)

    # count
    # jumlah_sesuai = 0
    # jumlah_tidak_sesuai = 0
    # jumlah_belum_review = 0
    # total_sampling = 0
    # biaya_sesuai = 0
    # biaya_tidak_sesuai = 0
    # biaya_belum_review = 0
    # biaya_sampling = 0
    try:
        sampling_data_klaim = SamplingDataKlaimCBG.objects.filter(register=instance)
        jumlah_sesuai = sampling_data_klaim.filter(status=StatusReviewChoices.Sesuai).count()
        jumlah_tidak_sesuai = sampling_data_klaim.filter(status=StatusReviewChoices.TidakSesuai).count()
        jumlah_belum_review = sampling_data_klaim.filter(status=StatusReviewChoices.Belum).count()
        total_sampling = sampling_data_klaim.count()

        biaya_sesuai = sampling_data_klaim.filter(status=StatusReviewChoices.Sesuai).aggregate(Sum('Biayaverifikasi'))[
            'Biayaverifikasi__sum']
        biaya_tidak_sesuai = \
            sampling_data_klaim.filter(status=StatusReviewChoices.TidakSesuai).aggregate(Sum('Biayaverifikasi'))[
                'Biayaverifikasi__sum']
        biaya_belum_review = \
            sampling_data_klaim.filter(status=StatusReviewChoices.Belum).aggregate(Sum('Biayaverifikasi'))[
                'Biayaverifikasi__sum']
        biaya_sampling = sampling_data_klaim.aggregate(Sum('Biayaverifikasi'))['Biayaverifikasi__sum']
    except Exception as e:
        messages.warning(request, f'Pengecekan Data Sampling Error, {e}')
        return redirect(request.headers.get('Referer'))

    if request.method == 'POST':
        form = FinalisasiRegisterPostKlaimForm(request.POST, instance=instance)
        if form.is_valid():
            if jumlah_belum_review != 0:
                messages.warning(request, 'Masih terdapat data sampling klaim yang belum direview')
                return redirect(request.headers.get('Referer'))
            obj = form.save()
            obj.is_final = True
            obj.save()
            sampling_data_klaim.update(is_final=True)
            messages.success(request, 'Finalisasi Register berhasil.')
    else:
        form = FinalisasiRegisterPostKlaimForm(instance=instance)

    context = {
        'form': form,
        'jumlah_sesuai': jumlah_sesuai,
        'jumlah_tidak_sesuai': jumlah_tidak_sesuai,
        'jumlah_belum_review': jumlah_belum_review,
        'total_sampling': total_sampling,
        'biaya_sesuai': biaya_sesuai,
        'biaya_tidak_sesuai': biaya_tidak_sesuai,
        'biaya_belum_review': biaya_belum_review,
        'biaya_sampling': biaya_sampling,
        'instance': instance,
    }
    return render(request, 'vpkaak/update_finalisasi_register_post_klaim.html', context)


@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
def kertas_kerja_koreksi(request):
    queryset = RegisterPostKlaim.objects.filter(
        user__kantorcabang=request.user.kantorcabang_set.all().first(),
        status=StatusChoices.Finalisasi)

    # filter
    myFilter = RegisterPostKlaimFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    # pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'register_list': queryset,
        'myFilter': myFilter,
    }
    return render(request, 'vpkaak/kertas_kerja_koreksi.html', context)


@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
def input_nomor_ba(request, pk):
    queryset = RegisterPostKlaim.objects.filter(user__kantorcabang=request.user.kantorcabang_set.all().first())
    instance = queryset.get(pk=pk)

    if request.method == 'POST':
        form = InputNomorBAForm(request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Nomor BA berhasil diupdate.')
            return redirect('/vpkaak/kertas-kerja-verifikasi')
    else:
        form = InputNomorBAForm(instance=instance)

    context = {
        'form': form,
    }
    return render(request, 'vpkaak/input_nomor_ba.html', context)


@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
def api_json_data_sampling_vpkaak(request):
    queryset = SamplingDataKlaimCBG.objects.filter(register__user__kantorcabang=request.user.kantorcabang_set.all().first(), is_from_kp=False, is_final=False).\
        values('register__nomor_register', 'status', 'Nmkclayan')
    data = list(queryset)
    return JsonResponse(data, safe=False)


@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
def monitoring_data_sampling_vpkaak(request):
    context = {}
    return render(request, 'vpkaak/monitoring_data_sampling_vpkaak.html', context=context)

# robot processing automation with selenium
@login_required
@check_device
@permissions(role=['verifikator', 'stafupk', 'supervisor'])
def rpa_cbg_vpkaak(request):
    data = request.POST
    queryset = None
    context = {
        'data_klaim': queryset,
        # 'register_klaim': register_klaim
    }

    # if 'register_klaim' in data_get:
    register_post_klaim = request.GET.get('nomor_register')
    # status = request.GET.get('status')
    # status_sinkron = request.GET.get('status_sinkron')
    kantor_cabang = request.user.kantorcabang_set.all().first().kode_cabang
    # ganti dengan ini kalo deploy
    # queryset = SamplingDataKlaimCBG.objects.filter(Kdkclayan=kantor_cabang,
    #                                                register__nomor_register=register_post_klaim,
    #                                                is_final=True,
    #                                                status=StatusReviewChoices.TidakSesuai).order_by('created_at')
    # untuk sementara
    queryset = SamplingDataKlaimCBG.objects.filter(Kdkclayan=kantor_cabang,
                                                   register__nomor_register=register_post_klaim,
                                                   status=StatusReviewChoices.TidakSesuai).order_by('created_at')
    # print(queryset)

    # filter
    myFilter = CBGKoreksiBoaFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs
    # print(queryset)
    context.update({
        'data_klaim': queryset,
        'register_post_klaim': register_post_klaim,
        'myFilter': myFilter,
    })

    if 'no_sep' in data:
        nosep = data['no_sep']
        keterangan_koreksi_boa = data['keterangan_koreksi_boa']
        status_koreksi_boa = data['status_koreksi_boa']
        try:
            klaim = SamplingDataKlaimCBG.objects.get(Nosjp=nosep)
            klaim.keterangan_koreksi_boa = keterangan_koreksi_boa
            klaim.tgl_koreksi_boa = datetime.datetime.today()
            klaim.status_koreksi_boa = status_koreksi_boa
            klaim.save()
            context.update({
                'klaim': klaim,
            })
            response_data = {
                'no_sep': nosep,
                'keterangan_koreksi_boa': keterangan_koreksi_boa,
                'status_koreksi_boa': status_koreksi_boa,
                'message': 'Berhasil Update',
            }
            return JsonResponse(response_data)
        except SamplingDataKlaimCBG.DoesNotExist:
            response_data = {
                'no_sep': nosep,
                'message': 'Data tidak ditemukan',
            }
            return JsonResponse(response_data, status=404)
    return render(request, 'vpkaak/rpa_cbg_vpkaak.html', context=context)


#################
# supervisorkp ##
#################
@login_required
@check_device
@permissions(role=['supervisorkp'])
def register_post_klaim_supervisorkp(request):
    queryset = RegisterPostKlaim.objects.filter(is_kp=True).order_by('-created_at')

    # filter
    myFilter = RegisterPostKlaimFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    # pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    if request.POST.get('add_register_post_klaim'):
        form = RegisterPostKlaimSupervisorKPForm(data=request.POST)
        if form.is_valid():
            jenis_audit = form.cleaned_data['jenis_audit']

            # validasi field
            if jenis_audit == 'AAK-FKRTL':
                required_fields = ['periode_awal', 'periode_akhir']
            elif jenis_audit == 'VPK-FKRTL':
                required_fields = ['bulan_beban']
            else:
                messages.warning(request, 'Jenis audit tidak valid')
                return redirect(request.headers.get('Referer'))

            for field_name in required_fields:
                if not form.cleaned_data.get(field_name):
                    messages.warning(request, f'Field {field_name} harus diisi')
                    return redirect(request.headers.get('Referer'))

            # deteksi VPK dengan buban ganda
            if jenis_audit == JenisAuditChoices.VPK:
                cek_vpk = RegisterPostKlaim.objects.filter(is_kp=True, bulan_beban=form.cleaned_data['bulan_beban'])
                if cek_vpk:
                    messages.warning(request,
                                     f'VPK dengan Bulan beban '
                                     f'{calendar.month_name[cek_vpk[0].bulan_beban.month]}-{cek_vpk[0].bulan_beban.year} '
                                     f'sudah ada')
                    return redirect(request.headers.get('Referer'))

            # Jika validasi berhasil, bentuk nomor register klaim, simpan data dan atur status menjadi Register
            if jenis_audit == 'AAK-FKRTL':
                jenis_audit_prefix = 'AAK-FKRTL'
            elif jenis_audit == 'VPK-FKRTL':
                jenis_audit_prefix = 'VPK-FKRTL'
            else:
                jenis_audit_prefix = 'OTHER'
            bulan_tahun = timezone.now().strftime("%m%y")
            tahun = timezone.now().strftime("%y")
            max_nomor_urut = RegisterPostKlaim.objects.filter(
                nomor_register__contains=f"{jenis_audit_prefix}/KP",
                nomor_register__endswith=tahun).aggregate(Max('nomor_register'))
            if max_nomor_urut['nomor_register__max']:
                nomor_urut = int(max_nomor_urut['nomor_register__max'].split('/')[0])
                nomor_urut += 1
            else:
                nomor_urut = 1

            register_post_klaim = form.save(commit=False)
            register_post_klaim.is_kp = True
            register_post_klaim.status = 'Register'
            register_post_klaim.user = request.user
            register_post_klaim.nomor_register = f"{nomor_urut:03d}/{jenis_audit_prefix}/KP/{bulan_tahun}"
            register_post_klaim.save()

            messages.success(request, 'Register berhasil dibuat')
            return redirect(request.headers.get('Referer'))
    else:
        form = RegisterPostKlaimSupervisorKPForm()

    context = {
        'form': form,
        'register_list': queryset,
        'myFilter': myFilter,
    }
    return render(request, 'vpkaak/register_supervisorkp.html', context)


@login_required
@check_device
@permissions(role=['supervisorkp'])
def import_sampling_data_klaim_supervisorkp(request):
    storage = TemporaryStorage()
    import_form = ImportSamplingDataKlaimForm()
    if request.method == 'POST' and request.POST.get('action') == 'import':
        import_form = ImportSamplingDataKlaimForm(files=request.FILES, data=request.POST)
        if import_form.is_valid():
            list_data_import = ['Nokapst', 'Tgldtgsjp', 'Tglplgsjp', 'Nosjp', 'Tglpelayanan', 'Kdkrlayan', 'Kdkclayan',
                                'Nmkclayan', 'Kddati2Layan', 'Nmdati2Layan', 'Kdppklayan', 'Nmppklayan', 'Nmtkp',
                                'Kdinacbgs', 'Nminacbgs', 'Kddiagprimer', 'Nmdiagprimer', 'Diagsekunder', 'Procedure',
                                'Klsrawat', 'Nmjnspulang', 'kddokter', 'nmdokter', 'Umur', 'kdsa', 'kdsd',
                                'deskripsisd', 'kdsi', 'deskripsisi', 'kdsp', 'deskripsisp', 'kdsr', 'deskripsisr',
                                'Tarifgroup', 'tarifsa', 'tarifsd', 'tarifsi', 'tarifsp', 'tarifsr', 'Biayaverifikasi',
                                'Kodersmenkes', 'Kelasrsmenkes', 'Jkpst', 'redflag', 'deskripsi_redflag', 'source']
            nomor_register = import_form.cleaned_data.get('register')
            register = RegisterPostKlaim.objects.get(nomor_register=nomor_register)
            file_name = f'{uuid.uuid4()}-{int(round(time.time() * 1000))}.xlsx'
            storage.save(name=file_name, content=import_form.cleaned_data.get('file'))
            get_password = request.POST.get('password')
            if get_password != '':
                unlocked_file = io.BytesIO()
                try:
                    with open(storage.path(name=file_name), "rb") as file:
                        excel_file = msoffcrypto.OfficeFile(file)
                        excel_file.load_key(password=get_password)
                        excel_file.decrypt(unlocked_file)
                    data_frame = pd.read_excel(unlocked_file, usecols=list_data_import)
                except InvalidKeyError:
                    messages.warning(request, 'Password File Excel yang Anda masukkan salah!')
                    return redirect(request.headers.get('Referer'))
                except DecryptionError:
                    messages.warning(request, 'File Excel seharusnya tidak memiliki password!')
                    return redirect(request.headers.get('Referer'))
                except Exception as e:
                    messages.warning(request, f'Terjadi Kesalahan pada saat import File. Keterangan error : {e}')
                    return redirect(request.headers.get('Referer'))
            else:
                try:
                    data_frame = pd.read_excel(storage.path(name=file_name), usecols=list_data_import)
                except XLRDError:
                    messages.warning(request,
                                     f'File yang Anda import memiliki password. Silahkan masukkan password file!')
                    return redirect(request.headers.get('Referer'))
                except Exception as e:
                    messages.warning(request, f'Terjadi kesalahan pada saat import File. Keterangan error : {e}')
                    return redirect(request.headers.get('Referer'))
            data_frame = data_frame.replace(np.nan, None)
            data_frame['register'] = register
            data_frame['Kdkclayan'] = data_frame['Kdkclayan'].astype('str').apply(lambda x: x.zfill(4))
            data_frame['is_from_kp'] = True
            # data_frame['faskes'] = register.faskes
            # data_frame['TGLPULANG'] = pd.to_datetime(data_frame['TGLPULANG'])
            # data_frame['bupel'] = data_frame['TGLPULANG'].dt.to_period('M').dt.to_timestamp()
            # data_frame['bupel'] = pd.to_datetime(data_frame['bupel'])
            register.password_file_excel = get_password
            register.save()

            # ini jadi ga semua retrieval dan penulisan di database tidak semua di kasih transaction.atomic
            # with transaction.atomic():
            try:
                obj_list = []
                for _, row in data_frame.iterrows():
                    data_klaim = SamplingDataKlaimCBG()
                    try:
                        data_klaim = SamplingDataKlaimCBG(**dict(row))
                        data_klaim.full_clean()  # Validate the object
                        obj_list.append(data_klaim)
                        # ini saya hapus mas jadi ini tiap iterasi buat list bulk jadi bikin space di db ke lock for a while kalo kebanyakan
                        # DataKlaimCBG.objects.bulk_create(obj_list)
                    except TypeError as e:
                        messages.warning(request,
                                         f'Terjadi kesalahan pada saat import File. Keterangan error : {e}')
                        return redirect(request.headers.get('Referer'))
                    except Exception as e:
                        messages.warning(request,
                                         f'Kesalahan terjadi pada No SEP : {data_klaim.Nosjp}. Keterangan error : {e}')
                        return redirect(request.headers.get('Referer'))

                # ini data penulisan dipisah mas
                with transaction.atomic():
                    SamplingDataKlaimCBG.objects.bulk_create(obj_list)
                    # ini data retrieval
                    valid_data = SamplingDataKlaimCBG.objects.filter(id__in=[obj.id for obj in obj_list])
                    # invalid_data = SamplingDataKlaimCBG.objects.filter(id__in=[obj.id for obj in obj_list
                    #                                                            if
                    #                                                            obj.NOSEP[
                    #                                                            :8] != register.faskes.kode_ppk or
                    #                                                            obj.bupel.month != register.bulan_pelayanan.month or
                    #                                                            obj.bupel.year != register.bulan_pelayanan.year])
                    df_valid = pd.DataFrame.from_records(valid_data.values())
                    total_data_valid = len(df_valid)
                    df_valid = df_valid.head()
                    # df_invalid = pd.DataFrame.from_records(invalid_data.values())
                    # total_data_invalid = len(df_invalid)
                    transaction.set_rollback(True)

            except IntegrityError as e:
                messages.info(request, f'Terjadi kesalahan saat mencoba menyimpan data : {e}')
                return redirect(request.headers.get('Referer'))
            except Exception as e:
                messages.info(request, f'Kesalahan terjadi pada saat import File. Keterangan error : {e}')
                return redirect(request.headers.get('Referer'))

            return render(request, 'vpkaak/preview_sampling_data_import_cbg.html',
                          {'preview_data_valid': df_valid,
                           'total_data_valid': total_data_valid,
                           # 'preview_data_invalid': df_invalid,
                           # 'total_data_invalid': total_data_invalid,
                           'file_name': file_name,
                           'register': nomor_register,
                           'password': get_password})

    if request.method == 'POST' and request.POST.get('action') == 'confirm':
        list_data_import = ['Nokapst', 'Tgldtgsjp', 'Tglplgsjp', 'Nosjp', 'Tglpelayanan', 'Kdkrlayan', 'Kdkclayan',
                            'Nmkclayan', 'Kddati2Layan', 'Nmdati2Layan', 'Kdppklayan', 'Nmppklayan', 'Nmtkp',
                            'Kdinacbgs', 'Nminacbgs', 'Kddiagprimer', 'Nmdiagprimer', 'Diagsekunder', 'Procedure',
                            'Klsrawat', 'Nmjnspulang', 'kddokter', 'nmdokter', 'Umur', 'kdsa', 'kdsd',
                            'deskripsisd', 'kdsi', 'deskripsisi', 'kdsp', 'deskripsisp', 'kdsr', 'deskripsisr',
                            'Tarifgroup', 'tarifsa', 'tarifsd', 'tarifsi', 'tarifsp', 'tarifsr', 'Biayaverifikasi',
                            'Kodersmenkes', 'Kelasrsmenkes', 'Jkpst', 'redflag', 'deskripsi_redflag', 'source']
        file_name = request.POST.get('file_name')
        nomor_register = request.POST.get('register')
        get_password = request.POST.get('password')
        register = RegisterPostKlaim.objects.get(nomor_register=nomor_register)
        if get_password != '':
            unlocked_file = io.BytesIO()
            try:
                with open(storage.path(name=file_name), "rb") as file:
                    excel_file = msoffcrypto.OfficeFile(file)
                    excel_file.load_key(password=request.POST.get('password'))
                    excel_file.decrypt(unlocked_file)
                data_frame = pd.read_excel(unlocked_file, usecols=list_data_import)
            except InvalidKeyError:
                messages.warning(request, 'Password File Excel yang Anda masukkan salah!')
                return redirect(request.headers.get('Referer'))
            except DecryptionError:
                messages.warning(request, 'File Excel seharusnya tidak memiliki password!')
                return redirect(request.headers.get('Referer'))
            except Exception as e:
                messages.warning(request, f'Terjadi Kesalahan pada saat import File. Keterangan error : {e}')
                return redirect(request.headers.get('Referer'))
        else:
            try:
                data_frame = pd.read_excel(storage.path(name=file_name), usecols=list_data_import)
            except XLRDError:
                messages.warning(request,
                                 f'File yang Anda import memiliki password. Silahkan masukkan password file!')
                return redirect(request.headers.get('Referer'))
            except Exception as e:
                messages.warning(request, f'Terjadi kesalahan pada saat import File. Keterangan error : {e}')
                return redirect(request.headers.get('Referer'))

        data_frame = data_frame.replace(np.nan, None)
        data_frame['register'] = register
        data_frame['Kdkclayan'] = data_frame['Kdkclayan'].astype('str').apply(lambda x: x.zfill(4))
        data_frame['is_from_kp'] = True

        # saya nambahin try and except disini mas jadi ga ngasih 500 tapi ngasih messages error kalo ada kesalahan
        try:
            with transaction.atomic():
                SamplingDataKlaimCBG.objects.bulk_create(
                    [SamplingDataKlaimCBG(**dict(row[1])) for row in data_frame.iterrows()])
            # register jadi verifikasi
            register.status = StatusChoices.Verifikasi
            register.save()
            messages.success(request, "Data Klaim Berhasil Di-import")
        except Exception as e:
            messages.error(request, f"Terjadi kesalahan: {e}")
            return redirect(request.headers.get('Referer'))

    context = {
        'import_form': import_form,
        # 'verifikator': verifikator,
    }
    return render(request, 'vpkaak/import_data_supervisorkp.html', context)


@login_required
@check_device
@permissions(role=['supervisorkp'])
def api_json_data_sampling_vpkaak_supervisorkp(request):
    queryset = SamplingDataKlaimCBG.objects.filter(is_from_kp=True, is_final=False).\
        values('register__nomor_register', 'status', 'Nmkclayan')
    data = list(queryset)
    return JsonResponse(data, safe=False)


@login_required
@check_device
@permissions(role=['supervisorkp'])
def monitoring_data_sampling_vpkaak_supervisorkp(request):
    context = {}
    return render(request, 'vpkaak/monitoring_data_sampling_vpkaak_supervisorkp.html', context=context)


@login_required
@check_device
@permissions(role=['supervisorkp'])
def finalisasi_register_post_klaim_supervisorkp(request):
    queryset = RegisterPostKlaim.objects.filter(is_kp=True, status=StatusChoices.Verifikasi)

    # filter
    myFilter = RegisterPostKlaimFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    # pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'register_list': queryset,
        'myFilter': myFilter,
    }
    return render(request, 'vpkaak/finalisasi_register_post_klaim_supervisorkp.html', context)


@login_required
@check_device
@permissions(role=['supervisorkp'])
def update_finalisasi_register_post_klaim_supervisorkp(request, pk):
    queryset = RegisterPostKlaim.objects.filter(is_kp=True)
    instance = queryset.get(pk=pk)

    try:
        sampling_data_klaim = SamplingDataKlaimCBG.objects.filter(register=instance)
        jumlah_sesuai = sampling_data_klaim.filter(status=StatusReviewChoices.Sesuai).count()
        jumlah_tidak_sesuai = sampling_data_klaim.filter(status=StatusReviewChoices.TidakSesuai).count()
        jumlah_belum_review = sampling_data_klaim.filter(status=StatusReviewChoices.Belum).count()
        total_sampling = sampling_data_klaim.count()

        biaya_sesuai = sampling_data_klaim.filter(status=StatusReviewChoices.Sesuai).aggregate(Sum('Biayaverifikasi'))[
            'Biayaverifikasi__sum']
        biaya_tidak_sesuai = \
            sampling_data_klaim.filter(status=StatusReviewChoices.TidakSesuai).aggregate(Sum('Biayaverifikasi'))[
                'Biayaverifikasi__sum']
        biaya_belum_review = \
            sampling_data_klaim.filter(status=StatusReviewChoices.Belum).aggregate(Sum('Biayaverifikasi'))[
                'Biayaverifikasi__sum']
        biaya_sampling = sampling_data_klaim.aggregate(Sum('Biayaverifikasi'))['Biayaverifikasi__sum']
    except Exception as e:
        messages.warning(request, f'Pengecekan Data Sampling Error, {e}')
        return redirect(request.headers.get('Referer'))

    if request.method == 'POST':
        form = FinalisasiRegisterPostKlaimForm(request.POST, instance=instance)
        if form.is_valid():
            if jumlah_belum_review != 0:
                messages.warning(request, 'Masih terdapat data sampling klaim yang belum direview')
                return redirect(request.headers.get('Referer'))
            obj = form.save()
            obj.is_final = True
            obj.save()
            sampling_data_klaim.update(is_final=True)
            messages.success(request, 'Finalisasi Register berhasil.')
    else:
        form = FinalisasiRegisterPostKlaimForm(instance=instance)

    context = {
        'form': form,
        'jumlah_sesuai': jumlah_sesuai,
        'jumlah_tidak_sesuai': jumlah_tidak_sesuai,
        'jumlah_belum_review': jumlah_belum_review,
        'total_sampling': total_sampling,
        'biaya_sesuai': biaya_sesuai,
        'biaya_tidak_sesuai': biaya_tidak_sesuai,
        'biaya_belum_review': biaya_belum_review,
        'biaya_sampling': biaya_sampling,
        'instance': instance,
    }
    return render(request, 'vpkaak/update_finalisasi_register_post_klaim_supervisorkp.html', context)