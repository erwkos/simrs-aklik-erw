import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm, AdminPasswordChangeForm
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl.workbook import Workbook

from faskes.models import KantorCabang
from klaim.models import RegisterKlaim, DataKlaimCBG
from supervisorkp.filters import UserSupervisorkpFilter
from supervisorkp.forms import EditUserSupervisorkpForm, CreateUserSupervisorkpForm
from user.decorators import check_device, permissions
from user.models import User
from verifikator.filters import DataKlaimCBGFilter


@login_required
@check_device
@permissions(role=['supervisorkp'])
def api_json_register_supervisorkp(request):
    queryset = RegisterKlaim.objects.all().values('faskes__nama', 'faskes__kantor_cabang__nama', 'bulan_pelayanan')
    data = list(queryset)
    return JsonResponse(data, safe=False)


@login_required
@check_device
@permissions(role=['supervisorkp'])
def monitoring_register_supervisorkp(request):
    context = {}
    return render(request, 'supervisorkp/monitoring_register_supervisorkp.html', context=context)


@login_required
@check_device
@permissions(role=['supervisorkp'])
def daftar_user_supervisorkp(request):
    queryset = User.objects.all().order_by('-date_joined')

    # filter
    myFilter = UserSupervisorkpFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    # pagination
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'register_list': queryset,
        'myFilter': myFilter,
    }
    return render(request, 'supervisorkp/daftar_user_supervisorkp.html', context=context)


@login_required
@check_device
@permissions(role=['supervisorkp'])
def update_user_supervisorkp(request, pk):
    instance = get_object_or_404(User, pk=pk)
    form = EditUserSupervisorkpForm(instance=instance)
    if request.method == 'POST':
        form = EditUserSupervisorkpForm(data=request.POST, instance=instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Data user berhasil diubah')
            return redirect('supervisorkp:daftar_user_supervisorkp')
        else:
            form = EditUserSupervisorkpForm()
            messages.warning(request, 'Data user tidak berhasil diubah')
    context = {'form': form,
               'instance': instance
               }
    return render(request, 'supervisorkp/update_user_supervisorkp.html', context=context)


@login_required
@check_device
@permissions(role=['supervisorkp'])
def add_user_supervisorkp(request):
    kc = KantorCabang.objects.all()
    if request.method == 'POST':
        form = CreateUserSupervisorkpForm(data=request.POST)
        if form.is_valid():
            add_kc = kc.filter(nama=request.POST.get('kc')).first()
            if add_kc is not None:
                try:
                    fm = form.save()
                    add_kc.user.add(fm)
                    messages.success(request, 'Data user berhasil ditambahkan')
                    return redirect(request.headers.get('Referer'))
                except Exception as e:
                    messages.warning(request, 'Terdapat error pada saat penambahan user :', str(e))
                    return redirect(request.headers.get('Referer'))
            elif add_kc is None:
                try:
                    form.save()
                    messages.success(request, 'Data user berhasil ditambahkan')
                    return redirect(request.headers.get('Referer'))
                except Exception as e:
                    messages.warning(request, 'Terdapat error pada saat penambahan user :', str(e))
                    return redirect(request.headers.get('Referer'))
        else:
            form = CreateUserSupervisorkpForm()
            messages.warning(request, 'Data user tidak berhasil ditambahkan')
    else:
        form = CreateUserSupervisorkpForm()
    context = {'form': form,
               'kc': kc}
    return render(request, 'supervisorkp/add_user_supervisorkp.html', context=context)


@login_required
@check_device
@permissions(role=['supervisorkp'])
def reset_password_supervisorkp(request, pk):
    user = get_object_or_404(User, pk=pk)
    form = AdminPasswordChangeForm(user)
    if request.method == 'POST':
        form = AdminPasswordChangeForm(user, data=request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Reset password berhasil')
            return redirect('supervisorkp:daftar_user_supervisorkp')
        else:
            form = AdminPasswordChangeForm(user)
            messages.warning(request, 'Reset password tidak berhasil')
    context = {'form': form,
               'instance': user
               }
    return render(request, 'supervisorkp/reset_password_supervisorkp.html', context=context)


@login_required
@check_device
@permissions(role=['supervisorkp'])
def daftar_data_klaim_cbg(request):
    queryset = DataKlaimCBG.objects.filter(status='Klaim',
                                           prosesklaim=True,
                                           prosespending=True)
    # filter
    myFilter = DataKlaimCBGFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    # export excel
    export = request.GET.get('export')
    if export == 'export':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-dataverifikasi.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Data Klaim CBG'

        # Define the titles for columns
        columns = [
            'id',
            'namars',
            'Status',
            'NOSEP',
            'prosesklaim',
            'prosespending',
            'prosesdispute',
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all movies
        for queryset in queryset:
            row_num += 1

            # Define the data for each cell in the row
            row = [
                queryset.id,
                queryset.faskes.nama,
                queryset.status,
                queryset.NOSEP,
                queryset.prosesklaim,
                queryset.prosespending,
                queryset.prosesdispute,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response

    # pagination
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'data_klaim': queryset,
        'myFilter': myFilter,
    }

    return render(request, 'supervisorkp/daftar_data_klaim_cbg.html', context)


