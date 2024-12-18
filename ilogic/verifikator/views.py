import io
import json
import re

import msoffcrypto
from Crypto.Cipher import AES
from Crypto.Random import get_random_bytes
from dal import autocomplete
from django.contrib import messages
from django.contrib.auth.models import Group
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Sum, Count, Max, Q
from django.http import JsonResponse, HttpResponseRedirect
from django.shortcuts import render, redirect, HttpResponse, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.template.loader import render_to_string
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.views.decorators.http import require_POST, require_GET
from msoffcrypto.exceptions import InvalidKeyError, DecryptionError
from openpyxl import Workbook
import pandas as pd
import numpy as np
import uuid
import time
import random
import datetime
from datetime import timedelta

from openpyxl.reader.excel import load_workbook
from tablib import Dataset
from xlrd import XLRDError

from faskes.models import Faskes
from klaim.filters import RegisterKlaimFaskesFilter
from klaim.models import (
    RegisterKlaim,
    DataKlaimCBG, KeteranganPendingDispute, SLA, DataKlaimObat
)
from klaim.resources import DataKlaimCBGResource, DataKlaimObatResource
from metafisik.models import DataKlaimCBGMetafisik
from .filters import DataKlaimCBGFilter, DownloadDataKlaimCBGFilter, DataKlaimObatFilter, DownloadDataKlaimObatFilter, \
    SinkronisasiVIBIVIDIFilter, FragmentasiFilter, ReadmisiFilter
from .forms import (
    StatusRegisterKlaimForm,
    ImportDataKlaimForm,
    DataKlaimVerifikatorForm, FinalisasiVerifikatorForm, HitungDataKlaimForm, KeteranganPendingForm,
    STATUS_CHOICES_DATA_KLAIM_VERIFIKATOR, UploadDataKlaimForm, PotongKlaimForm, DataKlaimForm
)
from user.decorators import permissions, check_device
from user.models import User
from klaim.choices import (
    StatusDataKlaimChoices,
    JenisPelayananChoices, StatusRegisterChoices, NamaJenisKlaimChoices, JenisPendingChoices, JenisDisputeChoices,
    StatusSinkronChoices
)
from .models import HitungDataKlaim
from .storages import TemporaryStorage
from collections import Counter, defaultdict
from .utils import pembagian_tugas
from django.db import IntegrityError


def replace_illegal_characters(text):
    # Mengganti karakter yang tidak diinginkan
    return re.sub(r'[^A-Za-z0-9\s.,-]', '', text)


@login_required
@check_device
@permissions(role=['verifikator'])
def daftar_register(request):
    queryset = RegisterKlaim.objects.filter(
        nomor_register_klaim__startswith=request.user.kantorcabang_set.all().first().kode_cabang).order_by('-tgl_aju')

    # filter
    myFilter = RegisterKlaimFaskesFilter(request.GET, queryset=queryset, request=request)
    myFilter.request = request
    queryset = myFilter.qs

    # pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'register_list': queryset,
        'myFilter': myFilter,
    }

    # (DataKlaimCBG.objects.filter(register_klaim__nomor_register_klaim='020123080001',
    #                                    status=StatusDataKlaimChoices.PROSES).
    #  update(status=StatusDataKlaimChoices.LAYAK))

    return render(request, 'verifikator/daftar_register.html', context)


@login_required
@check_device
@permissions(role=['verifikator'])
def detail_register(request, pk):
    kantor_cabang = request.user.kantorcabang_set.all().first()
    queryset = RegisterKlaim.objects.filter(nomor_register_klaim__startswith=kantor_cabang.kode_cabang)
    instance = queryset.get(id=pk)
    status_form = StatusRegisterKlaimForm(instance=instance)
    potongklaim_form = PotongKlaimForm(instance=instance)

    if request.method == 'POST' and request.POST.get('action') == 'potong_klaim':
        potongklaim_form = PotongKlaimForm(request.POST, instance=instance)
        if potongklaim_form.is_valid():
            potongklaim_form.save()
            messages.success(request, 'Update Flaging Potong Klaim Berhasil')
            return redirect(request.headers.get('Referer'))
        else:
            messages.warning(request, 'Update Flaging Potong Klaim Gagal')
            return redirect(request.headers.get('Referer'))

    if request.method == 'POST' and request.POST.get('action') == 'update_status':
        status_form = StatusRegisterKlaimForm(instance=instance, data=request.POST)
        if instance.verifikator != request.user:
            return HttpResponse(content="Anda Tidak Memiliki Hak Akses, Harap Menghubungi Admin!", status=403)

        if status_form.is_valid():
            status_form.save()

            # jika tgl ba lengkap diubah, maka tgl SLA verifikasi juga berubah
            if 'tgl_ba_lengkap' in status_form.changed_data:
                data_klaim = None
                if instance.jenis_klaim.nama == NamaJenisKlaimChoices.CBG_REGULER or instance.jenis_klaim.nama == NamaJenisKlaimChoices.CBG_SUSULAN:
                    data_klaim = DataKlaimCBG.objects.filter(
                        register_klaim__nomor_register_klaim=instance.nomor_register_klaim)
                elif instance.jenis_klaim.nama == NamaJenisKlaimChoices.OBAT_REGULER or instance.jenis_klaim.nama == NamaJenisKlaimChoices.OBAT_SUSULAN:
                    data_klaim = DataKlaimObat.objects.filter(
                        register_klaim__nomor_register_klaim=instance.nomor_register_klaim)
                if data_klaim:
                    sla = SLA.objects.filter(jenis_klaim=instance.jenis_klaim,
                                             kantor_cabang=request.user.kantorcabang_set.all().first()).first()
                    if sla:
                        if instance.tgl_ba_lengkap:
                            tgl_sla = instance.tgl_ba_lengkap + datetime.timedelta(days=sla.plus_hari_sla)
                            data_klaim.update(tgl_SLA=tgl_sla)
                    else:
                        if instance.tgl_ba_lengkap:
                            tgl_sla = instance.tgl_ba_lengkap + datetime.timedelta(days=6)
                            data_klaim.update(tgl_SLA=tgl_sla)

            if instance.jenis_klaim.nama == NamaJenisKlaimChoices.CBG_SUSULAN:
                data_klaim = DataKlaimCBG.objects.filter(faskes=instance.faskes,
                                                         bupel=instance.bulan_pelayanan,
                                                         status=StatusDataKlaimChoices.PEMBAHASAN)
                for data_klaim in data_klaim:
                    data_klaim.register_klaim = instance
                    data_klaim.tgl_SLA = None
                    data_klaim.status = StatusDataKlaimChoices.PROSES
                    data_klaim.prosesklaim = False
                    data_klaim.is_hitung = False
                    data_klaim.save()

                # status Tidak Layak tidak dapat diubah lagi status nya oleh Faskes
                data_klaim_tidak_layak = DataKlaimCBG.objects.filter(faskes=instance.faskes,
                                                                     bupel=instance.bulan_pelayanan,
                                                                     status=StatusDataKlaimChoices.TIDAK_LAYAK)
                for data_klaim_tidak_layak in data_klaim_tidak_layak:
                    data_klaim_tidak_layak.prosestidaklayak = True
                    data_klaim_tidak_layak.save()

            elif instance.jenis_klaim.nama == NamaJenisKlaimChoices.OBAT_SUSULAN:
                data_klaim = DataKlaimObat.objects.filter(faskes=instance.faskes,
                                                          bupel=instance.bulan_pelayanan,
                                                          status=StatusDataKlaimChoices.PEMBAHASAN)
                for data_klaim in data_klaim:
                    data_klaim.register_klaim = instance
                    data_klaim.tgl_SLA = None
                    data_klaim.status = StatusDataKlaimChoices.PROSES
                    data_klaim.prosesklaim = False
                    data_klaim.is_hitung = False
                    data_klaim.save()

                # status Tidak Layak tidak dapat diubah lagi status nya oleh Faskes
                data_klaim_tidak_layak = DataKlaimCBG.objects.filter(faskes=instance.faskes,
                                                                     bupel=instance.bulan_pelayanan,
                                                                     status=StatusDataKlaimChoices.TIDAK_LAYAK)
                for data_klaim_tidak_layak in data_klaim_tidak_layak:
                    data_klaim_tidak_layak.prosestidaklayak = True
                    data_klaim_tidak_layak.save()
            messages.success(request, "Data Berhasil Disimpan. Selanjutnya lakukan import data klaim. Terima Kasih")
            return redirect(request.headers.get('Referer'))

    context = {
        'register': instance,
        'status_form': status_form,
        'potongklaim_form': potongklaim_form,
    }
    return render(request, 'verifikator/detail_register.html', context)


@login_required
@check_device
@permissions(role=['verifikator'])
def import_data_klaim(request):
    storage = TemporaryStorage()
    import_form = ImportDataKlaimForm()
    dataklaim_form = DataKlaimForm()
    verifikator = list(User.objects.filter(
        kantorcabang__in=request.user.kantorcabang_set.all(),
        groups__name='verifikator',
        is_active=True,
        is_staff=True
    ))

    if request.method == 'POST' and request.POST.get('action') == 'import':
        import_form = ImportDataKlaimForm(files=request.FILES, data=request.POST)
        if import_form.is_valid():
            list_data_import = ['NOSEP', 'TGLSEP', 'TGLPULANG', 'JNSPEL', 'NOKARTU', 'NMPESERTA',
                                'POLI', 'KDINACBG', 'BYPENGAJUAN', 'ALGORITMA']
            nomor_register_klaim = import_form.cleaned_data.get('register')
            register = RegisterKlaim.objects.get(nomor_register_klaim=nomor_register_klaim)
            file_name = f'{uuid.uuid4()}-{int(round(time.time() * 1000))}.xlsx'
            storage.save(name=file_name, content=import_form.cleaned_data.get('file'))
            get_password = request.POST.get('password')

            # Load the Excel file
            try:
                if get_password:
                    unlocked_file = io.BytesIO()
                    with open(storage.path(name=file_name), "rb") as file:
                        excel_file = msoffcrypto.OfficeFile(file)
                        excel_file.load_key(password=get_password)
                        excel_file.decrypt(unlocked_file)
                    data_frame = pd.read_excel(unlocked_file, usecols=list_data_import)
                else:
                    data_frame = pd.read_excel(storage.path(name=file_name), usecols=list_data_import)
            except Exception as e:
                messages.warning(request, f'Terjadi kesalahan saat memuat file: {e}')
                return redirect('/verifikator/import-data-klaim')

            data_frame = data_frame.replace(np.nan, None)
            data_frame['register_klaim'] = register
            data_frame['faskes'] = register.faskes
            data_frame['TGLPULANG'] = pd.to_datetime(data_frame['TGLPULANG'])
            data_frame['bupel'] = pd.to_datetime(data_frame['TGLPULANG'].dt.to_period('M').dt.to_timestamp())
            register.password_file_excel = get_password
            register.save()

            # Data validation
            # Check for missing NOKARTU and NMPESERTA
            missing_nokartu = data_frame['NOKARTU'].isnull() | data_frame['NOKARTU'].eq('')
            missing_nmpeserta = data_frame['NMPESERTA'].isnull() | data_frame['NMPESERTA'].eq('')

            if missing_nokartu.any() or missing_nmpeserta.any():
                error_rows = data_frame[missing_nokartu | missing_nmpeserta]
                error_nosep = error_rows['NOSEP'].tolist()
                messages.error(request,
                               f'Terdapat data dengan NOKARTU atau NMPESERTA kosong untuk NOSEP: {", ".join(map(str, error_nosep))}')
                return redirect('/verifikator/import-data-klaim')

            # Check for duplicate NOSEP in the data_frame
            duplicate_nosep = data_frame['NOSEP'][data_frame['NOSEP'].duplicated()].tolist()
            if duplicate_nosep:
                messages.error(request,
                               f'Terdapat duplikasi NOSEP dalam file yang diimpor: {", ".join(map(str, duplicate_nosep))}')
                return redirect('/verifikator/import-data-klaim')

            # Check for NOSEP that already exist in the database
            existing_nosep = DataKlaimCBG.objects.filter(NOSEP__in=data_frame['NOSEP']).values_list('NOSEP', flat=True)
            if existing_nosep.exists():
                messages.error(request,
                               f'Terdapat NOSEP yang sudah pernah diimpor sebelumnya: {", ".join(existing_nosep)}')
                return redirect('/verifikator/import-data-klaim')

            # Data validation
            valid_data = data_frame[
                (data_frame['NOSEP'].str[:8] == register.faskes.kode_ppk) &
                (data_frame['bupel'].dt.month == register.bulan_pelayanan.month) &
                (data_frame['bupel'].dt.year == register.bulan_pelayanan.year)
            ]
            total_data_valid = len(valid_data)
            invalid_data = data_frame.drop(valid_data.index)
            total_data_invalid = len(invalid_data)

            return render(request, 'verifikator/cbg/preview_data_import_cbg.html', {
                'total_data_valid': total_data_valid,
                'preview_data_invalid': invalid_data,
                'total_data_invalid': total_data_invalid,
                'file_name': file_name,
                'register': nomor_register_klaim,
                'password': get_password
            })

    if request.method == 'POST' and request.POST.get('action') == 'confirm':
        list_data_import = ['NOSEP', 'TGLSEP', 'TGLPULANG', 'JNSPEL', 'NOKARTU', 'NMPESERTA',
                            'POLI', 'KDINACBG', 'BYPENGAJUAN', 'ALGORITMA']
        file_name = request.POST.get('file_name')
        nomor_register_klaim = request.POST.get('register')
        get_password = request.POST.get('password')
        register = RegisterKlaim.objects.get(nomor_register_klaim=nomor_register_klaim)

        # Load the Excel file
        try:
            if get_password:
                unlocked_file = io.BytesIO()
                with open(storage.path(name=file_name), "rb") as file:
                    excel_file = msoffcrypto.OfficeFile(file)
                    excel_file.load_key(password=get_password)
                    excel_file.decrypt(unlocked_file)
                data_frame = pd.read_excel(unlocked_file, usecols=list_data_import)
            else:
                data_frame = pd.read_excel(storage.path(name=file_name), usecols=list_data_import)
        except Exception as e:
            messages.warning(request, f'Terjadi kesalahan saat memuat file: {e}')
            return redirect('/verifikator/import-data-klaim')

        data_frame = data_frame.replace(np.nan, None)
        data_frame['register_klaim'] = register
        data_frame['faskes'] = register.faskes
        data_frame['TGLPULANG'] = pd.to_datetime(data_frame['TGLPULANG'])
        data_frame['bupel'] = pd.to_datetime(data_frame['TGLPULANG'].dt.to_period('M').dt.to_timestamp())

        # Filter valid data
        valid_data = data_frame[
            (data_frame['NOSEP'].str[:8] == register.faskes.kode_ppk) &
            (data_frame['bupel'].dt.month == register.bulan_pelayanan.month) &
            (data_frame['bupel'].dt.year == register.bulan_pelayanan.year)
        ]

        try:
            with transaction.atomic():
                register.has_import_data = True
                register.save()

            # Update JNSPEL based on KDINACBG for valid data
            valid_data['JNSPEL'] = valid_data.apply(
                lambda row: JenisPelayananChoices.RAWAT_INAP.value if str(row['KDINACBG']).endswith('I') else (
                    JenisPelayananChoices.RAWAT_JALAN.value if str(row['KDINACBG']).endswith('0') else row['JNSPEL']
                ), axis=1
            )

            # Create DataKlaimCBG instances
            data_records = valid_data.to_dict('records')
            data_klaim_objects = [DataKlaimCBG(**record) for record in data_records]
            DataKlaimCBG.objects.bulk_create(data_klaim_objects)

            # Assign verifikators
            queryset = DataKlaimCBG.objects.filter(register_klaim=register, status=StatusDataKlaimChoices.BELUM_VER)
            verifikator_ids = [v.id for v in verifikator]
            num_verifikators = len(verifikator_ids)

            # Function to assign verifikators
            def assign_verifikator(queryset, jenis_pelayanan, reverse_order=False):
                qs_pelayanan = queryset.filter(JNSPEL=jenis_pelayanan.value)
                if not qs_pelayanan.exists():
                    return

                participants = qs_pelayanan.values_list('NMPESERTA', 'NOKARTU').distinct()
                participant_list = list(participants)
                random.shuffle(participant_list)

                verifikator_order = verifikator_ids[::-1] if reverse_order else verifikator_ids

                total_participants = len(participant_list)
                base_chunk_size = total_participants // num_verifikators
                remainder = total_participants % num_verifikators

                chunks = []
                start = 0
                for i in range(num_verifikators):
                    end = start + base_chunk_size + (1 if i < remainder else 0)
                    chunks.append(participant_list[start:end])
                    start = end

                verifikator_participant_map = {
                    verifikator_order[i]: chunks[i] for i in range(len(chunks))
                }

                with transaction.atomic():
                    for verifikator_id, participant_chunk in verifikator_participant_map.items():
                        if not participant_chunk:
                            continue
                        nmpeserta_list = [p[0] for p in participant_chunk]
                        nokartu_list = [p[1] for p in participant_chunk]
                        qs_pelayanan.filter(
                            NMPESERTA__in=nmpeserta_list,
                            NOKARTU__in=nokartu_list
                        ).update(
                            verifikator_id=verifikator_id,
                            status=StatusDataKlaimChoices.PROSES
                        )

            # Assign for RAWAT_JALAN and RAWAT_INAP
            assign_verifikator(queryset, JenisPelayananChoices.RAWAT_JALAN, reverse_order=False)
            assign_verifikator(queryset, JenisPelayananChoices.RAWAT_INAP, reverse_order=True)

            # Update tgl_SLA
            queryset_proses = DataKlaimCBG.objects.filter(
                register_klaim=register, status=StatusDataKlaimChoices.PROSES
            )

            sla = SLA.objects.filter(
                jenis_klaim=register.jenis_klaim,
                kantor_cabang=register.faskes.kantor_cabang
            ).first()
            sla_days = sla.plus_hari_sla if sla else 6

            if register.tgl_ba_lengkap:
                tgl_sla = register.tgl_ba_lengkap + timedelta(days=sla_days)
            elif register.tgl_terima:
                tgl_sla = register.tgl_terima + timedelta(days=15)
            else:
                tgl_sla = None

            if tgl_sla:
                queryset_proses.update(tgl_SLA=tgl_sla)

            # Save the file to the register
            with transaction.atomic():
                register.file_data_klaim = storage.open(name=file_name)
                register.file_data_klaim.name = file_name
                register.save()

            messages.success(request, "Data CBG Berhasil Di-import")

        except Exception as e:
            messages.error(request, f"Terjadi kesalahan: {e}")
            return redirect('/verifikator/import-data-klaim')

    context = {
        'import_form': import_form,
        'dataklaim_form': dataklaim_form,
        'verifikator': verifikator,
    }
    return render(request, 'verifikator/cbg/import_data_klaim_cbg.html', context)

# def import_data_klaim(request):
#     storage = TemporaryStorage()
#     import_form = ImportDataKlaimForm()
#     dataklaim_form = DataKlaimForm()
#     verifikator = User.objects.filter(kantorcabang__in=request.user.kantorcabang_set.all(),
#                                       groups__name='verifikator',
#                                       is_active=True,
#                                       is_staff=True)
#     if request.method == 'POST' and request.POST.get('action') == 'import':
#         import_form = ImportDataKlaimForm(files=request.FILES, data=request.POST)
#         if import_form.is_valid():
#             list_data_import = ['NOSEP', 'TGLSEP', 'TGLPULANG', 'JNSPEL', 'NOKARTU', 'NMPESERTA',
#                                 'POLI', 'KDINACBG', 'BYPENGAJUAN', 'ALGORITMA']
#             nomor_register_klaim = import_form.cleaned_data.get('register')
#             register = RegisterKlaim.objects.get(nomor_register_klaim=nomor_register_klaim)
#             file_name = f'{uuid.uuid4()}-{int(round(time.time() * 1000))}.xlsx'
#             storage.save(name=file_name, content=import_form.cleaned_data.get('file'))
#             get_password = request.POST.get('password')
#             if get_password != '':
#                 unlocked_file = io.BytesIO()
#                 try:
#                     with open(storage.path(name=file_name), "rb") as file:
#                         excel_file = msoffcrypto.OfficeFile(file)
#                         excel_file.load_key(password=get_password)
#                         excel_file.decrypt(unlocked_file)
#                     data_frame = pd.read_excel(unlocked_file, usecols=list_data_import)
#                 except InvalidKeyError:
#                     messages.warning(request, 'Password File Excel yang Anda masukkan salah!')
#                     return redirect('/verifikator/import-data-klaim')
#                 except DecryptionError:
#                     messages.warning(request, 'File Excel seharusnya tidak memiliki password!')
#                     return redirect('/verifikator/import-data-klaim')
#                 except Exception as e:
#                     messages.warning(request, f'Terjadi Kesalahan pada saat import File. Keterangan error : {e}')
#                     return redirect('/verifikator/import-data-klaim')
#             else:
#                 try:
#                     data_frame = pd.read_excel(storage.path(name=file_name), usecols=list_data_import)
#                 except XLRDError:
#                     messages.warning(request,
#                                      f'File yang Anda import memiliki password. Silahkan masukkan password file!')
#                     return redirect('/verifikator/import-data-klaim')
#                 except Exception as e:
#                     messages.warning(request, f'Terjadi kesalahan pada saat import File. Keterangan error : {e}')
#                     return redirect('/verifikator/import-data-klaim')
#                 # data_frame = pd.read_excel(storage.path(name=file_name), usecols=list_data_import, engine='openpyxl', password='Qwerty1!')
#             data_frame = data_frame.replace(np.nan, None)
#             data_frame['register_klaim'] = register
#             data_frame['faskes'] = register.faskes
#             data_frame['TGLPULANG'] = pd.to_datetime(data_frame['TGLPULANG'])
#             data_frame['bupel'] = data_frame['TGLPULANG'].dt.to_period('M').dt.to_timestamp()
#             data_frame['bupel'] = pd.to_datetime(data_frame['bupel'])
#             register.password_file_excel = get_password
#             register.save()
#
#             # ini jadi ga semua retrieval dan penulisan di database tidak semua di kasih transaction.atomic
#             # with transaction.atomic():
#             try:
#                 obj_list = []
#                 for _, row in data_frame.iterrows():
#                     data_klaim = DataKlaimCBG()
#                     try:
#                         data_klaim = DataKlaimCBG(**dict(row))
#                         data_klaim.full_clean()  # Validate the object
#                         obj_list.append(data_klaim)
#                         # ini saya hapus mas jadi ini tiap iterasi buat list bulk jadi bikin space di db ke lock for a while kalo kebanyakan
#                         # DataKlaimCBG.objects.bulk_create(obj_list)
#                     except TypeError as e:
#                         messages.warning(request, f'Terjadi kesalahan pada saat import File. Keterangan error : {e}')
#                         return redirect('/verifikator/import-data-klaim')
#                     except Exception as e:
#                         messages.warning(request,
#                                          f'Kesalahan terjadi pada No SEP : {data_klaim.NOSEP}. Keterangan error : {e}')
#                         return redirect('/verifikator/import-data-klaim')
#
#                 # ini data penulisan dipisah mas
#                 with transaction.atomic():
#                     DataKlaimCBG.objects.bulk_create(obj_list)
#                     # ini data retrieval
#                     valid_data = DataKlaimCBG.objects.filter(id__in=[obj.id for obj in obj_list
#                                                                      if obj.NOSEP[:8] == register.faskes.kode_ppk and
#                                                                      obj.bupel.month == register.bulan_pelayanan.month and
#                                                                      obj.bupel.year == register.bulan_pelayanan.year])
#                     invalid_data = DataKlaimCBG.objects.filter(id__in=[obj.id for obj in obj_list
#                                                                        if obj.NOSEP[:8] != register.faskes.kode_ppk or
#                                                                        obj.bupel.month != register.bulan_pelayanan.month or
#                                                                        obj.bupel.year != register.bulan_pelayanan.year])
#                     df_valid = pd.DataFrame.from_records(valid_data.values())
#                     total_data_valid = len(df_valid)
#                     df_invalid = pd.DataFrame.from_records(invalid_data.values())
#                     total_data_invalid = len(df_invalid)
#                     transaction.set_rollback(True)
#
#             except IntegrityError as e:
#                 messages.info(request, f'Terjadi kesalahan saat mencoba menyimpan data : {e}')
#                 return redirect('/verifikator/import-data-klaim')
#             except Exception as e:
#                 messages.info(request, f'Kesalahan terjadi pada saat import File. Keterangan error : {e}')
#                 return redirect('/verifikator/import-data-klaim')
#
#             return render(request, 'verifikator/cbg/preview_data_import_cbg.html',
#                           {'preview_data_valid': df_valid,
#                            'total_data_valid': total_data_valid,
#                            'preview_data_invalid': df_invalid,
#                            'total_data_invalid': total_data_invalid,
#                            'file_name': file_name,
#                            'register': nomor_register_klaim,
#                            'password': get_password})
#
#     if request.method == 'POST' and request.POST.get('action') == 'confirm':
#         list_data_import = ['NOSEP', 'TGLSEP', 'TGLPULANG', 'JNSPEL', 'NOKARTU', 'NMPESERTA',
#                             'POLI', 'KDINACBG', 'BYPENGAJUAN', 'ALGORITMA']
#         file_name = request.POST.get('file_name')
#         nomor_register_klaim = request.POST.get('register')
#         get_password = request.POST.get('password')
#         register = RegisterKlaim.objects.get(nomor_register_klaim=nomor_register_klaim)
#         if get_password != '':
#             unlocked_file = io.BytesIO()
#             try:
#                 with open(storage.path(name=file_name), "rb") as file:
#                     excel_file = msoffcrypto.OfficeFile(file)
#                     excel_file.load_key(password=request.POST.get('password'))
#                     excel_file.decrypt(unlocked_file)
#                 data_frame = pd.read_excel(unlocked_file, usecols=list_data_import)
#             except InvalidKeyError:
#                 messages.warning(request, 'Password File Excel yang Anda masukkan salah!')
#                 return redirect('/verifikator/import-data-klaim')
#             except DecryptionError:
#                 messages.warning(request, 'File Excel seharusnya tidak memiliki password!')
#                 return redirect('/verifikator/import-data-klaim')
#             except Exception as e:
#                 messages.warning(request, f'Terjadi Kesalahan pada saat import File. Keterangan error : {e}')
#                 return redirect('/verifikator/import-data-klaim')
#         else:
#             try:
#                 data_frame = pd.read_excel(storage.path(name=file_name), usecols=list_data_import)
#             except XLRDError:
#                 messages.warning(request, f'File yang Anda import memiliki password. Silahkan masukkan password file!')
#                 return redirect('/verifikator/import-data-klaim')
#             except Exception as e:
#                 messages.warning(request, f'Terjadi kesalahan pada saat import File. Keterangan error : {e}')
#                 return redirect('/verifikator/import-data-klaim')
#
#         # data_frame = pd.read_excel(storage.path(name=file_name), usecols=list_data_import)
#         data_frame = data_frame.replace(np.nan, None)
#         data_frame['register_klaim'] = register
#         data_frame['faskes'] = register.faskes
#         data_frame['TGLPULANG'] = pd.to_datetime(data_frame['TGLPULANG'])
#         data_frame['bupel'] = data_frame['TGLPULANG'].dt.to_period('M').dt.to_timestamp()
#         data_frame['bupel'] = pd.to_datetime(data_frame['bupel'])
#         # saya nambahin try and except disini mas jadi ga ngasih 500 tapi ngasih messages error kalo ada kesalahan
#         try:
#             with transaction.atomic():
#                 register.has_import_data = True
#                 register.save()
#             with transaction.atomic():
#                 obj_list = DataKlaimCBG.objects.bulk_create(
#                     [DataKlaimCBG(**dict(row[1])) for row in data_frame.iterrows()]
#                 )
#
#             queryset = DataKlaimCBG.objects.filter(register_klaim=register, status=StatusDataKlaimChoices.BELUM_VER)
#
#             NMPESERTA_RJ = [obj.NMPESERTA for obj in queryset.filter(JNSPEL=JenisPelayananChoices.RAWAT_JALAN)]
#             list_nmpeserta_sort_freq_rawat_jalan = [item for items, c in Counter(NMPESERTA_RJ).most_common() for
#                                                     item in
#                                                     [items] * c]
#             list_nmpeserta_no_duplicate_rawat_jalan = list(dict.fromkeys(list_nmpeserta_sort_freq_rawat_jalan))
#             index = random.randrange(len(verifikator))
#
#             with transaction.atomic():
#                 for i in range(len(list_nmpeserta_no_duplicate_rawat_jalan)):
#                     queryset.filter(NMPESERTA=list_nmpeserta_no_duplicate_rawat_jalan[i],
#                                     JNSPEL=JenisPelayananChoices.RAWAT_JALAN). \
#                         update(verifikator=verifikator[index], status=StatusDataKlaimChoices.PROSES)
#                     if index == len(verifikator) - 1:
#                         index = 0
#                     else:
#                         index += 1
#
#             NMPESERTA_RI = [obj.NMPESERTA for obj in queryset.filter(JNSPEL=JenisPelayananChoices.RAWAT_INAP)]
#             list_nmpeserta_sort_freq_rawat_inap = [item for items, c in Counter(NMPESERTA_RI).most_common() for item
#                                                    in
#                                                    [items] * c]
#             list_nmpeserta_no_duplicate_rawat_inap = list(dict.fromkeys(list_nmpeserta_sort_freq_rawat_inap))
#             index = random.randrange(len(verifikator))
#
#             with transaction.atomic():
#                 for i in range(len(list_nmpeserta_no_duplicate_rawat_inap)):
#                     queryset.filter(NMPESERTA=list_nmpeserta_no_duplicate_rawat_inap[i],
#                                     JNSPEL=JenisPelayananChoices.RAWAT_INAP). \
#                         update(verifikator=verifikator[index], status=StatusDataKlaimChoices.PROSES)
#                     if index == len(verifikator) - 1:
#                         index = 0
#                     else:
#                         index += 1
#
#             queryset_proses = DataKlaimCBG.objects.filter(register_klaim=register, status=StatusDataKlaimChoices.PROSES)
#             # for q in q:
#             # q.save()
#
#             # penentuan SLA
#             sla = SLA.objects.filter(jenis_klaim=register.jenis_klaim,
#                                      kantor_cabang=register.faskes.kantor_cabang).first()
#             if sla:
#                 if register.tgl_ba_lengkap:
#                     queryset_proses.update(tgl_SLA=register.tgl_ba_lengkap + timedelta(days=sla.plus_hari_sla))
#                 elif register.tgl_terima:
#                     queryset_proses.update(tgl_SLA=register.tgl_terima + timedelta(days=15))
#             else:
#                 if register.tgl_ba_lengkap:
#                     queryset_proses.update(tgl_SLA=register.tgl_ba_lengkap + timedelta(days=6))
#                 elif register.tgl_terima:
#                     queryset_proses.update(tgl_SLA=register.tgl_terima + timedelta(days=15))
#             with transaction.atomic():
#                 register.file_data_klaim = storage.open(name=file_name)
#                 register.file_data_klaim.name = file_name
#                 register.save()
#             messages.success(request, "Data CBG Berhasil Di-import")
#         except Exception as e:
#             messages.error(request, f"Terjadi kesalahan: {e}")
#             return redirect('/verifikator/import-data-klaim')
#
#     context = {
#         'import_form': import_form,
#         'dataklaim_form': dataklaim_form,
#         'verifikator': verifikator,
#     }
#     return render(request, 'verifikator/cbg/import_data_klaim_cbg.html', context)


@login_required
@check_device
@permissions(role=['verifikator'])
def daftar_data_klaim(request):
    queryset = DataKlaimCBG.objects.filter(verifikator=request.user, prosesklaim=False).order_by('NMPESERTA', 'TGLSEP')

    # Apply filter
    myFilter = DataKlaimCBGFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    # Optimize queryset with select_related and prefetch_related
    queryset = queryset.select_related('faskes', 'register_klaim', 'verifikator').prefetch_related(
        'ket_pending_dispute', 'ket_jawaban_pending'
    )

    export = request.GET.get('export')
    if export == 'export':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-dataverifikasi.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Data Klaim CBG'

        # Define the titles for columns
        columns = [
            'namars',
            'status',
            'NOREG',
            'NOSEP',
            'TGLSEP',
            'TGLPULANG',
            'JNSPEL',
            'NOKARTU',
            'NMPESERTA',
            'POLI',
            'KDINACBG',
            'BYPENGAJUAN',
            'verifikator',
            'ALGORITMA',
            'jenis_pending',
            'jenis_dispute',
            'ket_pending',
            'ket_jawaban',
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all
        for item in queryset:
            row_num += 1

            # Fetch the last ket_pending_dispute and ket_jawaban_pending
            ket_pending_dispute_last = item.ket_pending_dispute.last()
            if ket_pending_dispute_last is None:
                ket_pending_disput_queryset = ''
            else:
                ket_pending_disput_queryset = '{0}, '.format(ket_pending_dispute_last)
                ket_pending_disput_queryset = replace_illegal_characters(ket_pending_disput_queryset)

            ket_jawaban_pending_last = item.ket_jawaban_pending.last()
            if ket_jawaban_pending_last is None:
                ket_jawaban_pending_queryset = ''
            else:
                ket_jawaban_pending_queryset = '{0}, '.format(ket_jawaban_pending_last)
                ket_jawaban_pending_queryset = replace_illegal_characters(ket_jawaban_pending_queryset)

            # Define the data for each cell in the row
            row = [
                item.faskes.nama,
                item.status,
                item.register_klaim.nomor_register_klaim,
                item.NOSEP,
                item.TGLSEP,
                item.TGLPULANG,
                item.JNSPEL,
                item.NOKARTU,
                item.NMPESERTA,
                item.POLI,
                item.KDINACBG,
                item.BYPENGAJUAN,
                item.verifikator.username,
                item.ALGORITMA,
                item.jenis_pending,
                item.jenis_dispute,
                ket_pending_disput_queryset,
                ket_jawaban_pending_queryset,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response

    if request.POST.get('import'):
        try:
            file = request.FILES['excel']
            if file.name.split('.')[-1] != 'xlsx':
                raise ValidationError('Hanya menerima file dengan ekstensi `.xlsx`')
        except Exception as e:
            messages.warning(request, f'Terdapat Error dalam proses upload dengan keterangan: {e}')
            return redirect(request.headers.get('Referer'))

        df_raw = pd.read_excel(file)

        # Replace NaN with empty strings
        df = df_raw.replace(np.nan, '')

        # Convert columns to title case
        df['status'] = df['status'].str.title()
        df['jenis_pending'] = df['jenis_pending'].str.title()
        df['jenis_dispute'] = df['jenis_dispute'].str.title()

        # Call the Resource Model
        data_claim_resource = DataKlaimCBGResource()

        # Load the pandas dataframe into a tablib dataset
        dataset = Dataset().load(df)

        # Check mandatory columns
        list_kolom_mandatory = ['status', 'NOSEP', 'jenis_pending', 'jenis_dispute', 'ket_pending']
        missing_columns = set(list_kolom_mandatory) - set(df.columns)
        if missing_columns:
            messages.warning(request, f'File yang diimport harus memiliki kolom {list_kolom_mandatory}')
            return redirect(request.headers.get('Referer'))

        # Check statuses
        valid_statuses = [StatusDataKlaimChoices.LAYAK, StatusDataKlaimChoices.PENDING,
                          StatusDataKlaimChoices.TIDAK_LAYAK, StatusDataKlaimChoices.DISPUTE]
        invalid_statuses = set(dataset['status']) - set(valid_statuses)
        if invalid_statuses:
            messages.warning(request, f'Status tidak valid: {invalid_statuses}')
            return redirect(request.headers.get('Referer'))

        # Check jenis_pending and jenis_dispute values
        valid_jenis_pending = [JenisPendingChoices.ADMINISTRASI, JenisPendingChoices.KODING,
                               JenisPendingChoices.STANDAR_PELAYANAN]
        invalid_jenis_pending = set(df['jenis_pending'].unique()) - set(valid_jenis_pending + [''])
        invalid_jenis_pending.discard('')
        if invalid_jenis_pending:
            messages.warning(request, f'Jenis pending tidak valid: {invalid_jenis_pending}')
            return redirect(request.headers.get('Referer'))

        valid_jenis_dispute = [JenisDisputeChoices.MEDIS, JenisDisputeChoices.KODING, JenisDisputeChoices.COB]
        invalid_jenis_dispute = set(df['jenis_dispute'].unique()) - set(valid_jenis_dispute + [''])
        invalid_jenis_dispute.discard('')
        if invalid_jenis_dispute:
            messages.warning(request, f'Jenis dispute tidak valid: {invalid_jenis_dispute}')
            return redirect(request.headers.get('Referer'))

        # Validations for Pending and Dispute statuses
        # When status is 'Pending', 'jenis_pending' must be filled
        pending_rows = df[df['status'] == StatusDataKlaimChoices.PENDING]
        if pending_rows['jenis_pending'].eq('').any():
            messages.warning(request, 'Untuk status "Pending", kolom "jenis_pending" harus diisi.')
            return redirect(request.headers.get('Referer'))

        # When status is 'Dispute', both 'jenis_pending' and 'jenis_dispute' must be filled
        dispute_rows = df[df['status'] == StatusDataKlaimChoices.DISPUTE]
        if dispute_rows['jenis_pending'].eq('').any():
            messages.warning(request, 'Untuk status "Dispute", kolom "jenis_pending" harus diisi.')
            return redirect(request.headers.get('Referer'))
        if dispute_rows['jenis_dispute'].eq('').any():
            messages.warning(request, 'Untuk status "Dispute", kolom "jenis_dispute" harus diisi.')
            return redirect(request.headers.get('Referer'))

        # Check ket_pending for Pending, Dispute, and Tidak Layak statuses
        statuses_requiring_ket_pending = [StatusDataKlaimChoices.PENDING, StatusDataKlaimChoices.DISPUTE,
                                          StatusDataKlaimChoices.TIDAK_LAYAK]
        ket_pending_required_rows = df[df['status'].isin(statuses_requiring_ket_pending)]
        if ket_pending_required_rows['ket_pending'].eq('').any():
            messages.warning(request, 'Keterangan pending harus diisi untuk status Pending, Dispute, dan Tidak Layak.')
            return redirect(request.headers.get('Referer'))

        # Get list of NOSEP
        NOSEP_list = df['NOSEP'].unique().tolist()

        # Query DataKlaimCBG objects
        data_klaim_qs = DataKlaimCBG.objects.filter(verifikator=request.user, prosesklaim=False, NOSEP__in=NOSEP_list)
        data_klaim_dict = {dk.NOSEP: dk for dk in data_klaim_qs}

        # Check for missing NOSEP
        missing_NOSEP = set(NOSEP_list) - set(data_klaim_dict.keys())
        if missing_NOSEP:
            messages.warning(request,
                             f'Terdapat NO SEP {missing_NOSEP} yang tidak terdapat dalam verifikasi klaim verifikator bersangkutan.')
            return redirect(request.headers.get('Referer'))

        # Check for DataKlaimCBG objects with status != 'Proses'
        not_proses_NOSEP = [dk.NOSEP for dk in data_klaim_qs if dk.status != StatusDataKlaimChoices.PROSES]
        if not_proses_NOSEP:
            messages.warning(request,
                             f'Terdapat NO SEP {not_proses_NOSEP} yang sudah diverifikasi. Status yang diimport tidak boleh dalam keadaan telah diverifikasi.')
            return redirect(request.headers.get('Referer'))

        try:
            with transaction.atomic():
                # Create KeteranganPendingDispute objects
                kpd_objects = []
                kpd_data = []  # List of tuples (NOSEP, KeteranganPendingDispute object)
                for index, row in df.iterrows():
                    ket_pending = row['ket_pending']
                    NOSEP = row['NOSEP']

                    # Create KeteranganPendingDispute object
                    kpd = KeteranganPendingDispute(ket_pending_dispute=ket_pending, verifikator=request.user)
                    kpd_objects.append(kpd)
                    kpd_data.append((NOSEP, kpd))

                # Bulk create KeteranganPendingDispute
                KeteranganPendingDispute.objects.bulk_create(kpd_objects)

                # Associate KeteranganPendingDispute with DataKlaimCBG
                for NOSEP, kpd in kpd_data:
                    dk = data_klaim_dict[NOSEP]
                    dk.ket_pending_dispute.add(kpd)

                # Import data using resource
                data_claim_resource.import_data(dataset, dry_run=False)

                # Delete empty ket_pending_dispute
                KeteranganPendingDispute.objects.filter(ket_pending_dispute='', verifikator=request.user).delete()

                messages.success(request, 'Data berhasil diimport.')
                return redirect(request.headers.get('Referer'))
        except Exception as e:
            messages.info(request, f'Terjadi kesalahan: {str(e)}')

    # Pagination
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'data_klaim': queryset,
        'myFilter': myFilter,
    }

    return render(request, 'verifikator/cbg/daftar_data_klaim_cbg.html', context)

# def daftar_data_klaim(request):
#     queryset = DataKlaimCBG.objects.filter(verifikator=request.user, prosesklaim=False).order_by('NMPESERTA', 'TGLSEP')
# 
#     # filter
#     myFilter = DataKlaimCBGFilter(request.GET, queryset=queryset)
#     queryset = myFilter.qs
#     export = request.GET.get('export')
#     if export == 'export':
#         response = HttpResponse(
#             content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
#         )
#         response['Content-Disposition'] = 'attachment; filename={date}-dataverifikasi.xlsx'.format(
#             date=datetime.datetime.now().strftime('%Y-%m-%d'),
#         )
#         workbook = Workbook()
# 
#         # Get active worksheet/tab
#         worksheet = workbook.active
#         worksheet.title = 'Data Klaim CBG'
# 
#         # Define the titles for columns
#         columns = [
#             'namars',
#             'status',
#             'NOREG',
#             'NOSEP',
#             'TGLSEP',
#             'TGLPULANG',
#             'JNSPEL',
#             'NOKARTU',
#             'NMPESERTA',
#             'POLI',
#             'KDINACBG',
#             'BYPENGAJUAN',
#             'verifikator',
#             'ALGORITMA',
#             'jenis_pending',
#             'jenis_dispute',
#             'ket_pending',
#             'ket_jawaban',
#         ]
#         row_num = 1
# 
#         # Assign the titles for each cell of the header
#         for col_num, column_title in enumerate(columns, 1):
#             cell = worksheet.cell(row=row_num, column=col_num)
#             cell.value = column_title
# 
#         # Iterate through all
#         for queryset in queryset:
#             row_num += 1
# 
#             if queryset.ket_pending_dispute.last() is None:
#                 ket_pending_disput_queryset = ''
#             else:
#                 ket_pending_disput_queryset = '{0}, '.format(queryset.ket_pending_dispute.last())
#                 ket_pending_disput_queryset = replace_illegal_characters(ket_pending_disput_queryset)
# 
#             if queryset.ket_jawaban_pending.last() is None:
#                 ket_jawaban_pending_queryset = ''
#             else:
#                 ket_jawaban_pending_queryset = '{0}, '.format(queryset.ket_jawaban_pending.last())
#                 ket_jawaban_pending_queryset = replace_illegal_characters(ket_jawaban_pending_queryset)
# 
#             # ket_pending_disput_queryset = ''
#             # for x in queryset.ket_pending_dispute.all():
#             #     ket_pending_disput_queryset += '{0}, '.format(x.ket_pending_dispute)
# 
#             # ket_jawaban_pending_queryset = ''
#             # for x in queryset.ket_jawaban_pending.all():
#             #     ket_jawaban_pending_queryset += '{0}, '.format(x.ket_jawaban_pending)
# 
#             # Define the data for each cell in the row
#             row = [
#                 queryset.faskes.nama,
#                 queryset.status,
#                 queryset.register_klaim.nomor_register_klaim,
#                 queryset.NOSEP,
#                 queryset.TGLSEP,
#                 queryset.TGLPULANG,
#                 queryset.JNSPEL,
#                 queryset.NOKARTU,
#                 queryset.NMPESERTA,
#                 queryset.POLI,
#                 queryset.KDINACBG,
#                 queryset.BYPENGAJUAN,
#                 queryset.verifikator.username,
#                 queryset.ALGORITMA,
#                 queryset.jenis_pending,
#                 queryset.jenis_dispute,
#                 ket_pending_disput_queryset,
#                 ket_jawaban_pending_queryset,
#             ]
# 
#             # Assign the data for each cell of the row
#             for col_num, cell_value in enumerate(row, 1):
#                 cell = worksheet.cell(row=row_num, column=col_num)
#                 cell.value = cell_value
# 
#         workbook.save(response)
#         return response
#     if request.POST.get('import'):
#         # ambil file excel dan buat menjadi dataframe
#         try:
#             file = request.FILES['excel']
#             if file.name.split('.')[-1] != 'xlsx':
#                 raise ValidationError('Hanya menerima file dengan ekstensi `.xlsx`')
#         except Exception as e:
#             messages.warning(request, f'Terdapat Error dalam proses upload dengan keterangan: {e}')
#             return redirect(request.headers.get('Referer'))
# 
#         df_raw = pd.read_excel(file)
# 
#         # replace nan dengan kosong
#         df = df_raw.replace(np.nan, '')
# 
#         # ubah nama kolom menjadi huruf besar
#         df['status'] = df['status'].str.title()
# 
#         # ubah nama jenis pending menjadi huruf besar
#         df['jenis_pending'] = df['jenis_pending'].str.title()
# 
#         # ubah nama jenis dispute menjadi huruf besar
#         df['jenis_dispute'] = df['jenis_dispute'].str.title()
# 
#         # Call the Student Resource Model and make its instance
#         data_claim_resource = DataKlaimCBGResource()
# 
#         # Load the pandas dataframe into a tablib dataset
#         dataset = Dataset().load(df)
# 
#         # cek column sudah sesuai
#         list_kolom_mandatory = ['status', 'NOSEP', 'jenis_pending', 'jenis_dispute', 'ket_pending']
#         list_status_df = df.columns.tolist()
#         for daftar in list_kolom_mandatory:
#             if daftar not in list_status_df:
#                 messages.warning(request, f'File yang diimport harus memiliki kolom {list_kolom_mandatory}')
#                 return redirect(request.headers.get('Referer'))
# 
#         # cek sep tersebut berstatus bukan proses
#         for status in dataset['status']:
#             if status == 'Proses':
#                 messages.warning(request, 'File yang diimport harus berstatus "Layak", "Pending, '
#                                           '"Dispute", "Tidak Layak"')
#                 return redirect(request.headers.get('Referer'))
# 
#         # cek status klaim
#         list_status_mandatory = [StatusDataKlaimChoices.LAYAK, StatusDataKlaimChoices.PENDING,
#                                  StatusDataKlaimChoices.TIDAK_LAYAK, StatusDataKlaimChoices.DISPUTE]
#         list_status_df = list(dict.fromkeys(dataset['status']))
#         for status in list_status_df:
#             if status not in list_status_mandatory:
#                 messages.warning(request, f'File yang diimport harus memiliki status yang sesuai. '
#                                           f'Mohon dapat dicek kembali.')
#                 return redirect(request.headers.get('Referer'))
# 
#         # cek jenis pending
#         list_jenis_mandatory = [JenisPendingChoices.ADMINISTRASI, JenisPendingChoices.KODING,
#                                 JenisPendingChoices.STANDAR_PELAYANAN]
#         list_jenis_pending_df = list(dict.fromkeys(dataset['jenis_pending']))
#         for i in list_jenis_pending_df:
#             if i == '':
#                 list_jenis_pending_df.remove('')
# 
#         for jenis in list_jenis_pending_df:
#             if jenis not in list_jenis_mandatory:
#                 messages.warning(request, f'File yang diimport harus memiliki jenis pending yang sesuai. '
#                                           f'Mohon dapat dicek kembali.')
#                 return redirect(request.headers.get('Referer'))
# 
#         # cek jenis dispute
#         list_jenis_dispute_mandatory = [JenisDisputeChoices.MEDIS, JenisDisputeChoices.KODING, JenisDisputeChoices.COB]
#         list_jenis_dispute_df = list(dict.fromkeys(dataset['jenis_dispute']))
#         for i in list_jenis_dispute_df:
#             if i == '':
#                 list_jenis_dispute_df.remove('')
# 
#         for jenis in list_jenis_dispute_df:
#             if jenis not in list_jenis_dispute_mandatory:
#                 messages.warning(request, f'File yang diimport harus memiliki jenis dispute yang sesuai. '
#                                           f'Mohon dapat dicek kembali.')
#                 return redirect(request.headers.get('Referer'))
# 
#         # cek jumlah jenis pending dan jenis dispute sama dengan yang distatus
#         jumlah_pending = df['status'].eq(StatusDataKlaimChoices.PENDING).sum()
#         jumlah_dispute = df['status'].eq(StatusDataKlaimChoices.DISPUTE).sum()
#         jumlah_tidak_layak = df['status'].eq(StatusDataKlaimChoices.TIDAK_LAYAK).sum()
#         jumlah_jenis_pending = df['jenis_pending'].apply(lambda x: isinstance(x, str) and len(x) > 0).sum().sum()
#         jumlah_jenis_dispute = df['jenis_dispute'].apply(lambda x: isinstance(x, str) and len(x) > 0).sum().sum()
#         jumlah_ket_pending = df['ket_pending'].apply(lambda x: isinstance(x, str) and len(x) > 0).sum().sum()
#         total_pending_dispute = jumlah_pending + jumlah_dispute
#         total_pending_dispute_tidak_layak = jumlah_pending + jumlah_dispute + jumlah_tidak_layak
# 
#         # cek jenis pending
#         if total_pending_dispute > jumlah_jenis_pending:
#             messages.warning(request, f'File yang diimport untuk status "Pending" dan "Dispute" '
#                                       f'harus diisi jenis pending')
#             return redirect(request.headers.get('Referer'))
#         # cek jenis dispute
#         elif jumlah_dispute > jumlah_jenis_dispute:
#             messages.warning(request, f'File yang diimport untuk status "Dispute" harus diisi jenis dispute')
#             return redirect(request.headers.get('Referer'))
#         # cek ket pending
#         elif total_pending_dispute_tidak_layak > jumlah_ket_pending:
#             messages.warning(request, f'File yang diimport untuk status "Pending" dan "Dispute" belum diisi '
#                                       f'keterangan pending')
#             return redirect(request.headers.get('Referer'))
# 
#         # cek sep tersebut memang milik verifikator dan yang diimport adalah status belum diverif
#         for i in dataset['NOSEP']:
#             queryset = DataKlaimCBG.objects.filter(verifikator=request.user, prosesklaim=False, NOSEP=i)
#             if not queryset.exists():
#                 messages.warning(request, f'Terdapat NO SEP {i} yang tidak terdapat dalam verifikasi klaim '
#                                           f'verifikator bersangkutan.')
#                 return redirect(request.headers.get('Referer'))
#             else:
#                 for a in queryset:
#                     if a.status != 'Proses':
#                         messages.warning(request, f'Terdapat NO SEP {a.NOSEP} yang sudah di verifikasi. '
#                                                   f'Status yang diimport tidak boleh dalam keadaan telah diverifikasi.')
#                         return redirect(request.headers.get('Referer'))
# 
#         try:
#             with transaction.atomic():
#                 dataset_new = Dataset().load(df)
#                 list_id = []
#                 ket_pending_dispute_excel = dataset['ket_pending']
#                 for i in ket_pending_dispute_excel:
#                     obj_ket_pending_dispute = KeteranganPendingDispute(ket_pending_dispute=i, verifikator=request.user)
#                     obj_ket_pending_dispute.save()
#                     list_id.append(obj_ket_pending_dispute.id)
# 
#                 df['ket_pending_dispute'] = list_id
# 
#                 for index, row in df.iterrows():
#                     obj = DataKlaimCBG.objects.filter(verifikator=request.user, prosesklaim=False,
#                                                       NOSEP=row['NOSEP']).first()
#                     obj.ket_pending_dispute.add(row['ket_pending_dispute'])
# 
#                 # import
#                 data_claim_resource.import_data(dataset_new, dry_run=False)
# 
#                 # delete yang tidak penting
#                 ket_pending_kosong = KeteranganPendingDispute.objects.filter(ket_pending_dispute='',
#                                                                              verifikator=request.user)
#                 ket_pending_kosong.delete()
#                 messages.success(request, 'Data berhasil diimport. {0}'.format(dataset_new))
#                 return redirect(request.headers.get('Referer'))
#         except Exception as e:
#             # Nampilih error
#             messages.info(request, f'Terjadi kesalahan: {str(e)}')
# 
#     # pagination
#     paginator = Paginator(queryset, 10)
#     page_number = request.GET.get('page')
#     queryset = paginator.get_page(page_number)
# 
#     context = {
#         'data_klaim': queryset,
#         'myFilter': myFilter,
#     }
# 
#     return render(request, 'verifikator/cbg/daftar_data_klaim_cbg.html', context)


@login_required
@check_device
@permissions(role=['verifikator'])
def cek_aksi_vibi_vidi(request):
    data = request.POST
    # data_get = request.GET
    # register_klaim = ""
    queryset = None
    context = {
        'data_klaim': queryset,
        # 'register_klaim': register_klaim
    }

    # if 'register_klaim' in data_get:
    register_klaim = request.GET.get('nomor_register_klaim')
    # status = request.GET.get('status')
    # status_sinkron = request.GET.get('status_sinkron')
    related_kantor_cabang = request.user.kantorcabang_set.all()
    queryset = DataKlaimCBG.objects.filter(
        # verifikator=request.user,
        verifikator__kantorcabang__in=related_kantor_cabang,
        register_klaim__nomor_register_klaim=register_klaim,
        # prosesklaim=False,
        # status=status,
        # status_sinkron=status_sinkron,
    ).order_by('NOSEP', 'TGLSEP')
    # print(queryset)

    # filter
    myFilter = SinkronisasiVIBIVIDIFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    context.update({
        'data_klaim': queryset,
        'register_klaim': register_klaim,
        'myFilter': myFilter,
    })

    if 'no_sep' in data:
        nosep = data['no_sep']
        status_vidi = data['status_vidi']
        try:
            klaim = DataKlaimCBG.objects.get(NOSEP=nosep)
            klaim.status_vidi = status_vidi
            # klaim.status_sinkron = 'Sinkron'
            if status_vidi == klaim.status:
                status_sinkron = StatusSinkronChoices.SINKRON
            elif status_vidi != klaim.status:
                status_sinkron = StatusSinkronChoices.TIDAK_SINKRON
            else:
                status_sinkron = ''
            klaim.status_sinkron = status_sinkron
            klaim.tgl_sinkron = datetime.datetime.today()
            klaim.save()
            context.update({
                'klaim': klaim,
            })
            response_data = {
                'no_sep': nosep,
                'status_vidi': status_vidi,
                'message': 'Berhasil Update',
            }
            return JsonResponse(response_data)
        except DataKlaimCBG.DoesNotExist:
            response_data = {
                'no_sep': nosep,
                'message': 'Data tidak ditemukan',
            }
            return JsonResponse(response_data, status=404)

    return render(request, 'verifikator/cbg/cek_aksi_vibi_vidi.html', context)


@login_required
@check_device
@permissions(role=['verifikator'])
def sinkronisasi_aksi_vibi_vidi(request):
    data = request.POST
    # data_get = request.GET
    # register_klaim = ""
    queryset = None
    context = {
        'data_klaim': queryset,
        # 'register_klaim': register_klaim
    }

    # if 'register_klaim' in data_get:
    register_klaim = request.GET.get('nomor_register_klaim')
    # status = request.GET.get('status')
    # status_sinkron = request.GET.get('status_sinkron')
    related_kantor_cabang = request.user.kantorcabang_set.all()
    queryset = DataKlaimCBG.objects.filter(
        # verifikator=request.user,
        verifikator__kantorcabang__in=related_kantor_cabang,
        register_klaim__nomor_register_klaim=register_klaim,
        verifikator=request.user).order_by('NOSEP', 'TGLSEP')
    # print(queryset)

    # filter
    myFilter = SinkronisasiVIBIVIDIFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    context.update({
        'data_klaim': queryset,
        'register_klaim': register_klaim,
        'myFilter': myFilter,
    })

    if 'no_sep' in data:
        nosep = data['no_sep']
        status_vidi = data['status_vidi']
        try:
            klaim = DataKlaimCBG.objects.get(NOSEP=nosep)
            klaim.status_vidi = status_vidi
            # klaim.status_sinkron = 'Sinkron'
            if status_vidi == klaim.status:
                status_sinkron = StatusSinkronChoices.SINKRON
            elif status_vidi != klaim.status:
                status_sinkron = StatusSinkronChoices.TIDAK_SINKRON
            else:
                status_sinkron = ''
            klaim.status_sinkron = status_sinkron
            klaim.tgl_sinkron = datetime.datetime.today()
            klaim.save()
            context.update({
                'klaim': klaim,
            })
            response_data = {
                'no_sep': nosep,
                'status_vidi': status_vidi,
                'message': 'Berhasil Update',
            }
            return JsonResponse(response_data)
        except DataKlaimCBG.DoesNotExist:
            response_data = {
                'no_sep': nosep,
                'message': 'Data tidak ditemukan',
            }
            return JsonResponse(response_data, status=404)

    return render(request, 'verifikator/cbg/sinkronisasi_aksi_vibi_vidi.html', context)


@csrf_exempt
@login_required
@check_device
@permissions(role=['verifikator'])
def sinkronisasi_vibi_vidi_download(request):
    data = request.POST
    if data.get('no_sep'):
        nosep = data.get('no_sep')
        status_vidi = data.get('status_vidi')
        status_sinkron = data.get('status_sinkron')
        queryset = DataKlaimCBG.objects.get(NOSEP=nosep)
        queryset.status_vidi = status_vidi
        # queryset.status_sinkron = 'Sinkron'
        queryset.save()
        response_data = {
            'no_sep': nosep,
            'status_vidi': status_vidi,
            'status_sinkron': status_sinkron,
            'message': 'Berhasil Update',
        }
        return JsonResponse(response_data)
    queryset = DataKlaimCBG.objects.filter(verifikator=request.user, prosesklaim=False).order_by('NOSEP', 'TGLSEP')
    context = {
        'data_klaim': queryset,
    }
    return render(request, 'verifikator/cbg/sinkronisasi_vibi_vidi_download.html', context)


# @login_required
# @check_device
# @permissions(role=['verifikator'])
# def detail_data_klaim(request, pk):
#     dataklaimcbg = DataKlaimCBG.objects.filter(verifikator=request.user).get(id=pk)
#     listchoicestatus = STATUS_CHOICES_DATA_KLAIM_VERIFIKATOR
#     context = {
#         "dataklaimcbg": dataklaimcbg,
#         "listchoicestatus": listchoicestatus,
#     }
#     return render(request, 'verifikator/detail_data_klaim_cbg.html', context)


# def edit_detail_data_klaim(request):
#     data = request.POST
#     if data.get('id_dataklaimcbg'):
#         dataklaimcbg = DataKlaimCBG.objects.get(id=data.get('id_dataklaimcbg'))
#
#         if dataklaimcbg.is_hitung is False:
#             add_hitung_data_klaim = HitungDataKlaim(nomor_register_klaim=dataklaimcbg.register_klaim.nomor_register_klaim,
#                                                 jenis_klaim=dataklaimcbg.register_klaim.jenis_klaim,
#                                                 verifikator=dataklaimcbg.verifikator)
#             add_hitung_data_klaim.save()
#             dataklaimcbg.is_hitung = True
#             dataklaimcbg.save()
#
#         keterangan = data.get('keterangan')
#         keterangan_pending_dispute = KeteranganPendingDispute(
#             ket_pending_dispute=keterangan
#         )
#         keterangan_pending_dispute.save()
#         dataklaimcbg.ket_pending_dispute.add(keterangan_pending_dispute)
#         dataklaimcbg.save()
#
#         next = request.POST.get('next', '/')
#         return HttpResponseRedirect(next)
#     else:
#         messages.info(request, "Hmmm sepertinya ada yang salah.")
#         return redirect('/')


@login_required
@check_device
@permissions(role=['verifikator'])
def detail_data_klaim(request, pk):
    queryset = DataKlaimCBG.objects.filter(verifikator=request.user)
    instance = queryset.get(pk=pk)
    data_klaim_form = DataKlaimVerifikatorForm(instance=instance)
    data_klaim_pending_form = KeteranganPendingForm()

    data_klaim_cbg_metafisik = DataKlaimCBGMetafisik.objects.filter(nosjp=instance.NOSEP).first()

    if request.method == 'POST':
        data_klaim_form = DataKlaimVerifikatorForm(instance=instance, data=request.POST)
        keterangan = KeteranganPendingForm(request.POST or None)
        if data_klaim_form.is_valid():
            if instance.is_hitung is False:
                HitungDataKlaim.objects.create(nomor_register_klaim=instance.register_klaim.nomor_register_klaim,
                                               jenis_klaim=instance.register_klaim.jenis_klaim,
                                               verifikator=instance.verifikator)
                instance.is_hitung = True
                instance.save()
            if instance.status != StatusDataKlaimChoices.LAYAK:
                obj_keterangan = keterangan.save()
                obj_keterangan.verifikator = instance.verifikator
                obj_keterangan.save()
                instance.ket_pending_dispute.add(obj_keterangan)
            data_klaim_form.save()
            next = request.POST.get('next', '/')
            messages.success(request, f'NO SEP {instance.NOSEP} berhasil diupdate.')
            return HttpResponseRedirect(next)

    context = {
        'data_klaim': instance,
        'data_klaim_form': data_klaim_form,
        'data_klaim_pending_form': data_klaim_pending_form,
        'meta': data_klaim_cbg_metafisik
    }
    return render(request, 'verifikator/cbg/detail_data_klaim_cbg.html', context)


@login_required
@check_device
@permissions(role=['verifikator'])
def finalisasi_data_klaim(request):
    queryset = RegisterKlaim.objects.filter(
        nomor_register_klaim__startswith=request.user.kantorcabang_set.all().first().kode_cabang,
        status=StatusRegisterChoices.VERIFIKASI).order_by('created_at')

    # filter
    myFilter = RegisterKlaimFaskesFilter(request.GET, queryset=queryset, request=request)
    queryset = myFilter.qs

    # pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'register_list': queryset,
        'myFilter': myFilter,
    }

    return render(request, 'verifikator/finalisasi_data_klaim.html', context)


@login_required
@check_device
@permissions(role=['verifikator'])
def update_finalisasi_data_klaim(request, pk):
    kantor_cabang = request.user.kantorcabang_set.all().first()
    queryset = RegisterKlaim.objects.filter(
        nomor_register_klaim__startswith=kantor_cabang.kode_cabang)
    instance = queryset.get(id=pk)

    # count
    data_klaim = None
    jumlah_proses = 0
    jumlah_layak = 0
    jumlah_pending = 0
    jumlah_dispute = 0
    jumlah_tidak_layak = 0
    jumlah_klaim = 0
    total_klaim = 0
    biaya_proses = 0
    biaya_layak = 0
    biaya_pending = 0
    biaya_dispute = 0
    biaya_tidak_layak = 0
    biaya_klaim = 0
    try:
        if instance.jenis_klaim.nama == NamaJenisKlaimChoices.CBG_REGULER or instance.jenis_klaim.nama == NamaJenisKlaimChoices.CBG_SUSULAN:
            data_klaim = DataKlaimCBG.objects.filter(register_klaim=instance)
            jumlah_proses = data_klaim.filter(status=StatusDataKlaimChoices.PROSES).count()
            jumlah_layak = data_klaim.filter(status=StatusDataKlaimChoices.LAYAK).count()
            jumlah_pending = data_klaim.filter(status=StatusDataKlaimChoices.PENDING).count()
            jumlah_dispute = data_klaim.filter(status=StatusDataKlaimChoices.DISPUTE).count()
            jumlah_tidak_layak = data_klaim.filter(status=StatusDataKlaimChoices.TIDAK_LAYAK).count()
            jumlah_klaim = data_klaim.filter(status=StatusDataKlaimChoices.KLAIM).count()
            total_klaim = data_klaim.count()

            biaya_proses = data_klaim.filter(status=StatusDataKlaimChoices.PROSES).aggregate(Sum('BYPENGAJUAN'))[
                'BYPENGAJUAN__sum']
            biaya_layak = data_klaim.filter(status=StatusDataKlaimChoices.LAYAK).aggregate(Sum('BYPENGAJUAN'))[
                'BYPENGAJUAN__sum']
            biaya_pending = data_klaim.filter(status=StatusDataKlaimChoices.PENDING).aggregate(Sum('BYPENGAJUAN'))[
                'BYPENGAJUAN__sum']
            biaya_dispute = data_klaim.filter(status=StatusDataKlaimChoices.DISPUTE).aggregate(Sum('BYPENGAJUAN'))[
                'BYPENGAJUAN__sum']
            biaya_tidak_layak = \
            data_klaim.filter(status=StatusDataKlaimChoices.TIDAK_LAYAK).aggregate(Sum('BYPENGAJUAN'))[
                'BYPENGAJUAN__sum']
            biaya_klaim = data_klaim.aggregate(Sum('BYPENGAJUAN'))['BYPENGAJUAN__sum']
        elif instance.jenis_klaim.nama == NamaJenisKlaimChoices.OBAT_REGULER or instance.jenis_klaim.nama == NamaJenisKlaimChoices.OBAT_SUSULAN:
            data_klaim = DataKlaimObat.objects.filter(register_klaim=instance)
            jumlah_proses = data_klaim.filter(status=StatusDataKlaimChoices.PROSES).count()
            jumlah_layak = data_klaim.filter(status=StatusDataKlaimChoices.LAYAK).count()
            jumlah_pending = data_klaim.filter(status=StatusDataKlaimChoices.PENDING).count()
            jumlah_dispute = data_klaim.filter(status=StatusDataKlaimChoices.DISPUTE).count()
            jumlah_tidak_layak = data_klaim.filter(status=StatusDataKlaimChoices.TIDAK_LAYAK).count()
            jumlah_klaim = data_klaim.filter(status=StatusDataKlaimChoices.KLAIM).count()
            total_klaim = data_klaim.count()
            biaya_proses = data_klaim.filter(status=StatusDataKlaimChoices.PROSES).aggregate(Sum('ByVerApt'))[
                'ByVerApt__sum']
            biaya_layak = data_klaim.filter(status=StatusDataKlaimChoices.LAYAK).aggregate(Sum('ByVerApt'))[
                'ByVerApt__sum']
            biaya_pending = data_klaim.filter(status=StatusDataKlaimChoices.PENDING).aggregate(Sum('ByVerApt'))[
                'ByVerApt__sum']
            biaya_dispute = data_klaim.filter(status=StatusDataKlaimChoices.DISPUTE).aggregate(Sum('ByVerApt'))[
                'ByVerApt__sum']
            biaya_tidak_layak = data_klaim.filter(status=StatusDataKlaimChoices.TIDAK_LAYAK).aggregate(Sum('ByVerApt'))[
                'ByVerApt__sum']
            biaya_klaim = data_klaim.aggregate(Sum('ByVerApt'))['ByVerApt__sum']
    except Exception as e:
        HttpResponse(content=e)

    status_form = FinalisasiVerifikatorForm(instance=instance)
    if request.method == 'POST' and request.POST.get('action') == 'update_status':
        status_form = FinalisasiVerifikatorForm(instance=instance, data=request.POST)
        print(instance.no_ba_lengkap)
        if instance.verifikator != request.user:
            return HttpResponse(content="Anda Tidak Memiliki Hak Akses, Harap Menghubungi Admin!", status=403)
        if instance.sisa_klaim > 0 and instance.no_ba_lengkap is None:
            messages.warning(request, "Tidak dapat difinalisasi! Masih ada sisa klaim yang belum diverifikasi "
                                      "dan BA Lengkap belum diupdate. Terima Kasih.")
        elif instance.sisa_klaim > 0:
            messages.warning(request, "Tidak dapat difinalisasi! Masih ada sisa klaim yang belum diverifikasi. "
                                      "Terima Kasih.")
        elif instance.no_ba_lengkap is None:
            messages.warning(request, "Tidak dapat difinalisasi! BA Lengkap belum diupdate. Terima Kasih.")
        elif status_form.is_valid() and instance.sisa_klaim == 0:
            status_form.save()
            if instance.is_final is False:
                instance.is_final = True
                instance.save()
            if (instance.jenis_klaim.nama == NamaJenisKlaimChoices.CBG_REGULER
                    or instance.jenis_klaim.nama == NamaJenisKlaimChoices.CBG_SUSULAN
                    or instance.jenis_klaim.nama == NamaJenisKlaimChoices.OBAT_REGULER
                    or instance.jenis_klaim.nama == NamaJenisKlaimChoices.OBAT_SUSULAN):
                data_klaim.filter(status=StatusDataKlaimChoices.LAYAK).update(status=StatusDataKlaimChoices.KLAIM,
                                                                              prosespending=False,
                                                                              prosesdispute=False)
                data_klaim.filter(status=StatusDataKlaimChoices.PENDING).update(prosespending=True)
                data_klaim.filter(status=StatusDataKlaimChoices.DISPUTE).update(prosespending=True, prosesdispute=True)
                data_klaim.update(prosesklaim=True)
            messages.success(request, "Register dan Data Klaim berhasil difinalisasi. Terima Kasih.")
            return redirect(request.headers.get('Referer'))

    context = {
        'register': instance,
        'status_form': status_form,
        'jumlah_proses': jumlah_proses,
        'jumlah_layak': jumlah_layak,
        'jumlah_pending': jumlah_pending,
        'jumlah_dispute': jumlah_dispute,
        'jumlah_tidak_layak': jumlah_tidak_layak,
        'jumlah_klaim': jumlah_klaim,
        'total_klaim': total_klaim,
        'biaya_proses': biaya_proses,
        'biaya_layak': biaya_layak,
        'biaya_pending': biaya_pending,
        'biaya_dispute': biaya_dispute,
        'biaya_tidak_layak': biaya_tidak_layak,
        'biaya_klaim': biaya_klaim,
    }
    return render(request, 'verifikator/update_finalisasi_data_klaim.html', context)


@login_required
@check_device
@permissions(role=['verifikator'])
def update_data_klaim_cbg(request, pk):
    data = dict()
    data_klaim = get_object_or_404(DataKlaimCBG, id=pk)

    if request.method == 'POST':
        form = DataKlaimVerifikatorForm(request.POST, instance=data_klaim)
        if form.is_valid():
            form.save()
            data['form_is_valid'] = True
        else:
            data['form_is_valid'] = False
    else:
        form = DataKlaimVerifikatorForm(instance=data_klaim)
    context = {
        'form': form,
        'data_klaim': data_klaim,
    }
    data['html_form'] = render_to_string(request=request, template_name='verifikator/update_data_klaim.html',
                                         context=context)
    return JsonResponse(data)


@login_required
@check_device
@permissions(role=['verifikator'])
def download_data_cbg(request):
    # initial relasi pada kantor cabang
    related_kantor_cabang = request.user.kantorcabang_set.all()
    queryset = DataKlaimCBG.objects.filter(verifikator__kantorcabang__in=related_kantor_cabang)

    # filter
    myFilter = DownloadDataKlaimCBGFilter(request.GET, queryset=queryset, request=request)
    queryset = myFilter.qs

    kantor_cabang = request.user.kantorcabang_set.all().first()

    # fitur download
    download = request.GET.get('download')
    bupel_month = request.GET.get('bupel_month')
    bupel_year = request.GET.get('bupel_year')
    faskes = request.GET.get('faskes')
    nomor_register = request.GET.get('nomor_register_klaim')
    try:
        if nomor_register != '':
            if download == 'download':
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                response['Content-Disposition'] = 'attachment; filename={date}-dataverifikasi.xlsx'.format(
                    date=datetime.datetime.now().strftime('%Y-%m-%d'),
                )
                workbook = Workbook()

                # Get active worksheet/tab
                worksheet = workbook.active
                worksheet.title = 'Data Klaim CBG'

                # Define the titles for columns
                columns = [
                    'namars',
                    'status',
                    'NOREG',
                    'NOSEP',
                    'TGLSEP',
                    'TGLPULANG',
                    'JNSPEL',
                    'NOKARTU',
                    'NMPESERTA',
                    'POLI',
                    'KDINACBG',
                    'BYPENGAJUAN',
                    'ALGORITMA',
                    'verifikator',
                    'jenis_pending',
                    'jenis_dispute',
                    'ket_pending',
                    'ket_jawaban',
                ]
                row_num = 1

                # Assign the titles for each cell of the header
                for col_num, column_title in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title

                # Iterate through all movies
                for queryset in queryset:
                    row_num += 1

                    if queryset.ket_pending_dispute.last() is None:
                        ket_pending_disput_queryset = ''
                    else:
                        ket_pending_disput_queryset = '{0}, '.format(queryset.ket_pending_dispute.last())
                        ket_pending_disput_queryset = replace_illegal_characters(ket_pending_disput_queryset)

                    if queryset.ket_jawaban_pending.last() is None:
                        ket_jawaban_pending_queryset = ''
                    else:
                        ket_jawaban_pending_queryset = '{0}, '.format(queryset.ket_jawaban_pending.last())
                        ket_jawaban_pending_queryset = replace_illegal_characters(ket_jawaban_pending_queryset)

                    # ket_pending_disput_queryset = ''
                    # for x in queryset.ket_pending_dispute.all():
                    #     ket_pending_disput_queryset += '{0}, '.format(x.ket_pending_dispute)
                    #
                    # ket_jawaban_pending_queryset = ''
                    # for x in queryset.ket_jawaban_pending.all():
                    #     ket_jawaban_pending_queryset += '{0}, '.format(x.ket_jawaban_pending)

                    # Define the data for each cell in the row
                    row = [
                        queryset.faskes.nama,
                        queryset.status,
                        queryset.register_klaim.nomor_register_klaim,
                        queryset.NOSEP,
                        queryset.TGLSEP,
                        queryset.TGLPULANG,
                        queryset.JNSPEL,
                        queryset.NOKARTU,
                        queryset.NMPESERTA,
                        queryset.POLI,
                        queryset.KDINACBG,
                        queryset.BYPENGAJUAN,
                        queryset.ALGORITMA,
                        queryset.verifikator.username,
                        queryset.jenis_pending,
                        queryset.jenis_dispute,
                        ket_pending_disput_queryset,
                        ket_jawaban_pending_queryset,
                    ]

                    # Assign the data for each cell of the row
                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value

                workbook.save(response)
                return response
        elif bupel_month != '' and bupel_year != '' and faskes != '':
            if download == 'download':
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                response['Content-Disposition'] = 'attachment; filename={date}-dataverifikasi.xlsx'.format(
                    date=datetime.datetime.now().strftime('%Y-%m-%d'),
                )
                workbook = Workbook()

                # Get active worksheet/tab
                worksheet = workbook.active
                worksheet.title = 'Data Klaim CBG'

                # Define the titles for columns
                columns = [
                    'namars',
                    'status',
                    'NOREG',
                    'NOSEP',
                    'TGLSEP',
                    'TGLPULANG',
                    'JNSPEL',
                    'NOKARTU',
                    'NMPESERTA',
                    'POLI',
                    'KDINACBG',
                    'BYPENGAJUAN',
                    'ALGORITMA',
                    'verifikator',
                    'jenis_pending',
                    'jenis_dispute',
                    'ket_pending',
                    'ket_jawaban',
                ]
                row_num = 1

                # Assign the titles for each cell of the header
                for col_num, column_title in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title

                # Iterate through all movies
                for queryset in queryset:
                    row_num += 1

                    if queryset.ket_pending_dispute.last() is None:
                        ket_pending_disput_queryset = ''
                    else:
                        ket_pending_disput_queryset = '{0}, '.format(queryset.ket_pending_dispute.last())
                        ket_pending_disput_queryset = replace_illegal_characters(ket_pending_disput_queryset)

                    if queryset.ket_jawaban_pending.last() is None:
                        ket_jawaban_pending_queryset = ''
                    else:
                        ket_jawaban_pending_queryset = '{0}, '.format(queryset.ket_jawaban_pending.last())
                        ket_jawaban_pending_queryset = replace_illegal_characters(ket_jawaban_pending_queryset)

                    # ket_pending_disput_queryset = ''
                    # for x in queryset.ket_pending_dispute.all():
                    #     ket_pending_disput_queryset += '{0}, '.format(x.ket_pending_dispute)
                    #
                    # ket_jawaban_pending_queryset = ''
                    # for x in queryset.ket_jawaban_pending.all():
                    #     ket_jawaban_pending_queryset += '{0}, '.format(x.ket_jawaban_pending)

                    # Define the data for each cell in the row
                    row = [
                        queryset.faskes.nama,
                        queryset.status,
                        queryset.register_klaim.nomor_register_klaim,
                        queryset.NOSEP,
                        queryset.TGLSEP,
                        queryset.TGLPULANG,
                        queryset.JNSPEL,
                        queryset.NOKARTU,
                        queryset.NMPESERTA,
                        queryset.POLI,
                        queryset.KDINACBG,
                        queryset.BYPENGAJUAN,
                        queryset.ALGORITMA,
                        queryset.verifikator.username,
                        queryset.jenis_pending,
                        queryset.jenis_dispute,
                        ket_pending_disput_queryset,
                        ket_jawaban_pending_queryset,
                    ]

                    # Assign the data for each cell of the row
                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value

                workbook.save(response)
                return response
        else:
            messages.warning(request, 'Bulan, Tahun, dan Faskes harus diisi atau Nomor Register Klaim harus diisi!')
    except Exception as e:
        messages.warning(request, "Terjadi Kesalahan Dalam Download Data, dengan Keterangan: " + str(e))

    context = {
        'myFilter': myFilter,
    }
    return render(request, 'verifikator/cbg/download_data_cbg.html', context)


class RumahSakitAutocomplete(autocomplete.Select2QuerySetView):
    # def __init__(self, *args, **kwargs):
    #     request = kwargs.pop('request', None)  # Dapatkan 'request' dari kwargs
    #     super().__init__(*args, **kwargs)
    def get_queryset(self, *args, **kwargs):
        request = kwargs.pop('request', None)  # Dapatkan 'request' dari kwargs
        super().__init__(*args, **kwargs)
        # Don't forget to filter out results depending on the visitor !
        if not self.request.user.is_authenticated:
            return Faskes.objects.none()

        qs = Faskes.objects.filter(kantor_cabang__in=self.request.user.kantorcabang_set.all())

        if self.q:
            qs = qs.filter(nama__icontains=self.q)
        return qs


#######################
# VERIFIKASI KLAIM OBAT
#######################


@login_required
@check_device
@permissions(role=['verifikator'])
def import_data_klaim_obat(request):
    storage = TemporaryStorage()
    import_form = ImportDataKlaimForm()
    verifikator = User.objects.filter(kantorcabang__in=request.user.kantorcabang_set.all(),
                                      groups__name='verifikator',
                                      is_active=True,
                                      is_staff=True)
    if request.method == 'POST' and request.POST.get('action') == 'import':
        import_form = ImportDataKlaimForm(files=request.FILES, data=request.POST)
        if import_form.is_valid():
            list_data_import = ['KdJenis', 'No SEP Apotek', 'No SEP Asal Resep', 'No Kartu', 'Nama Peserta', 'No Resep',
                                'Tgl Resep', 'By Tag Apt', 'By Ver Apt', 'rufil']
            nomor_register_klaim = import_form.cleaned_data.get('register')
            register = RegisterKlaim.objects.get(nomor_register_klaim=nomor_register_klaim)
            file_name = f'{uuid.uuid4()}-{int(round(time.time() * 1000))}.xlsx'
            storage.save(name=file_name, content=import_form.cleaned_data.get('file'))
            get_password = request.POST.get('password')
            if get_password != '':
                unlocked_file = io.BytesIO()
                try:
                    with open(storage.path(name=file_name), "rb") as file:
                        excel_file = msoffcrypto.OfficeFile(file)
                        excel_file.load_key(password=get_password)
                        excel_file.decrypt(unlocked_file)
                    data_frame = pd.read_excel(unlocked_file, usecols=list_data_import)
                except InvalidKeyError:
                    messages.warning(request, 'Password File Excel yang Anda masukkan salah!')
                    return redirect('/verifikator/import-data-klaim-obat')
                except DecryptionError:
                    messages.warning(request, 'File Excel seharusnya tidak memiliki password!')
                    return redirect('/verifikator/import-data-klaim-obat')
                except Exception as e:
                    messages.warning(request, f'Terjadi Kesalahan pada saat import File. Keterangan error : {e}')
                    return redirect('/verifikator/import-data-klaim-obat')
            else:
                try:
                    data_frame = pd.read_excel(storage.path(name=file_name), usecols=list_data_import)
                except XLRDError:
                    messages.warning(request,
                                     f'File yang Anda import memiliki password. Silahkan masukkan password file!')
                    return redirect('/verifikator/import-data-klaim-obat')
                except Exception as e:
                    messages.warning(request, f'Terjadi kesalahan pada saat import File. Keterangan error : {e}')
                    return redirect('/verifikator/import-data-klaim-obat')
                # data_frame = pd.read_excel(storage.path(name=file_name), usecols=list_data_import, engine='openpyxl', password='Qwerty1!')
            data_frame = data_frame.replace(np.nan, None)
            data_frame['register_klaim'] = register
            data_frame['faskes'] = register.faskes
            data_frame['Tgl Resep'] = pd.to_datetime(data_frame['Tgl Resep'])
            data_frame['bupel'] = data_frame['Tgl Resep'].dt.to_period('M').dt.to_timestamp()
            data_frame['bupel'] = pd.to_datetime(data_frame['bupel'])
            data_frame = data_frame.rename(columns={'No SEP Apotek': 'NoSEPApotek',
                                                    'No SEP Asal Resep': 'NoSEPAsalResep',
                                                    'No Kartu': 'NoKartu',
                                                    'Nama Peserta': 'NamaPeserta',
                                                    'No Resep': 'NoResep',
                                                    'Tgl Resep': 'TglResep',
                                                    'By Tag Apt': 'ByTagApt',
                                                    'By Ver Apt': 'ByVerApt'
                                                    })
            register.password_file_excel = get_password
            register.save()

            # with transaction.atomic():
            try:
                obj_list = []
                for _, row in data_frame.iterrows():
                    data_klaim = DataKlaimObat()
                    try:
                        data_klaim = DataKlaimObat(**dict(row))
                        data_klaim.full_clean()  # Validate the object
                        obj_list.append(data_klaim)
                    except TypeError as e:
                        messages.warning(request, f'Terjadi kesalahan pada saat import File. Keterangan error : {e}')
                        return redirect('/verifikator/import-data-klaim-obat')
                    except Exception as e:
                        messages.warning(request,
                                         f'Kesalahan terjadi pada No SEP Apotek : {data_klaim.NoSEPApotek}. Keterangan error : {e}')
                        return redirect('/verifikator/import-data-klaim-obat')
                # DataKlaimObat.objects.bulk_create(obj_list)

                with transaction.atomic():
                    DataKlaimObat.objects.bulk_create(obj_list)
                    valid_data = DataKlaimObat.objects.filter(id__in=[obj.id for obj in obj_list
                                                                      if
                                                                      obj.NoSEPAsalResep[:8] == register.faskes.kode_ppk and
                                                                      obj.bupel.month == register.bulan_pelayanan.month and
                                                                      obj.bupel.year == register.bulan_pelayanan.year])
                    invalid_data = DataKlaimObat.objects.filter(id__in=[obj.id for obj in obj_list
                                                                        if obj.NoSEPAsalResep[
                                                                           :8] != register.faskes.kode_ppk or
                                                                        obj.bupel.month != register.bulan_pelayanan.month or
                                                                        obj.bupel.year != register.bulan_pelayanan.year])

                    df_valid = pd.DataFrame.from_records(valid_data.values())
                    total_data_valid = len(df_valid)
                    df_invalid = pd.DataFrame.from_records(invalid_data.values())
                    total_data_invalid = len(df_invalid)
                    transaction.set_rollback(True)

            except IntegrityError:
                messages.info(request, 'Terjadi kesalahan saat mencoba menyimpan data.')
                return redirect('/verifikator/import-data-klaim-obat')
            except Exception as e:
                messages.info(request, f'Kesalahan terjadi pada saat import File. Keterangan error : {e}')
                return redirect('/verifikator/import-data-klaim-obat')

            return render(request, 'verifikator/obat/preview_data_import_obat.html',
                          {'preview_data_valid': df_valid,
                           'total_data_valid': total_data_valid,
                           'preview_data_invalid': df_invalid,
                           'total_data_invalid': total_data_invalid,
                           'file_name': file_name,
                           'register': nomor_register_klaim,
                           'password': get_password})

    if request.method == 'POST' and request.POST.get('action') == 'confirm':
        list_data_import = ['KdJenis', 'No SEP Apotek', 'No SEP Asal Resep', 'No Kartu', 'Nama Peserta', 'No Resep',
                            'Tgl Resep', 'By Tag Apt', 'By Ver Apt', 'rufil']
        file_name = request.POST.get('file_name')
        nomor_register_klaim = request.POST.get('register')
        get_password = request.POST.get('password')
        register = RegisterKlaim.objects.get(nomor_register_klaim=nomor_register_klaim)
        if get_password != '':
            unlocked_file = io.BytesIO()
            try:
                with open(storage.path(name=file_name), "rb") as file:
                    excel_file = msoffcrypto.OfficeFile(file)
                    excel_file.load_key(password=request.POST.get('password'))
                    excel_file.decrypt(unlocked_file)
                data_frame = pd.read_excel(unlocked_file, usecols=list_data_import)
            except InvalidKeyError:
                messages.warning(request, 'Password File Excel yang Anda masukkan salah!')
                return redirect('/verifikator/import-data-klaim-obat')
            except DecryptionError:
                messages.warning(request, 'File Excel seharusnya tidak memiliki password!')
                return redirect('/verifikator/import-data-klaim-obat')
            except Exception as e:
                messages.warning(request, f'Terjadi Kesalahan pada saat import File. Keterangan error : {e}')
                return redirect('/verifikator/import-data-klaim-obat')
        else:
            try:
                data_frame = pd.read_excel(storage.path(name=file_name), usecols=list_data_import)
            except XLRDError:
                messages.warning(request, f'File yang Anda import memiliki password. Silahkan masukkan password file!')
                return redirect('/verifikator/import-data-klaim-obat')
            except Exception as e:
                messages.warning(request, f'Terjadi kesalahan pada saat import File. Keterangan error : {e}')
                return redirect('/verifikator/import-data-klaim-obat')

        # data_frame = pd.read_excel(storage.path(name=file_name), usecols=list_data_import)
        data_frame = data_frame.replace(np.nan, None)
        data_frame['register_klaim'] = register
        data_frame['faskes'] = register.faskes
        data_frame['Tgl Resep'] = pd.to_datetime(data_frame['Tgl Resep'])
        data_frame['bupel'] = data_frame['Tgl Resep'].dt.to_period('M').dt.to_timestamp()
        data_frame['bupel'] = pd.to_datetime(data_frame['bupel'])
        data_frame = data_frame.rename(columns={'No SEP Apotek': 'NoSEPApotek',
                                                'No SEP Asal Resep': 'NoSEPAsalResep',
                                                'No Kartu': 'NoKartu',
                                                'Nama Peserta': 'NamaPeserta',
                                                'No Resep': 'NoResep',
                                                'Tgl Resep': 'TglResep',
                                                'By Tag Apt': 'ByTagApt',
                                                'By Ver Apt': 'ByVerApt'
                                                })
        try:
            with transaction.atomic():
                register.has_import_data = True
                register.save()
            with transaction.atomic():
                obj_list = DataKlaimObat.objects.bulk_create(
                    [DataKlaimObat(**dict(row[1])) for row in data_frame.iterrows()]
                )

                queryset = DataKlaimObat.objects.filter(register_klaim=register, status=StatusDataKlaimChoices.BELUM_VER)

                index = random.randrange(verifikator.count())

                # with transaction.atomic():
                objs_to_update = []
                for kd_jenis in [1, 2, 3]:
                    objs = queryset.filter(KdJenis=kd_jenis)
                    if objs.exists():
                        for obj in objs:
                            obj.verifikator = verifikator[index]
                            obj.status = StatusDataKlaimChoices.PROSES
                            objs_to_update.append(obj)
                            index = (index + 1) % verifikator.count()
                if objs_to_update:
                    DataKlaimObat.objects.bulk_update(objs_to_update, ['verifikator', 'status'])

                queryset_proses = DataKlaimObat.objects.filter(register_klaim=register,
                                                        status=StatusDataKlaimChoices.PROSES)

                # penentuan SLA
                sla = SLA.objects.filter(jenis_klaim=register.jenis_klaim,
                                         kantor_cabang=register.faskes.kantor_cabang).first()
                if sla:
                    if register.tgl_ba_lengkap:
                        queryset_proses.update(tgl_SLA=register.tgl_ba_lengkap + timedelta(days=sla.plus_hari_sla))
                    elif register.tgl_terima:
                        queryset_proses.update(tgl_SLA=register.tgl_terima + timedelta(days=15))
                else:
                    if register.tgl_ba_lengkap:
                        queryset_proses.update(tgl_SLA=register.tgl_ba_lengkap + timedelta(days=6))
                    elif register.tgl_terima:
                        queryset_proses.update(tgl_SLA=register.tgl_terima + timedelta(days=15))

            messages.success(request, "Data Obat Berhasil Di-import")
            # # Obat PRB
            # with transaction.atomic():
            #     for obj in queryset.filter(KdJenis=1):
            #         obj.verifikator = verifikator[index]
            #         obj.status = StatusDataKlaimChoices.PROSES
            #         obj.save()
            #         if index == verifikator.count() - 1:
            #             index = 0
            #         else:
            #             index += 1
            #
            # # Obat Kronis
            # with transaction.atomic():
            #     for obj in queryset.filter(KdJenis=2):
            #         obj.verifikator = verifikator[index]
            #         obj.status = StatusDataKlaimChoices.PROSES
            #         obj.save()
            #         if index == verifikator.count() - 1:
            #             index = 0
            #         else:
            #             index += 1
            #
            # # obat kemoterapi
            # with transaction.atomic():
            #     for obj in queryset.filter(KdJenis=3):
            #         obj.verifikator = verifikator[index]
            #         obj.status = StatusDataKlaimChoices.PROSES
            #         obj.save()
            #         if index == verifikator.count() - 1:
            #             index = 0
            #         else:
            #             index += 1
        except Exception as e:
            messages.error(request, f"Terjadi kesalahan pada saat import data. Keterangan error : {e}")
            return redirect('/verifikator/import-data-klaim-obat')

    context = {
        'import_form': import_form,
        'verifikator': verifikator,
    }
    return render(request, 'verifikator/obat/import_data_klaim_obat.html', context)


@login_required
@check_device
@permissions(role=['verifikator'])
def daftar_data_klaim_obat(request):
    queryset = DataKlaimObat.objects.filter(verifikator=request.user, prosesklaim=False).order_by('NamaPeserta',
                                                                                                  'TglResep')

    # filter
    myFilter = DataKlaimObatFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs
    export = request.GET.get('export')
    if export == 'export':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-dataverifikasi.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Data Klaim Obat'

        # Define the titles for columns
        columns = [
            'namars',
            'status',
            'NoReg',
            'NoSEPApotek',
            'NoSEPAsalResep',
            'TglResep',
            'KdJenis',
            'NoKartu',
            'NamaPeserta',
            'ByTagApt',
            'ByVerApt',
            'verifikator',
            'rufil',
            'jenis_pending',
            'jenis_dispute',
            'ket_pending',
            'ket_jawaban',
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all
        for queryset in queryset:
            row_num += 1

            if queryset.ket_pending_dispute.last() is None:
                ket_pending_disput_queryset = ''
            else:
                ket_pending_disput_queryset = '{0}, '.format(queryset.ket_pending_dispute.last())

            if queryset.ket_jawaban_pending.last() is None:
                ket_jawaban_pending_queryset = ''
            else:
                ket_jawaban_pending_queryset = '{0}, '.format(queryset.ket_jawaban_pending.last())

            # ket_pending_disput_queryset = ''
            # for x in queryset.ket_pending_dispute.all():
            #     ket_pending_disput_queryset += '{0}, '.format(x.ket_pending_dispute)
            #
            # ket_jawaban_pending_queryset = ''
            # for x in queryset.ket_jawaban_pending.all():
            #     ket_jawaban_pending_queryset += '{0}, '.format(x.ket_jawaban_pending)

            # Define the data for each cell in the row
            row = [
                queryset.faskes.nama,
                queryset.status,
                queryset.register_klaim.nomor_register_klaim,
                queryset.NoSEPApotek,
                queryset.NoSEPAsalResep,
                queryset.TglResep,
                queryset.KdJenis,
                queryset.NoKartu,
                queryset.NamaPeserta,
                queryset.ByTagApt,
                queryset.ByVerApt,
                queryset.verifikator.username,
                queryset.rufil,
                queryset.jenis_pending,
                queryset.jenis_dispute,
                ket_pending_disput_queryset,
                ket_jawaban_pending_queryset,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response

    if request.POST.get('import'):

        # ambil file excel dan buat menjadi dataframe
        try:
            file = request.FILES['excel']
            if file.name.split('.')[-1] != 'xlsx':
                raise ValidationError('Hanya menerima file dengan ekstensi `.xlsx`')
        except Exception as e:
            messages.warning(request, f'Terdapat Error dalam proses upload dengan keterangan: {e}')
            return redirect(request.headers.get('Referer'))

        df_raw = pd.read_excel(file)

        # replace nan dengan kosong
        df = df_raw.replace(np.nan, '')

        # ubah nama kolom menjadi huruf besar
        df['status'] = df['status'].str.title()

        # ubah nama jenis pending menjadi huruf besar
        df['jenis_pending'] = df['jenis_pending'].str.title()

        # ubah nama jenis dispute menjadi huruf besar
        df['jenis_dispute'] = df['jenis_dispute'].str.title()

        # Call the Student Resource Model and make its instance
        data_claim_resource = DataKlaimObatResource()

        # Load the pandas dataframe into a tablib dataset
        dataset = Dataset().load(df)

        # cek column sudah sesuai
        list_kolom_mandatory = ['status', 'NoSEPApotek', 'jenis_pending', 'jenis_dispute', 'ket_pending']
        list_status_df = df.columns.tolist()
        for daftar in list_kolom_mandatory:
            if daftar not in list_status_df:
                messages.warning(request, f'File yang diimport harus memiliki kolom {list_kolom_mandatory}')
                return redirect(request.headers.get('Referer'))

        # cek sep tersebut berstatus bukan proses
        for status in dataset['status']:
            if status == 'Proses':
                messages.warning(request, 'File yang diimport harus berstatus "Layak", "Pending, '
                                          '"Dispute", "Tidak Layak"')
                return redirect(request.headers.get('Referer'))

        # cek status klaim
        list_status_mandatory = [StatusDataKlaimChoices.LAYAK, StatusDataKlaimChoices.PENDING,
                                 StatusDataKlaimChoices.TIDAK_LAYAK, StatusDataKlaimChoices.DISPUTE]
        list_status_df = list(dict.fromkeys(dataset['status']))
        for status in list_status_df:
            if status not in list_status_mandatory:
                messages.warning(request, f'File yang diimport harus memiliki status yang sesuai. '
                                          f'Mohon dapat dicek kembali.')
                return redirect(request.headers.get('Referer'))

        # cek jenis pending
        list_jenis_mandatory = [JenisPendingChoices.ADMINISTRASI, JenisPendingChoices.KODING,
                                JenisPendingChoices.STANDAR_PELAYANAN]
        list_jenis_pending_df = list(dict.fromkeys(dataset['jenis_pending']))
        for i in list_jenis_pending_df:
            if i == '':
                list_jenis_pending_df.remove('')

        for jenis in list_jenis_pending_df:
            if jenis not in list_jenis_mandatory:
                messages.warning(request, f'File yang diimport harus memiliki jenis pending yang sesuai. '
                                          f'Mohon dapat dicek kembali.')
                return redirect(request.headers.get('Referer'))

        # cek jenis dispute
        list_jenis_dispute_mandatory = [JenisDisputeChoices.MEDIS, JenisDisputeChoices.KODING, JenisDisputeChoices.COB]
        list_jenis_dispute_df = list(dict.fromkeys(dataset['jenis_dispute']))
        for i in list_jenis_dispute_df:
            if i == '':
                list_jenis_dispute_df.remove('')

        for jenis in list_jenis_dispute_df:
            if jenis not in list_jenis_dispute_mandatory:
                messages.warning(request, f'File yang diimport harus memiliki jenis dispute yang sesuai. '
                                          f'Mohon dapat dicek kembali.')
                return redirect(request.headers.get('Referer'))

        # cek jumlah jenis pending dan jenis dispute sama dengan yang distatus
        jumlah_pending = df['status'].eq(StatusDataKlaimChoices.PENDING).sum()
        jumlah_dispute = df['status'].eq(StatusDataKlaimChoices.DISPUTE).sum()
        jumlah_tidak_layak = df['status'].eq(StatusDataKlaimChoices.TIDAK_LAYAK).sum()
        jumlah_jenis_pending = df['jenis_pending'].apply(lambda x: isinstance(x, str) and len(x) > 0).sum().sum()
        jumlah_jenis_dispute = df['jenis_dispute'].apply(lambda x: isinstance(x, str) and len(x) > 0).sum().sum()
        jumlah_ket_pending = df['ket_pending'].apply(lambda x: isinstance(x, str) and len(x) > 0).sum().sum()
        total_pending_dispute = jumlah_pending + jumlah_dispute
        total_pending_dispute_tidak_layak = jumlah_pending + jumlah_dispute + jumlah_tidak_layak

        # cek jenis pending
        if total_pending_dispute > jumlah_jenis_pending:
            messages.warning(request, f'File yang diimport untuk status "Pending" dan "Dispute" '
                                      f'harus diisi jenis pending')
            return redirect(request.headers.get('Referer'))

        # cek jenis dispute
        elif jumlah_dispute > jumlah_jenis_dispute:
            messages.warning(request, f'File yang diimport untuk status "Dispute" harus diisi jenis dispute')
            return redirect(request.headers.get('Referer'))

        # cek ket pending
        elif total_pending_dispute_tidak_layak > jumlah_ket_pending:
            messages.warning(request, f'File yang diimport untuk status "Pending" dan "Dispute" belum diisi '
                                      f'keterangan pending')
            return redirect(request.headers.get('Referer'))

        # cek sep tersebut memang milik verifikator dan yang diimport adalah status belum diverif
        for i in dataset['NoSEPApotek']:
            queryset = DataKlaimObat.objects.filter(verifikator=request.user, prosesklaim=False, NoSEPApotek=i)
            if not queryset.exists():
                messages.warning(request, f'Terdapat NO SEP Apotek {i} yang tidak terdapat dalam verifikasi klaim '
                                          f'verifikator bersangkutan.')
                return redirect(request.headers.get('Referer'))
            else:
                for a in queryset:
                    if a.status != 'Proses':
                        messages.warning(request, f'Terdapat NO SEP Apotek {a.NoSEPApotek} yang sudah di verifikasi. '
                                                  f'Status yang diimport tidak boleh dalam keadaan telah diverifikasi.')
                        return redirect(request.headers.get('Referer'))

        try:
            with transaction.atomic():
                dataset_new = Dataset().load(df)
                list_id = []
                ket_pending_dispute_excel = dataset['ket_pending']
                for i in ket_pending_dispute_excel:
                    obj_ket_pending_dispute = KeteranganPendingDispute(ket_pending_dispute=i, verifikator=request.user)
                    obj_ket_pending_dispute.save()
                    list_id.append(obj_ket_pending_dispute.id)

                df['ket_pending_dispute'] = list_id

                for index, row in df.iterrows():
                    obj = DataKlaimObat.objects.filter(verifikator=request.user, prosesklaim=False,
                                                       NoSEPApotek=row['NoSEPApotek']).first()
                    obj.ket_pending_dispute.add(row['ket_pending_dispute'])
                data_claim_resource.import_data(dataset_new, dry_run=False)

                # delete yang tidak penting
                ket_pending_kosong = KeteranganPendingDispute.objects.filter(ket_pending_dispute='',
                                                                             verifikator=request.user)
                ket_pending_kosong.delete()
                messages.success(request, 'Data berhasil diimport. {0}'.format(dataset_new))
                return redirect(request.headers.get('Referer'))
        except Exception as e:
            # Nampilih error
            messages.info(request, f'Terjadi kesalahan: {str(e)}')

    # pagination
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'data_klaim': queryset,
        'myFilter': myFilter,
    }

    return render(request, 'verifikator/obat/daftar_data_klaim_obat.html', context)


@login_required
@check_device
@permissions(role=['verifikator'])
def detail_data_klaim_obat(request, pk):
    queryset = DataKlaimObat.objects.filter(verifikator=request.user)
    instance = queryset.get(pk=pk)
    data_klaim_form = DataKlaimVerifikatorForm(instance=instance)
    data_klaim_pending_form = KeteranganPendingForm()

    if request.method == 'POST':
        data_klaim_form = DataKlaimVerifikatorForm(instance=instance, data=request.POST)
        keterangan = KeteranganPendingForm(request.POST or None)
        if data_klaim_form.is_valid():
            if instance.is_hitung is False:
                HitungDataKlaim.objects.create(nomor_register_klaim=instance.register_klaim.nomor_register_klaim,
                                               jenis_klaim=instance.register_klaim.jenis_klaim,
                                               verifikator=instance.verifikator)
                instance.is_hitung = True
                instance.save()
            if instance.status != StatusDataKlaimChoices.LAYAK:
                obj_keterangan = keterangan.save()
                obj_keterangan.verifikator = instance.verifikator
                obj_keterangan.save()
                instance.ket_pending_dispute.add(obj_keterangan)
            data_klaim_form.save()
            next = request.POST.get('next', '/')
            messages.success(request, f'NO SEP {instance.NoSEPApotek} berhasil diupdate.')
            return HttpResponseRedirect(next)

    context = {
        'data_klaim': instance,
        'data_klaim_form': data_klaim_form,
        'data_klaim_pending_form': data_klaim_pending_form,
    }
    return render(request, 'verifikator/obat/detail_data_klaim_obat.html', context)


@login_required
@check_device
@permissions(role=['verifikator'])
def download_data_obat(request):
    # initial relasi pada kantor cabang
    related_kantor_cabang = request.user.kantorcabang_set.all()
    queryset = DataKlaimObat.objects.filter(verifikator__kantorcabang__in=related_kantor_cabang)

    # filter
    myFilter = DownloadDataKlaimObatFilter(request.GET, queryset=queryset, request=request)
    queryset = myFilter.qs

    kantor_cabang = request.user.kantorcabang_set.all().first()

    # fitur download
    download = request.GET.get('download')
    bupel_month = request.GET.get('bupel_month')
    bupel_year = request.GET.get('bupel_year')
    faskes = request.GET.get('faskes')
    nomor_register = request.GET.get('nomor_register_klaim')
    try:
        if nomor_register != '':
            if download == 'download':
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                response['Content-Disposition'] = 'attachment; filename={date}-dataverifikasi.xlsx'.format(
                    date=datetime.datetime.now().strftime('%Y-%m-%d'),
                )
                workbook = Workbook()

                # Get active worksheet/tab
                worksheet = workbook.active
                worksheet.title = 'Data Klaim Obat'

                # Define the titles for columns
                columns = [
                    'namars',
                    'status',
                    'NoReg',
                    'NoSEPApotek',
                    'NoSEPAsalResep',
                    'TglResep',
                    'KdJenis',
                    'NoKartu',
                    'NamaPeserta',
                    'ByTagApt',
                    'ByVerApt',
                    'verifikator',
                    'rufil',
                    'jenis_pending',
                    'jenis_dispute',
                    'ket_pending',
                    'ket_jawaban',
                ]
                row_num = 1

                # Assign the titles for each cell of the header
                for col_num, column_title in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title

                # Iterate through all movies
                for queryset in queryset:
                    row_num += 1

                    if queryset.ket_pending_dispute.last() is None:
                        ket_pending_disput_queryset = ''
                    else:
                        ket_pending_disput_queryset = '{0}, '.format(queryset.ket_pending_dispute.last())

                    if queryset.ket_jawaban_pending.last() is None:
                        ket_jawaban_pending_queryset = ''
                    else:
                        ket_jawaban_pending_queryset = '{0}, '.format(queryset.ket_jawaban_pending.last())

                    # ket_pending_disput_queryset = ''
                    # for x in queryset.ket_pending_dispute.all():
                    #     ket_pending_disput_queryset += '{0}, '.format(x.ket_pending_dispute)
                    #
                    # ket_jawaban_pending_queryset = ''
                    # for x in queryset.ket_jawaban_pending.all():
                    #     ket_jawaban_pending_queryset += '{0}, '.format(x.ket_jawaban_pending)

                    # Define the data for each cell in the row
                    row = [
                        queryset.faskes.nama,
                        queryset.status,
                        queryset.register_klaim.nomor_register_klaim,
                        queryset.NoSEPApotek,
                        queryset.NoSEPAsalResep,
                        queryset.TglResep,
                        queryset.KdJenis,
                        queryset.NoKartu,
                        queryset.NamaPeserta,
                        queryset.ByTagApt,
                        queryset.ByVerApt,
                        queryset.verifikator.username,
                        queryset.rufil,
                        queryset.jenis_pending,
                        queryset.jenis_dispute,
                        ket_pending_disput_queryset,
                        ket_jawaban_pending_queryset,
                    ]

                    # Assign the data for each cell of the row
                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value

                workbook.save(response)
                return response
        elif bupel_month != '' and bupel_year != '' and faskes != '':
            if download == 'download':
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                response['Content-Disposition'] = 'attachment; filename={date}-dataverifikasi.xlsx'.format(
                    date=datetime.datetime.now().strftime('%Y-%m-%d'),
                )
                workbook = Workbook()

                # Get active worksheet/tab
                worksheet = workbook.active
                worksheet.title = 'Data Klaim Obat'

                # Define the titles for columns
                columns = [
                    'namars',
                    'status',
                    'NoReg',
                    'NoSEPApotek',
                    'NoSEPAsalResep',
                    'TglResep',
                    'KdJenis',
                    'NoKartu',
                    'NamaPeserta',
                    'ByTagApt',
                    'ByVerApt',
                    'verifikator',
                    'rufil',
                    'jenis_pending',
                    'jenis_dispute',
                    'ket_pending',
                    'ket_jawaban',
                ]
                row_num = 1

                # Assign the titles for each cell of the header
                for col_num, column_title in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title

                # Iterate through all movies
                for queryset in queryset:
                    row_num += 1

                    if queryset.ket_pending_dispute.last() is None:
                        ket_pending_disput_queryset = ''
                    else:
                        ket_pending_disput_queryset = '{0}, '.format(queryset.ket_pending_dispute.last())

                    if queryset.ket_jawaban_pending.last() is None:
                        ket_jawaban_pending_queryset = ''
                    else:
                        ket_jawaban_pending_queryset = '{0}, '.format(queryset.ket_jawaban_pending.last())

                    # ket_pending_disput_queryset = ''
                    # for x in queryset.ket_pending_dispute.all():
                    #     ket_pending_disput_queryset += '{0}, '.format(x.ket_pending_dispute)
                    #
                    # ket_jawaban_pending_queryset = ''
                    # for x in queryset.ket_jawaban_pending.all():
                    #     ket_jawaban_pending_queryset += '{0}, '.format(x.ket_jawaban_pending)

                    # Define the data for each cell in the row
                    row = [
                        queryset.faskes.nama,
                        queryset.status,
                        queryset.register_klaim.nomor_register_klaim,
                        queryset.NoSEPApotek,
                        queryset.NoSEPAsalResep,
                        queryset.TglResep,
                        queryset.KdJenis,
                        queryset.NoKartu,
                        queryset.NamaPeserta,
                        queryset.ByTagApt,
                        queryset.ByVerApt,
                        queryset.verifikator.username,
                        queryset.rufil,
                        queryset.jenis_pending,
                        queryset.jenis_dispute,
                        ket_pending_disput_queryset,
                        ket_jawaban_pending_queryset,
                    ]

                    # Assign the data for each cell of the row
                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value

                workbook.save(response)
                return response
        else:
            messages.warning(request, 'Bulan, Tahun, dan Faskes harus diisi atau Nomor Register Klaim harus isi!')
    except Exception as e:
        messages.warning(request, "Terjadi Kesalahan Dalam Download Data, dengan Keterangan: " + str(e))

    context = {
        'myFilter': myFilter,
    }
    return render(request, 'verifikator/obat/download_data_obat.html', context)


@csrf_protect
@require_POST
@login_required
@check_device
@permissions(role=['verifikator'])
def simpan_status_fragmentasi(request):
    try:
        data = json.loads(request.body)
        no_sep = data.get('no_sep')
        status = data.get('status')  # Bisa 'Layak', 'Pending', 'Tidak Layak', 'Dispute' atau null

        # Validasi data yang diterima
        if not no_sep:
            return JsonResponse({'error': 'Parameter no_sep diperlukan.'}, status=400)

        # Validasi status jika diberikan
        valid_status = ["Layak", "Pending", "Tidak Layak", "Dispute"]
        if status and status not in valid_status:
            return JsonResponse({'error': 'Status tidak valid.'}, status=400)

        # Cari objek DataKlaimCBG berdasarkan NOSEP
        try:
            data_klaim = DataKlaimCBG.objects.get(NOSEP=no_sep)
        except DataKlaimCBG.DoesNotExist:
            return JsonResponse({'error': 'Data Klaim tidak ditemukan.'}, status=404)

        # Update status jika diberikan
        if status:
            data_klaim.status = status
        else:
            data_klaim.status = ''  # Atur sesuai kebutuhan

        if status in ['Pending', 'Dispute', 'Tidak Layak']:
            # Tentukan jenis_pending dan jenis_dispute jika perlu
            data_klaim.jenis_pending = JenisPendingChoices.ADMINISTRASI
            if status == 'Dispute':
                data_klaim.jenis_dispute = JenisDisputeChoices.KODING

            # Cek apakah KeteranganPendingDispute sudah ada
            ket_pending_dispute_text = 'Potensi Fragmentasi'
            keterangan_pending_dispute = KeteranganPendingDispute.objects.create(
                ket_pending_dispute=ket_pending_dispute_text,
                verifikator=request.user
            )
            data_klaim.ket_pending_dispute.add(keterangan_pending_dispute)

        data_klaim.save()

        return JsonResponse({'message': 'Status berhasil disimpan.'}, status=200)

    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON tidak valid.'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@check_device
@permissions(role=['verifikator'])
def detail_data_klaim_verifkhusus(request, no_sep):
    # Menggunakan get_object_or_404 untuk mencari instance berdasarkan NOSEP dan verifikator
    instance = get_object_or_404(DataKlaimCBG, verifikator=request.user, NOSEP=no_sep)

    # Inisialisasi form dengan instance
    data_klaim_form = DataKlaimVerifikatorForm(instance=instance)
    data_klaim_pending_form = KeteranganPendingForm()

    # Mendapatkan data metafisik terkait NOSEP
    data_klaim_cbg_metafisik = DataKlaimCBGMetafisik.objects.filter(nosjp=no_sep).first()

    context = {
        'data_klaim': instance,
        'data_klaim_form': data_klaim_form,
        'data_klaim_pending_form': data_klaim_pending_form,
        'meta': data_klaim_cbg_metafisik
    }
    return render(request, 'verifikator/cbg/detail_data_klaim_verifkhusus.html', context)


@csrf_protect
@require_POST
@login_required
@check_device
@permissions(role=['verifikator'])
def simpan_status_readmisi(request):
    try:
        data = json.loads(request.body)
        no_sep = data.get('no_sep')
        status = data.get('status')

        if not no_sep:
            return JsonResponse({'error': 'Parameter no_sep diperlukan.'}, status=400)

        valid_status = ["Layak", "Pending", "Tidak Layak", "Dispute"]
        if status not in valid_status:
            return JsonResponse({'error': 'Status tidak valid.'}, status=400)

        try:
            data_klaim = DataKlaimCBG.objects.get(NOSEP=no_sep, verifikator=request.user)
        except DataKlaimCBG.DoesNotExist:
            return JsonResponse({'error': 'Data Klaim tidak ditemukan.'}, status=404)

        data_klaim.status = status

        if status in ['Pending', 'Dispute', 'Tidak Layak']:
            data_klaim.jenis_pending = JenisPendingChoices.ADMINISTRASI
            if status == 'Dispute':
                data_klaim.jenis_dispute = JenisDisputeChoices.KODING

            ket_pending_dispute_text = 'Potensi Readmisi'
            keterangan_pending_dispute = KeteranganPendingDispute.objects.create(
                ket_pending_dispute=ket_pending_dispute_text,
                verifikator=request.user
            )
            data_klaim.ket_pending_dispute.add(keterangan_pending_dispute)

        data_klaim.save()

        return JsonResponse({'message': 'Status berhasil disimpan.'}, status=200)

    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON tidak valid.'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@csrf_protect
@require_POST
@login_required
@check_device
@permissions(role=['verifikator'])
def simpan_status_rehabilitasi(request):
    try:
        data = json.loads(request.body)
        no_sep = data.get('no_sep')
        status = data.get('status')

        if not no_sep:
            return JsonResponse({'error': 'Parameter no_sep diperlukan.'}, status=400)

        valid_status = ["Layak", "Pending", "Tidak Layak", "Dispute"]
        if status not in valid_status:
            return JsonResponse({'error': 'Status tidak valid.'}, status=400)

        try:
            data_klaim = DataKlaimCBG.objects.get(NOSEP=no_sep, verifikator=request.user)
        except DataKlaimCBG.DoesNotExist:
            return JsonResponse({'error': 'Data Klaim tidak ditemukan.'}, status=404)

        data_klaim.status = status

        if status in ['Pending', 'Dispute', 'Tidak Layak']:
            data_klaim.jenis_pending = JenisPendingChoices.ADMINISTRASI
            if status == 'Dispute':
                data_klaim.jenis_dispute = JenisDisputeChoices.KODING

            ket_pending_dispute_text = 'Ketidaksesuaian Pengajuan Klaim Rehabilitasi Medis'
            keterangan_pending_dispute = KeteranganPendingDispute.objects.create(
                ket_pending_dispute=ket_pending_dispute_text,
                verifikator=request.user
            )
            data_klaim.ket_pending_dispute.add(keterangan_pending_dispute)

        data_klaim.save()

        return JsonResponse({'message': 'Status berhasil disimpan.'}, status=200)

    except json.JSONDecodeError:
        return JsonResponse({'error': 'JSON tidak valid.'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@check_device
@permissions(role=['verifikator'])
def fragmentasi(request):
    queryset = DataKlaimCBG.objects.filter(
        verifikator=request.user,
        prosesklaim=False,
        JNSPEL=JenisPelayananChoices.RAWAT_JALAN
    ).order_by('NMPESERTA', 'TGLSEP')

    # Terapkan filter dengan menambahkan parameter 'user'
    myFilter = FragmentasiFilter(request.GET, queryset=queryset, user=request.user)
    queryset = myFilter.qs

    # Kelompokkan data berdasarkan kombinasi unik NOKARTU dan NMPESERTA
    grouped_queryset = queryset.values(
        'NOKARTU',
        'NMPESERTA',
        'register_klaim__nomor_register_klaim',
        'faskes__nama'
    ).annotate(
        jumlah_kunjungan=Count('NOSEP'),
        jumlah_biaya=Sum('BYPENGAJUAN'),
    ).filter(jumlah_kunjungan__gt=1).order_by('-jumlah_kunjungan')

    # Pagination: ubah jumlah item per halaman menjadi 10
    paginator = Paginator(grouped_queryset, 10)
    page_number = request.GET.get('page')
    grouped_queryset = paginator.get_page(page_number)

    context = {
        'data_klaim': grouped_queryset,
        'myFilter': myFilter,
    }
    return render(request, 'verifikator/cbg/fragmentasi.html', context)


@login_required
@check_device
@permissions(role=['verifikator'])
def readmisi(request):
    queryset = DataKlaimCBG.objects.filter(
        verifikator=request.user,
        prosesklaim=False,
        JNSPEL=JenisPelayananChoices.RAWAT_INAP
    ).order_by('NMPESERTA', 'TGLSEP')

    # Terapkan filter dengan menambahkan parameter 'user'
    myFilter = FragmentasiFilter(request.GET, queryset=queryset, user=request.user)
    queryset = myFilter.qs

    # Kelompokkan data berdasarkan kombinasi unik NOKARTU dan NMPESERTA
    grouped_queryset = queryset.values(
        'NOKARTU',
        'NMPESERTA',
        'register_klaim__nomor_register_klaim',
        'faskes__nama'
    ).annotate(
        jumlah_kunjungan=Count('NOSEP'),
        jumlah_biaya=Sum('BYPENGAJUAN'),
    ).filter(jumlah_kunjungan__gt=1).order_by('-jumlah_kunjungan')

    # Pagination: ubah jumlah item per halaman menjadi 10
    paginator = Paginator(grouped_queryset, 10)
    page_number = request.GET.get('page')
    grouped_queryset = paginator.get_page(page_number)

    context = {
        'data_klaim': grouped_queryset,
        'myFilter': myFilter,
    }
    return render(request, 'verifikator/cbg/readmisi.html', context)


@login_required
@check_device
@permissions(role=['verifikator'])
def rehabilitasi(request):
    queryset = DataKlaimCBG.objects.filter(
        verifikator=request.user,
        prosesklaim=False
    ).filter(
        Q(POLI='IRM') | Q(KDINACBG__in=['Z-3-12-0', 'M-3-16-0'])
    ).order_by('NMPESERTA', 'TGLSEP')

    # Terapkan filter dengan menambahkan parameter 'user'
    myFilter = FragmentasiFilter(request.GET, queryset=queryset, user=request.user)
    queryset = myFilter.qs

    # Kelompokkan data berdasarkan kombinasi unik NOKARTU dan NMPESERTA
    grouped_queryset = queryset.values(
        'NOKARTU',
        'NMPESERTA',
        'register_klaim__nomor_register_klaim',
        'faskes__nama'
    ).annotate(
        jumlah_kunjungan=Count('NOSEP'),
        jumlah_biaya=Sum('BYPENGAJUAN'),
    ).filter(jumlah_kunjungan__gt=1).order_by('-jumlah_kunjungan')

    # Pagination: ubah jumlah item per halaman menjadi 10
    paginator = Paginator(grouped_queryset, 10)
    page_number = request.GET.get('page')
    grouped_queryset = paginator.get_page(page_number)

    context = {
        'data_klaim': grouped_queryset,
        'myFilter': myFilter,
    }
    return render(request, 'verifikator/cbg/rehabilitasi.html', context)


@login_required
@check_device
@permissions(role=['verifikator'])
@require_GET
def get_fragmentasi_detail(request):
    nokartu = request.GET.get('nokartu')
    nmpeserta = request.GET.get('nmpeserta')
    page = request.GET.get('page', 1)
    items_per_page = 26

    if not nokartu or not nmpeserta:
        return JsonResponse({'error': 'Parameter nokartu dan nmpeserta diperlukan.'}, status=400)

    queryset = DataKlaimCBG.objects.filter(
        verifikator=request.user,
        prosesklaim=False,
        JNSPEL=JenisPelayananChoices.RAWAT_JALAN,
        NOKARTU=nokartu,
        NMPESERTA=nmpeserta
    ).order_by('TGLSEP')

    paginator = Paginator(queryset, items_per_page)
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        return JsonResponse({'error': 'Halaman tidak valid.'}, status=400)

    detail_data = [{
        "rs": data.faskes.nama,
        "no_sep": data.NOSEP,
        "in": data.TGLSEP.strftime('%d-%m-%Y') if data.TGLSEP else '',
        "out": data.TGLPULANG.strftime('%d-%m-%Y') if data.TGLPULANG else '',
        "poli": data.POLI,
        "cbg": data.KDINACBG,
        "tarif": data.BYPENGAJUAN,
        "current_status": data.status,
        "status_options": ["Layak", "Pending", "Tidak Layak", "Dispute"]
    } for data in page_obj]

    return JsonResponse({
        'detail_data': detail_data,
        'has_next': page_obj.has_next(),
        'has_previous': page_obj.has_previous(),
        'num_pages': paginator.num_pages,
        'current_page': page_obj.number,
    })


@login_required
@check_device
@permissions(role=['verifikator'])
@require_GET
def get_readmisi_detail(request):
    nokartu = request.GET.get('nokartu')
    nmpeserta = request.GET.get('nmpeserta')
    page = request.GET.get('page', 1)
    items_per_page = 26

    if not nokartu or not nmpeserta:
        return JsonResponse({'error': 'Parameter nokartu dan nmpeserta diperlukan.'}, status=400)

    queryset = DataKlaimCBG.objects.filter(
        verifikator=request.user,
        prosesklaim=False,
        JNSPEL=JenisPelayananChoices.RAWAT_INAP,
        NOKARTU=nokartu,
        NMPESERTA=nmpeserta
    ).order_by('TGLSEP')

    paginator = Paginator(queryset, items_per_page)
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        return JsonResponse({'error': 'Halaman tidak valid.'}, status=400)

    detail_data = [{
        "rs": data.faskes.nama,
        "no_sep": data.NOSEP,
        "in": data.TGLSEP.strftime('%d-%m-%Y') if data.TGLSEP else '',
        "out": data.TGLPULANG.strftime('%d-%m-%Y') if data.TGLPULANG else '',
        "poli": data.POLI,
        "cbg": data.KDINACBG,
        "tarif": data.BYPENGAJUAN,
        "current_status": data.status,
        "status_options": ["Layak", "Pending", "Tidak Layak", "Dispute"]
    } for data in page_obj]

    return JsonResponse({
        'detail_data': detail_data,
        'has_next': page_obj.has_next(),
        'has_previous': page_obj.has_previous(),
        'num_pages': paginator.num_pages,
        'current_page': page_obj.number,
    })


@login_required
@check_device
@permissions(role=['verifikator'])
@require_GET
def get_rehabilitasi_detail(request):
    nokartu = request.GET.get('nokartu')
    nmpeserta = request.GET.get('nmpeserta')
    page = request.GET.get('page', 1)
    items_per_page = 26

    if not nokartu or not nmpeserta:
        return JsonResponse({'error': 'Parameter nokartu dan nmpeserta diperlukan.'}, status=400)

    queryset = DataKlaimCBG.objects.filter(
        verifikator=request.user,
        prosesklaim=False,
        NOKARTU=nokartu,
        NMPESERTA=nmpeserta
    ).filter(
        Q(POLI='IRM') | Q(KDINACBG__in=['Z-3-12-0', 'M-3-16-0'])
    ).order_by('TGLSEP')

    paginator = Paginator(queryset, items_per_page)
    try:
        page_obj = paginator.page(page)
    except (PageNotAnInteger, EmptyPage):
        return JsonResponse({'error': 'Halaman tidak valid.'}, status=400)

    detail_data = [{
        "rs": data.faskes.nama,
        "no_sep": data.NOSEP,
        "in": data.TGLSEP.strftime('%d-%m-%Y') if data.TGLSEP else '',
        "out": data.TGLPULANG.strftime('%d-%m-%Y') if data.TGLPULANG else '',
        "poli": data.POLI,
        "cbg": data.KDINACBG,
        "tarif": data.BYPENGAJUAN,
        "current_status": data.status,
        "status_options": ["Layak", "Pending", "Tidak Layak", "Dispute"]
    } for data in page_obj]

    return JsonResponse({
        'detail_data': detail_data,
        'has_next': page_obj.has_next(),
        'has_previous': page_obj.has_previous(),
        'num_pages': paginator.num_pages,
        'current_page': page_obj.number,
    })