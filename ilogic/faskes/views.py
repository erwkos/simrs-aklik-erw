import calendar
import datetime

from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from openpyxl.workbook import Workbook

from user.decorators import permissions, check_device
from klaim.filters import RegisterKlaimFaskesFilter, RegisterKlaimKhususFaskesFilter
from verifikator.filters import DataKlaimCBGFilter, DataKlaimCBGFaskesFilter, DataKlaimObatFaskesFilter
from .models import (
    Kepwil,
    KantorCabang,
    Faskes
)
from klaim.models import (
    RegisterKlaim, JenisKlaim, DataKlaimCBG, DataKlaimObat
)
from klaim.choices import (
    StatusRegisterChoices, NamaJenisKlaimChoices, StatusDataKlaimChoices
)
from .forms import (
    RegisterKlaimForm,
    UpdateRegisterKlaimForm, DataKlaimCBGFaskesForm, JawabanPendingDisputeForm, UpdateRegisterKlaimDisableForm,
    DataKlaimObatFaskesForm
)


@login_required
@check_device
@permissions(role=['faskes'])
def register(request):
    user = request.user
    today = datetime.date.today()
    if request.method == 'POST':
        form = RegisterKlaimForm(data=request.POST)
        if form.is_valid():
            jenis_klaim = JenisKlaim.objects.filter(nama=form.cleaned_data.get('jenis_klaim'))[0]
            if form.cleaned_data.get('bulan_pelayanan').year >= today.year and form.cleaned_data.get('bulan_pelayanan').month >= today.month:
                messages.warning(request, f"Pengajuan {form.cleaned_data.get('jenis_klaim')} dengan bulan pelayanan "
                                          f"{calendar.month_name[form.cleaned_data.get('bulan_pelayanan').month]} "
                                          f"{form.cleaned_data.get('bulan_pelayanan').year} "
                                          f"masih belum tersedia! Mohon cek kembali bulan pelayanan yang diajukan. Terima Kasih.")
                return render(request, 'faskes/input.html', {'form': RegisterKlaimForm()})
            elif jenis_klaim.nama == NamaJenisKlaimChoices.CBG_REGULER or jenis_klaim.nama == NamaJenisKlaimChoices.OBAT_REGULER:
                if RegisterKlaim.objects.filter(
                    jenis_klaim__nama=form.cleaned_data.get('jenis_klaim'),
                    bulan_pelayanan__year=form.cleaned_data.get('bulan_pelayanan').year,
                    bulan_pelayanan__month=form.cleaned_data.get('bulan_pelayanan').month,
                    faskes=request.user.faskes_set.all().first()
                ).exists(): # and form.cleaned_data.get('is_pengajuan_ulang') is False: # untuk klaim reguler hanya bisa diajukan 1x per rs
                    messages.warning(request, f"{form.cleaned_data.get('jenis_klaim')} dengan bulan pelayanan "
                                              f"{calendar.month_name[form.cleaned_data.get('bulan_pelayanan').month]} "
                                              f"{form.cleaned_data.get('bulan_pelayanan').year} "
                                              f"sudah pernah diajukan sebelumnya! Terima Kasih.")
                    return render(request, 'faskes/input.html', {'form': RegisterKlaimForm()})
            elif RegisterKlaim.objects.filter(
                    jenis_klaim__nama=form.cleaned_data.get('jenis_klaim'),
                    bulan_pelayanan__year=form.cleaned_data.get('bulan_pelayanan').year,
                    bulan_pelayanan__month=form.cleaned_data.get('bulan_pelayanan').month,
                    faskes=request.user.faskes_set.all().first()).exclude(status=StatusRegisterChoices.SELESAI).exists():
                    messages.warning(request, f"{form.cleaned_data.get('jenis_klaim')} dengan bulan pelayanan "
                                              f"{calendar.month_name[form.cleaned_data.get('bulan_pelayanan').month]} "
                                              f"{form.cleaned_data.get('bulan_pelayanan').year} "
                                              f"sedang tahap pengajuan/verifikasi! Terima Kasih.")
                    return render(request, 'faskes/input.html', {'form': RegisterKlaimForm()})


                # return render(request, 'faskes/input.html', {'form': RegisterKlaimForm()})

            # if jenis_klaim.nama == NamaJenisKlaimChoices.CBG_SUSULAN or jenis_klaim.nama == NamaJenisKlaimChoices.OBAT_SUSULAN:
            #     DataKlaimCBG.objects.filter(bupel__month=form.cleaned_data.get('bulan_pelayanan').month,
            #                                 bupel__year=form.cleaned_data.get('bulan_pelayanan').year,
            #                                 status=StatusDataKlaimChoices.PEMBAHASAN).\
            #         update(prosesklaim=False)

            messages.success(request, f"{form.cleaned_data.get('jenis_klaim')} dengan bulan pelayanan "
                                      f"{calendar.month_name[form.cleaned_data.get('bulan_pelayanan').month]} "
                                      f"{form.cleaned_data.get('bulan_pelayanan').year} "
                                      f"berhasil diajukan! Terima Kasih.")
            data = form.cleaned_data
            data['faskes'] = user.faskes_set.first()
            data['nomor_register_klaim'] = RegisterKlaim.generate_kode_register(faskes=user.faskes_set.all().first())
            data['tgl_aju'] = datetime.datetime.now()
            RegisterKlaim.objects.create(**data)
    return render(request, 'faskes/input.html', {'form': RegisterKlaimForm()})


@login_required
@check_device
@permissions(role=['faskes'])
def daftar_register(request):
    queryset = RegisterKlaim.objects.filter(faskes=request.user.faskes_set.all().first()).order_by('-tgl_aju')

    # filter
    myFilter = RegisterKlaimKhususFaskesFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    # pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'register_list': queryset,
        'myFilter': myFilter,
    }

    return render(request, 'faskes/daftar_register.html', context)


@login_required
@check_device
@permissions(role=['faskes'])
def detail_register(request, pk):
    queryset = RegisterKlaim.objects.filter(faskes=request.user.faskes_set.all().first())
    instance = queryset.get(pk=pk)
    form = UpdateRegisterKlaimForm(instance=instance)
    form_disable = UpdateRegisterKlaimDisableForm(instance=instance)
    if request.method == 'POST' and instance.status == StatusRegisterChoices.DIKEMBALIKAN:
        form = UpdateRegisterKlaimForm(instance=instance, data=request.POST)
        # if not form.has_changed():
        #     messages.warning(request, 'Tidak ada perubahan data')
        if form.is_valid():
            jenis_klaim = JenisKlaim.objects.filter(nama=form.cleaned_data.get('jenis_klaim'))[0]
            if jenis_klaim.nama == NamaJenisKlaimChoices.CBG_REGULER or jenis_klaim.nama == NamaJenisKlaimChoices.OBAT_REGULER:
                if RegisterKlaim.objects.filter(
                    jenis_klaim__nama=form.cleaned_data.get('jenis_klaim'),
                    bulan_pelayanan__year=form.cleaned_data.get('bulan_pelayanan').year,
                    bulan_pelayanan__month=form.cleaned_data.get('bulan_pelayanan').month,
                    faskes=request.user.faskes_set.all().first()
                ).exclude(status=StatusRegisterChoices.DIKEMBALIKAN).exists(): # and form.cleaned_data.get('is_pengajuan_ulang') is False: # untuk klaim reguler hanya bisa diajukan 1x per rs
                    messages.warning(request, f"{form.cleaned_data.get('jenis_klaim')} dengan bulan pelayanan "
                                              f"{calendar.month_name[form.cleaned_data.get('bulan_pelayanan').month]} "
                                              f"{form.cleaned_data.get('bulan_pelayanan').year} "
                                              f"sudah pernah diajukan sebelumnya! Terima Kasih.")
                    return redirect(request.headers.get('Referer'))
            elif RegisterKlaim.objects.filter(
                    jenis_klaim__nama=form.cleaned_data.get('jenis_klaim'),
                    bulan_pelayanan__year=form.cleaned_data.get('bulan_pelayanan').year,
                    bulan_pelayanan__month=form.cleaned_data.get('bulan_pelayanan').month,
                    faskes=request.user.faskes_set.all().first()).exclude(Q(status=StatusRegisterChoices.DIKEMBALIKAN) | Q(status=StatusRegisterChoices.SELESAI)).exists():
                messages.warning(request, f"{form.cleaned_data.get('jenis_klaim')} dengan bulan pelayanan "
                                          f"{calendar.month_name[form.cleaned_data.get('bulan_pelayanan').month]} "
                                          f"{form.cleaned_data.get('bulan_pelayanan').year} "
                                          f"sedang tahap pengajuan/verifikasi! Terima Kasih.")
                return redirect(request.headers.get('Referer'))

            form.save()
            instance.status = StatusRegisterChoices.PENGAJUAN

            # ketika diajukan maka reset tgl dengan tanggal saat ini
            instance.tgl_aju = datetime.datetime.now()

            instance.save()
            messages.success(request, 'Status Klaim berhasil diajukan kembali. Terima Kasih.')
            return redirect(request.headers.get('Referer'))

    context = {
        'register': instance,
        'form': form,
        'form_disable': form_disable,
    }
    return render(request, 'faskes/detail_register.html', context)


@login_required
@check_device
@permissions(role=['faskes'])
def daftar_data_klaim_pending_dispute_cbg(request):
    queryset = DataKlaimCBG.objects.filter(faskes=request.user.faskes_set.all().first(), prosesklaim=True,
                                           prosespending=True, prosestidaklayak=False).order_by('NMPESERTA', 'TGLSEP')

    # filter
    myFilter = DataKlaimCBGFaskesFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    # export excel
    export = request.GET.get('export')
    if export == 'export':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-dataverifikasi.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Data Klaim CBG'

        # Define the titles for columns
        columns = [
            'namars',
            'Status',
            'NOREG',
            'NOSEP',
            'TGLSEP',
            'TGLPULANG',
            'JNSPEL',
            'NOKARTU',
            'NMPESERTA',
            'POLI',
            'KDINACBG',
            'BYPENGAJUAN',
            'jenis_pending',
            'jenis_dispute',
            'ket_pending',
            'ket_jawaban',
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all movies
        for queryset in queryset:
            row_num += 1

            if queryset.ket_pending_dispute.last() is None:
                ket_pending_disput_queryset = ''
            else:
                ket_pending_disput_queryset = '{0}, '.format(queryset.ket_pending_dispute.last())

            if queryset.ket_jawaban_pending.last() is None:
                ket_jawaban_pending_queryset = ''
            else:
                ket_jawaban_pending_queryset = '{0}, '.format(queryset.ket_jawaban_pending.last())

            # ket_pending_disput_queryset = ''
            # for x in queryset.ket_pending_dispute.all():
            #     ket_pending_disput_queryset += '{0}, '.format(x.ket_pending_dispute)
            #
            # ket_jawaban_pending_queryset = ''
            # for x in queryset.ket_jawaban_pending.all():
            #     ket_jawaban_pending_queryset += '{0}, '.format(x.ket_jawaban_pending)

            # Define the data for each cell in the row
            row = [
                queryset.faskes.nama,
                queryset.status,
                queryset.register_klaim.nomor_register_klaim,
                queryset.NOSEP,
                queryset.TGLSEP,
                queryset.TGLPULANG,
                queryset.JNSPEL,
                queryset.NOKARTU,
                queryset.NMPESERTA,
                queryset.POLI,
                queryset.KDINACBG,
                queryset.BYPENGAJUAN,
                queryset.jenis_pending,
                queryset.jenis_dispute,
                ket_pending_disput_queryset,
                ket_jawaban_pending_queryset,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response

    # pagination
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'data_klaim': queryset,
        'myFilter': myFilter,
    }

    return render(request, 'faskes/daftar_data_klaim_pending_dispute_cbg.html', context)


@login_required
@check_device
@permissions(role=['faskes'])
def detail_data_klaim_pending_dispute_cbg(request, pk):
    queryset = DataKlaimCBG.objects.filter(faskes=request.user.faskes_set.all().first(), prosesklaim=True,
                                           prosespending=True, prosestidaklayak=False).order_by('NMPESERTA', 'TGLSEP')
    instance = queryset.get(pk=pk)
    data_klaim_form = DataKlaimCBGFaskesForm(instance=instance)
    jawaban_pending_dispute_form = JawabanPendingDisputeForm()

    if request.method == 'POST':
        data_klaim_form = DataKlaimCBGFaskesForm(instance=instance, data=request.POST)
        jawaban = JawabanPendingDisputeForm(request.POST or None)
        if data_klaim_form.is_valid():
            data_klaim_form.save()
            # if jawaban_pending_dispute_form.is_valid():
            obj_jawaban = jawaban.save()
            obj_jawaban.user_faskes = request.user
            obj_jawaban.save()
            instance.ket_jawaban_pending.add(obj_jawaban)
            next = request.POST.get('next', '/')
            return HttpResponseRedirect(next)
    context = {
        'data_klaim': instance,
        'data_klaim_form': data_klaim_form,
        'jawaban_pending_dispute_form': jawaban_pending_dispute_form,
    }
    return render(request, 'faskes/detail_data_klaim_pending_dispute_cbg.html', context)


@login_required
@check_device
@permissions(role=['faskes'])
def daftar_data_klaim_pending_dispute_obat(request):
    queryset = DataKlaimObat.objects.filter(faskes=request.user.faskes_set.all().first(), prosesklaim=True,
                                            prosespending=True, prosestidaklayak=False).order_by('NamaPeserta', 'TglResep')

    # filter
    myFilter = DataKlaimObatFaskesFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    # export excel
    export = request.GET.get('export')
    if export == 'export':
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-dataverifikasi.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Data Klaim Obat'

        # Define the titles for columns
        columns = [
            'namars',
            'status',
            'NoReg',
            'NoSEPApotek',
            'NoSEPAsalResep',
            'TglResep',
            'KdJenis',
            'NoKartu',
            'NamaPeserta',
            'ByTagApt',
            'ByVerApt',
            'verifikator',
            'rufil',
            'jenis_pending',
            'jenis_dispute',
            'ket_pending',
            'ket_jawaban',
        ]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        # Iterate through all movies
        for queryset in queryset:
            row_num += 1

            if queryset.ket_pending_dispute.last() is None:
                ket_pending_disput_queryset = ''
            else:
                ket_pending_disput_queryset = '{0}, '.format(queryset.ket_pending_dispute.last())

            if queryset.ket_jawaban_pending.last() is None:
                ket_jawaban_pending_queryset = ''
            else:
                ket_jawaban_pending_queryset = '{0}, '.format(queryset.ket_jawaban_pending.last())


            # ket_pending_disput_queryset = ''
            # for x in queryset.ket_pending_dispute.all():
            #     ket_pending_disput_queryset += '{0}, '.format(x.ket_pending_dispute)
            #
            # ket_jawaban_pending_queryset = ''
            # for x in queryset.ket_jawaban_pending.all():
            #     ket_jawaban_pending_queryset += '{0}, '.format(x.ket_jawaban_pending)

            # Define the data for each cell in the row
            row = [
                queryset.faskes.nama,
                queryset.status,
                queryset.register_klaim.nomor_register_klaim,
                queryset.NoSEPApotek,
                queryset.NoSEPAsalResep,
                queryset.TglResep,
                queryset.KdJenis,
                queryset.NoKartu,
                queryset.NamaPeserta,
                queryset.ByTagApt,
                queryset.ByVerApt,
                queryset.verifikator.username,
                queryset.rufil,
                queryset.jenis_pending,
                queryset.jenis_dispute,
                ket_pending_disput_queryset,
                ket_jawaban_pending_queryset,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)
        return response

    # pagination
    paginator = Paginator(queryset, 10)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'data_klaim': queryset,
        'myFilter': myFilter,
    }

    return render(request, 'faskes/daftar_data_klaim_pending_dispute_obat.html', context)


@login_required
@check_device
@permissions(role=['faskes'])
def detail_data_klaim_pending_dispute_obat(request, pk):
    queryset = DataKlaimObat.objects.filter(faskes=request.user.faskes_set.all().first(), prosesklaim=True,
                                            prosespending=True, prosestidaklayak=False)
    instance = queryset.get(pk=pk)
    data_klaim_form = DataKlaimCBGFaskesForm(instance=instance)
    jawaban_pending_dispute_form = JawabanPendingDisputeForm()

    if request.method == 'POST':
        data_klaim_form = DataKlaimObatFaskesForm(instance=instance, data=request.POST)
        jawaban = JawabanPendingDisputeForm(request.POST or None)
        if data_klaim_form.is_valid():
            data_klaim_form.save()
            # if jawaban_pending_dispute_form.is_valid():
            obj_jawaban = jawaban.save()
            obj_jawaban.user_faskes = request.user
            obj_jawaban.save()
            instance.ket_jawaban_pending.add(obj_jawaban)
            next = request.POST.get('next', '/')
            return HttpResponseRedirect(next)
    context = {
        'data_klaim': instance,
        'data_klaim_form': data_klaim_form,
        'jawaban_pending_dispute_form': jawaban_pending_dispute_form,
    }
    return render(request, 'faskes/detail_data_klaim_pending_dispute_obat.html', context)
