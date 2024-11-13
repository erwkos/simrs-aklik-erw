import datetime
import random
from datetime import timedelta
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Sum, Count, Q, Case, When, Value, CharField
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse_lazy
from openpyxl.workbook import Workbook

from faskes.models import KantorCabang
from klaim.choices import StatusRegisterChoices, StatusDataKlaimChoices, JenisPelayananChoices, NamaJenisKlaimChoices
from klaim.filters import RegisterKlaimFaskesFilter
from klaim.models import DataKlaimCBG, RegisterKlaim, SLA, DataKlaimObat
from supervisor.forms import PilihVerifikatorRegisterKlaimSupervisorForm, IsActiveForm, SLAUpdateForm, \
    SLACreateForm
from user.decorators import permissions, check_device
from user.models import User
from verifikator.filters import DownloadDataKlaimCBGFilter, DownloadDataKlaimObatFilter
from verifikator.forms import StatusRegisterKlaimForm
from collections import Counter, defaultdict


# Create your views here.

@login_required
@check_device
@permissions(role=['supervisor'])
def daftar_register_supervisor(request):
    queryset = RegisterKlaim.objects.filter(
        nomor_register_klaim__startswith=request.user.kantorcabang_set.all().first().kode_cabang).order_by('-tgl_aju')

    # filter
    myFilter = RegisterKlaimFaskesFilter(request.GET, queryset=queryset, request=request)
    queryset = myFilter.qs

    # pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'register_list': queryset,
        'myFilter': myFilter,
    }
    return render(request, 'verifikator/daftar_register.html', context)


@login_required
@check_device
@permissions(role=['supervisor'])
def detail_register_supervisor(request, pk):
    kantor_cabang = request.user.kantorcabang_set.all().first()
    queryset = RegisterKlaim.objects.filter(
        nomor_register_klaim__startswith=kantor_cabang.kode_cabang)
    instance = queryset.get(id=pk)
    status_form = PilihVerifikatorRegisterKlaimSupervisorForm(instance=instance)
    # pilih_verifikator_form = PilihVerifikatorRegisterKlaimSupervisorForm(instance=instance)
    status_form.fields['verifikator'].queryset = kantor_cabang.user.filter(
        groups__in=Group.objects.filter(name='verifikator'))
    # alasan_dikembalikan_form = AlasanDikembalikanForm(instance=instance)
    if request.method == 'POST': # and request.POST.get('action') == 'pilih_verifikator':
        pilih_verifikator_form = PilihVerifikatorRegisterKlaimSupervisorForm(instance=instance, data=request.POST)
        if pilih_verifikator_form.is_valid():
            pilih_verifikator_form.save()
            messages.success(request, "PIC Berhasil diganti. Selanjutnya dapat memberikan informasi ke Verifikator. "
                                      "Terima Kasih")
            return redirect(request.headers.get('Referer'))
    context = {
        'register': instance,
        'status_form': status_form

    }
    return render(request, 'verifikator/detail_register.html', context)


@login_required
@check_device
@permissions(role=['supervisor'])
def daftar_pembagian_ulang_verifikasi_cbg(request):
    queryset = RegisterKlaim.objects.filter(
        nomor_register_klaim__startswith=request.user.kantorcabang_set.all().first().kode_cabang,
        status=StatusRegisterChoices.VERIFIKASI,
        jenis_klaim__nama__in=(NamaJenisKlaimChoices.CBG_REGULER, NamaJenisKlaimChoices.CBG_SUSULAN),
    )
    # data_klaim = DataKlaimCBG.objects.filter(
    #     register_klaim__in=queryset
    # )

    # filter
    myFilter = RegisterKlaimFaskesFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    # pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'register_list': queryset,
        'myFilter': myFilter,
    }
    return render(request, 'supervisor/daftar_pembagian_ulang_verifikasi_cbg.html', context)


@login_required
@check_device
@permissions(role=['supervisor'])
def daftar_pembagian_ulang_verifikasi_obat(request):
    queryset = RegisterKlaim.objects.filter(
        nomor_register_klaim__startswith=request.user.kantorcabang_set.all().first().kode_cabang,
        status=StatusRegisterChoices.VERIFIKASI,
        jenis_klaim__nama__in=(NamaJenisKlaimChoices.OBAT_REGULER, NamaJenisKlaimChoices.OBAT_SUSULAN),
    )

    # filter
    myFilter = RegisterKlaimFaskesFilter(request.GET, queryset=queryset)
    queryset = myFilter.qs

    # pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    queryset = paginator.get_page(page_number)

    context = {
        'register_list': queryset,
        'myFilter': myFilter,
    }
    return render(request, 'supervisor/daftar_pembagian_ulang_verifikasi_obat.html', context)


@login_required
@check_device
@permissions(role=['supervisor'])
def update_pembagian_ulang_verifikasi_cbg(request, pk):
    kantor_cabang = request.user.kantorcabang_set.all().first()
    queryset_awal = RegisterKlaim.objects.filter(nomor_register_klaim__startswith=kantor_cabang.kode_cabang)
    register = queryset_awal.get(id=pk)
    verifikator_awal = register.faskes.kantor_cabang.user.filter(groups__name='verifikator')
    verifikator_bagi_ulang_per_verifikator = DataKlaimCBG.objects.filter(
        register_klaim=register,
        status=StatusDataKlaimChoices.PROSES
    ).values(
        'verifikator__id', 'verifikator__first_name', 'verifikator__last_name'
    ).distinct()

    # Count data klaim
    data_klaim = DataKlaimCBG.objects.filter(register_klaim=register)
    jumlah_proses = data_klaim.filter(status=StatusDataKlaimChoices.PROSES).count()
    jumlah_layak = data_klaim.filter(status=StatusDataKlaimChoices.LAYAK).count()
    jumlah_pending = data_klaim.filter(status=StatusDataKlaimChoices.PENDING).count()
    jumlah_dispute = data_klaim.filter(status=StatusDataKlaimChoices.DISPUTE).count()
    jumlah_tidak_layak = data_klaim.filter(status=StatusDataKlaimChoices.TIDAK_LAYAK).count()
    jumlah_klaim = data_klaim.filter(status=StatusDataKlaimChoices.KLAIM).count()
    total_klaim = data_klaim.count()

    # Calculate biaya
    biaya_proses = data_klaim.filter(status=StatusDataKlaimChoices.PROSES).aggregate(Sum('BYPENGAJUAN'))['BYPENGAJUAN__sum']
    biaya_layak = data_klaim.filter(status=StatusDataKlaimChoices.LAYAK).aggregate(Sum('BYPENGAJUAN'))['BYPENGAJUAN__sum']
    biaya_pending = data_klaim.filter(status=StatusDataKlaimChoices.PENDING).aggregate(Sum('BYPENGAJUAN'))['BYPENGAJUAN__sum']
    biaya_dispute = data_klaim.filter(status=StatusDataKlaimChoices.DISPUTE).aggregate(Sum('BYPENGAJUAN'))['BYPENGAJUAN__sum']
    biaya_tidak_layak = data_klaim.filter(status=StatusDataKlaimChoices.TIDAK_LAYAK).aggregate(Sum('BYPENGAJUAN'))['BYPENGAJUAN__sum']
    biaya_klaim = data_klaim.aggregate(Sum('BYPENGAJUAN'))['BYPENGAJUAN__sum']

    # Reassign all verifikators
    if request.method == 'POST' and request.POST.get('action') == 'bagi_ulang':
        try:
            verifikator_eksisting = register.faskes.kantor_cabang.user.filter(groups__name='verifikator')
            verifikator_ids = []
            for v in verifikator_eksisting:
                if request.POST.get(str(v.id)):
                    verifikator_ids.append(v.id)

            if not verifikator_ids:
                messages.error(request, "Tidak ada verifikator yang dipilih.")
                return redirect(request.path_info)

            num_verifikators = len(verifikator_ids)
            queryset = DataKlaimCBG.objects.filter(register_klaim=register, status=StatusDataKlaimChoices.PROSES)

            # Update JNSPEL based on KDINACBG if necessary
            with transaction.atomic():
                queryset_unknown_jnspel = queryset.filter(
                    ~Q(JNSPEL=JenisPelayananChoices.RAWAT_JALAN.value) &
                    ~Q(JNSPEL=JenisPelayananChoices.RAWAT_INAP.value)
                )

                if queryset_unknown_jnspel.exists():
                    queryset_unknown_jnspel.update(
                        JNSPEL=Case(
                            When(KDINACBG__endswith='I', then=Value(JenisPelayananChoices.RAWAT_INAP.value)),
                            When(KDINACBG__endswith='0', then=Value(JenisPelayananChoices.RAWAT_JALAN.value)),
                            default=Value(JenisPelayananChoices.UNKNOWN.value),
                            output_field=CharField(),
                        )
                    )

            # Assign participants per service type
            def assign_verifikator_per_service_type(jenis_pelayanan, reverse_order=False):
                qs_pelayanan = queryset.filter(JNSPEL=jenis_pelayanan.value)
                if not qs_pelayanan.exists():
                    return

                participants = qs_pelayanan.values_list('NMPESERTA', 'NOKARTU').distinct()
                participant_list = list(participants)
                random.shuffle(participant_list)

                verifikator_order = verifikator_ids[::-1] if reverse_order else verifikator_ids

                total_participants = len(participant_list)
                base_chunk_size = total_participants // num_verifikators
                remainder = total_participants % num_verifikators

                chunks = []
                start = 0
                for i in range(num_verifikators):
                    end = start + base_chunk_size + (1 if i < remainder else 0)
                    chunks.append(participant_list[start:end])
                    start = end

                verifikator_participant_map = {
                    verifikator_order[i]: chunks[i] for i in range(len(chunks))
                }

                with transaction.atomic():
                    for verifikator_id, participant_chunk in verifikator_participant_map.items():
                        if not participant_chunk:
                            continue
                        nmpeserta_list = [p[0] for p in participant_chunk]
                        nokartu_list = [p[1] for p in participant_chunk]
                        qs_pelayanan.filter(
                            NMPESERTA__in=nmpeserta_list,
                            NOKARTU__in=nokartu_list
                        ).update(
                            verifikator_id=verifikator_id,
                            status=StatusDataKlaimChoices.PROSES
                        )

            # Assign for RAWAT_JALAN
            assign_verifikator_per_service_type(
                JenisPelayananChoices.RAWAT_JALAN,
                reverse_order=False  # Assign in normal order
            )

            # Assign for RAWAT_INAP
            assign_verifikator_per_service_type(
                JenisPelayananChoices.RAWAT_INAP,
                reverse_order=True  # Assign in reverse order
            )

            messages.success(request, "Pembagian Ulang Klaim CBG Berhasil")
        except Exception as e:
            messages.warning(request, f"Pembagian Ulang Klaim CBG Gagal dengan keterangan: {e}")

    # Reassign per verifikator
    if request.method == 'POST' and request.POST.get('action') == 'bagi_ulang_per_verifikator':
        try:
            verifikator_eksisting = register.faskes.kantor_cabang.user.filter(groups__name='verifikator')
            verifikator_ids = []
            for v in verifikator_eksisting:
                if request.POST.get(str(v.id)):
                    verifikator_ids.append(v.id)
            verifikator_terpilih = request.POST.get('verifikator_terpilih')

            if not verifikator_ids:
                messages.error(request, "Tidak ada verifikator yang dipilih.")
                return redirect(request.path_info)

            num_verifikators = len(verifikator_ids)
            queryset = DataKlaimCBG.objects.filter(
                register_klaim=register,
                status=StatusDataKlaimChoices.PROSES,
                verifikator_id=verifikator_terpilih
            )

            # Update JNSPEL based on KDINACBG if necessary
            with transaction.atomic():
                queryset_unknown_jnspel = queryset.filter(
                    ~Q(JNSPEL=JenisPelayananChoices.RAWAT_JALAN.value) &
                    ~Q(JNSPEL=JenisPelayananChoices.RAWAT_INAP.value)
                )

                if queryset_unknown_jnspel.exists():
                    queryset_unknown_jnspel.update(
                        JNSPEL=Case(
                            When(KDINACBG__endswith='I', then=Value(JenisPelayananChoices.RAWAT_INAP.value)),
                            When(KDINACBG__endswith='0', then=Value(JenisPelayananChoices.RAWAT_JALAN.value)),
                            default=Value(JenisPelayananChoices.UNKNOWN.value),
                            output_field=CharField(),
                        )
                    )

            # Assign participants per service type
            def assign_verifikator_per_service_type(jenis_pelayanan, reverse_order=False):
                qs_pelayanan = queryset.filter(JNSPEL=jenis_pelayanan.value)
                if not qs_pelayanan.exists():
                    return

                participants = qs_pelayanan.values_list('NMPESERTA', 'NOKARTU').distinct()
                participant_list = list(participants)
                random.shuffle(participant_list)

                verifikator_order = verifikator_ids[::-1] if reverse_order else verifikator_ids

                total_participants = len(participant_list)
                base_chunk_size = total_participants // num_verifikators
                remainder = total_participants % num_verifikators

                chunks = []
                start = 0
                for i in range(num_verifikators):
                    end = start + base_chunk_size + (1 if i < remainder else 0)
                    chunks.append(participant_list[start:end])
                    start = end

                verifikator_participant_map = {
                    verifikator_order[i]: chunks[i] for i in range(len(chunks))
                }

                with transaction.atomic():
                    for verifikator_id, participant_chunk in verifikator_participant_map.items():
                        if not participant_chunk:
                            continue
                        nmpeserta_list = [p[0] for p in participant_chunk]
                        nokartu_list = [p[1] for p in participant_chunk]
                        qs_pelayanan.filter(
                            NMPESERTA__in=nmpeserta_list,
                            NOKARTU__in=nokartu_list
                        ).update(
                            verifikator_id=verifikator_id,
                            status=StatusDataKlaimChoices.PROSES
                        )

            # Assign for RAWAT_JALAN
            assign_verifikator_per_service_type(
                JenisPelayananChoices.RAWAT_JALAN,
                reverse_order=False
            )

            # Assign for RAWAT_INAP
            assign_verifikator_per_service_type(
                JenisPelayananChoices.RAWAT_INAP,
                reverse_order=True
            )

            messages.success(request, "Pembagian Ulang Klaim CBG Berhasil")
        except Exception as e:
            messages.warning(request, f"Pembagian Ulang Klaim CBG Gagal dengan keterangan: {e}")

    context = {
        'register': register,
        'verifikator': verifikator_awal,
        'verifikator_bagi': verifikator_bagi_ulang_per_verifikator,
        'jumlah_proses': jumlah_proses,
        'jumlah_layak': jumlah_layak,
        'jumlah_pending': jumlah_pending,
        'jumlah_dispute': jumlah_dispute,
        'jumlah_tidak_layak': jumlah_tidak_layak,
        'jumlah_klaim': jumlah_klaim,
        'total_klaim': total_klaim,
        'biaya_proses': biaya_proses,
        'biaya_layak': biaya_layak,
        'biaya_pending': biaya_pending,
        'biaya_dispute': biaya_dispute,
        'biaya_tidak_layak': biaya_tidak_layak,
        'biaya_klaim': biaya_klaim,
    }
    return render(request, 'supervisor/update_pembagian_ulang_verifikasi_cbg.html', context)

# def update_pembagian_ulang_verifikasi_cbg(request, pk):
#     kantor_cabang = request.user.kantorcabang_set.all().first()
#     queryset_awal = RegisterKlaim.objects.filter(nomor_register_klaim__startswith=kantor_cabang.kode_cabang)
#     register = queryset_awal.get(id=pk)
#     verifikator_awal = register.faskes.kantor_cabang.user.filter(groups__name='verifikator')
#     verifikator_bagi_ulang_per_verifikator = DataKlaimCBG.objects.filter(register_klaim=register, status=StatusDataKlaimChoices.PROSES).values(
#         'verifikator__id', 'verifikator__first_name', 'verifikator__last_name').distinct()
#
#     # count
#     data_klaim = DataKlaimCBG.objects.filter(register_klaim=register)
#     jumlah_proses = data_klaim.filter(status=StatusDataKlaimChoices.PROSES).count()
#     jumlah_layak = data_klaim.filter(status=StatusDataKlaimChoices.LAYAK).count()
#     jumlah_pending = data_klaim.filter(status=StatusDataKlaimChoices.PENDING).count()
#     jumlah_dispute = data_klaim.filter(status=StatusDataKlaimChoices.DISPUTE).count()
#     jumlah_tidak_layak = data_klaim.filter(status=StatusDataKlaimChoices.TIDAK_LAYAK).count()
#     jumlah_klaim = data_klaim.filter(status=StatusDataKlaimChoices.KLAIM).count()
#     total_klaim = data_klaim.count()
#
#     # biaya
#     biaya_proses = data_klaim.filter(status=StatusDataKlaimChoices.PROSES).aggregate(Sum('BYPENGAJUAN'))[
#         'BYPENGAJUAN__sum']
#     biaya_layak = data_klaim.filter(status=StatusDataKlaimChoices.LAYAK).aggregate(Sum('BYPENGAJUAN'))[
#         'BYPENGAJUAN__sum']
#     biaya_pending = data_klaim.filter(status=StatusDataKlaimChoices.PENDING).aggregate(Sum('BYPENGAJUAN'))[
#         'BYPENGAJUAN__sum']
#     biaya_dispute = data_klaim.filter(status=StatusDataKlaimChoices.DISPUTE).aggregate(Sum('BYPENGAJUAN'))[
#         'BYPENGAJUAN__sum']
#     biaya_tidak_layak = data_klaim.filter(status=StatusDataKlaimChoices.TIDAK_LAYAK).aggregate(Sum('BYPENGAJUAN'))[
#         'BYPENGAJUAN__sum']
#     biaya_klaim = data_klaim.aggregate(Sum('BYPENGAJUAN'))['BYPENGAJUAN__sum']
#
#     # Bagi Ulang semua verifikatoar
#     if request.method == 'POST' and request.POST.get('action') == 'bagi_ulang':
#         try:
#             verifikator_eksisting = register.faskes.kantor_cabang.user.filter(groups__name='verifikator')
#             verifikator = []
#             for i in verifikator_eksisting:
#                 get_verifikator = request.POST.get(str(i))
#                 if get_verifikator is not None:
#                     verifikator.append(get_verifikator)
#             queryset = DataKlaimCBG.objects.filter(register_klaim=register, status=StatusDataKlaimChoices.PROSES)
#
#             NMPESERTA_RJ = [obj.NMPESERTA for obj in queryset.filter(JNSPEL=JenisPelayananChoices.RAWAT_JALAN)]
#             list_nmpeserta_sort_freq_rawat_jalan = [item for items, c in Counter(NMPESERTA_RJ).most_common() for item in
#                                                     [items] * c]
#             list_nmpeserta_no_duplicate_rawat_jalan = list(dict.fromkeys(list_nmpeserta_sort_freq_rawat_jalan))
#
#             index = random.randrange(len(verifikator))
#             for i in range(len(list_nmpeserta_no_duplicate_rawat_jalan)):
#                 queryset.filter(NMPESERTA=list_nmpeserta_no_duplicate_rawat_jalan[i],
#                                 JNSPEL=JenisPelayananChoices.RAWAT_JALAN). \
#                     update(verifikator=verifikator[index], status=StatusDataKlaimChoices.PROSES)
#                 if index == len(verifikator) - 1:
#                     index = 0
#                 else:
#                     index += 1
#
#             NMPESERTA_RI = [obj.NMPESERTA for obj in queryset.filter(JNSPEL=JenisPelayananChoices.RAWAT_INAP)]
#             list_nmpeserta_sort_freq_rawat_inap = [item for items, c in Counter(NMPESERTA_RI).most_common() for item in
#                                                    [items] * c]
#             list_nmpeserta_no_duplicate_rawat_inap = list(dict.fromkeys(list_nmpeserta_sort_freq_rawat_inap))
#
#             index = random.randrange(len(verifikator))
#             for i in range(len(list_nmpeserta_no_duplicate_rawat_inap)):
#                 queryset.filter(NMPESERTA=list_nmpeserta_no_duplicate_rawat_inap[i],
#                                 JNSPEL=JenisPelayananChoices.RAWAT_INAP). \
#                     update(verifikator=verifikator[index], status=StatusDataKlaimChoices.PROSES)
#                 if index == len(verifikator) - 1:
#                     index = 0
#                 else:
#                     index += 1
#             messages.success(request, "Pembagian Ulang Klaim CBG Berhasil")
#         except Exception as e:
#             messages.warning(request, f"Pembagian Ulang Klaim CBG Gagal dengan keterangan : {e}")
#
#     # Bagi Ulang Per Verifikator
#     if request.method == 'POST' and request.POST.get('action') == 'bagi_ulang_per_verifikator':
#         try:
#             verifikator_eksisting = register.faskes.kantor_cabang.user.filter(groups__name='verifikator')
#             verifikator = []
#             for i in verifikator_eksisting:
#                 get_verifikator = request.POST.get(str(i))
#                 if get_verifikator is not None:
#                     verifikator.append(get_verifikator)
#             verifikator_terpilih = request.POST.get('verifikator_terpilih')
#             queryset = DataKlaimCBG.objects.filter(register_klaim=register, status=StatusDataKlaimChoices.PROSES,
#                                                    verifikator=verifikator_terpilih)
#
#             NMPESERTA_RJ = [obj.NMPESERTA for obj in queryset.filter(JNSPEL=JenisPelayananChoices.RAWAT_JALAN)]
#             list_nmpeserta_sort_freq_rawat_jalan = [item for items, c in Counter(NMPESERTA_RJ).most_common() for item in
#                                                     [items] * c]
#             list_nmpeserta_no_duplicate_rawat_jalan = list(dict.fromkeys(list_nmpeserta_sort_freq_rawat_jalan))
#
#             index = random.randrange(len(verifikator))
#             for i in range(len(list_nmpeserta_no_duplicate_rawat_jalan)):
#                 queryset.filter(NMPESERTA=list_nmpeserta_no_duplicate_rawat_jalan[i],
#                                 JNSPEL=JenisPelayananChoices.RAWAT_JALAN). \
#                     update(verifikator=verifikator[index], status=StatusDataKlaimChoices.PROSES)
#                 if index == len(verifikator) - 1:
#                     index = 0
#                 else:
#                     index += 1
#
#             NMPESERTA_RI = [obj.NMPESERTA for obj in queryset.filter(JNSPEL=JenisPelayananChoices.RAWAT_INAP)]
#             list_nmpeserta_sort_freq_rawat_inap = [item for items, c in Counter(NMPESERTA_RI).most_common() for item in
#                                                    [items] * c]
#             list_nmpeserta_no_duplicate_rawat_inap = list(dict.fromkeys(list_nmpeserta_sort_freq_rawat_inap))
#
#             index = random.randrange(len(verifikator))
#             for i in range(len(list_nmpeserta_no_duplicate_rawat_inap)):
#                 queryset.filter(NMPESERTA=list_nmpeserta_no_duplicate_rawat_inap[i],
#                                 JNSPEL=JenisPelayananChoices.RAWAT_INAP). \
#                     update(verifikator=verifikator[index], status=StatusDataKlaimChoices.PROSES)
#                 if index == len(verifikator) - 1:
#                     index = 0
#                 else:
#                     index += 1
#
#             messages.success(request, "Pembagian Ulang Klaim CBG Berhasil")
#         except Exception as e:
#             messages.warning(request, f"Pembagian Ulang Klaim CBG Gagal dengan Keterangan : {e}")
#
#     context = {
#         'register': register,
#         'verifikator': verifikator_awal,
#         'verifikator_bagi': verifikator_bagi_ulang_per_verifikator,
#         'jumlah_proses': jumlah_proses,
#         'jumlah_layak': jumlah_layak,
#         'jumlah_pending': jumlah_pending,
#         'jumlah_dispute': jumlah_dispute,
#         'jumlah_tidak_layak': jumlah_tidak_layak,
#         'jumlah_klaim': jumlah_klaim,
#         'total_klaim': total_klaim,
#         'biaya_proses': biaya_proses,
#         'biaya_layak': biaya_layak,
#         'biaya_pending': biaya_pending,
#         'biaya_dispute': biaya_dispute,
#         'biaya_tidak_layak': biaya_tidak_layak,
#         'biaya_klaim': biaya_klaim,
#     }
#     return render(request, 'supervisor/update_pembagian_ulang_verifikasi_cbg.html', context)


@login_required
@check_device
@permissions(role=['supervisor'])
def update_pembagian_ulang_verifikasi_cbg_per_verifikator(request, pk):
    kantor_cabang = request.user.kantorcabang_set.all().first()
    queryset_awal = RegisterKlaim.objects.filter(nomor_register_klaim__startswith=kantor_cabang.kode_cabang)
    register = queryset_awal.get(id=pk)
    verifikator_awal = register.faskes.kantor_cabang.user.filter(groups__name='verifikator')

    # count
    data_klaim = DataKlaimCBG.objects.filter(register_klaim=register)
    jumlah_proses = data_klaim.filter(status=StatusDataKlaimChoices.PROSES).count()
    jumlah_layak = data_klaim.filter(status=StatusDataKlaimChoices.LAYAK).count()
    jumlah_pending = data_klaim.filter(status=StatusDataKlaimChoices.PENDING).count()
    jumlah_dispute = data_klaim.filter(status=StatusDataKlaimChoices.DISPUTE).count()
    jumlah_tidak_layak = data_klaim.filter(status=StatusDataKlaimChoices.TIDAK_LAYAK).count()
    jumlah_klaim = data_klaim.filter(status=StatusDataKlaimChoices.KLAIM).count()
    total_klaim = data_klaim.count()

    # biaya
    biaya_proses = data_klaim.filter(status=StatusDataKlaimChoices.PROSES).aggregate(Sum('BYPENGAJUAN'))[
        'BYPENGAJUAN__sum']
    biaya_layak = data_klaim.filter(status=StatusDataKlaimChoices.LAYAK).aggregate(Sum('BYPENGAJUAN'))[
        'BYPENGAJUAN__sum']
    biaya_pending = data_klaim.filter(status=StatusDataKlaimChoices.PENDING).aggregate(Sum('BYPENGAJUAN'))[
        'BYPENGAJUAN__sum']
    biaya_dispute = data_klaim.filter(status=StatusDataKlaimChoices.DISPUTE).aggregate(Sum('BYPENGAJUAN'))[
        'BYPENGAJUAN__sum']
    biaya_tidak_layak = data_klaim.filter(status=StatusDataKlaimChoices.TIDAK_LAYAK).aggregate(Sum('BYPENGAJUAN'))[
        'BYPENGAJUAN__sum']
    biaya_klaim = data_klaim.aggregate(Sum('BYPENGAJUAN'))['BYPENGAJUAN__sum']

    if request.method == 'POST' and request.POST.get('action') == 'bagi_ulang':
        verifikator_eksisting = register.faskes.kantor_cabang.user.filter(groups__name='verifikator')
        verifikator = []
        for i in verifikator_eksisting:
            get_verifikator = request.POST.get(str(i))
            if get_verifikator is not None:
                verifikator.append(get_verifikator)
        verifikator_terpilih = request.POST.get('verifikator_terpilih')
        queryset = DataKlaimCBG.objects.filter(register_klaim=register, status=StatusDataKlaimChoices.PROSES,
                                               verifikator=verifikator_terpilih)

        NMPESERTA_RJ = [obj.NMPESERTA for obj in queryset.filter(JNSPEL=JenisPelayananChoices.RAWAT_JALAN)]
        list_nmpeserta_sort_freq_rawat_jalan = [item for items, c in Counter(NMPESERTA_RJ).most_common() for item in
                                                [items] * c]
        list_nmpeserta_no_duplicate_rawat_jalan = list(dict.fromkeys(list_nmpeserta_sort_freq_rawat_jalan))

        index = random.randrange(len(verifikator))
        for i in range(len(list_nmpeserta_no_duplicate_rawat_jalan)):
            queryset.filter(NMPESERTA=list_nmpeserta_no_duplicate_rawat_jalan[i],
                            JNSPEL=JenisPelayananChoices.RAWAT_JALAN). \
                update(verifikator=verifikator[index], status=StatusDataKlaimChoices.PROSES)
            if index == len(verifikator) - 1:
                index = 0
            else:
                index += 1

        NMPESERTA_RI = [obj.NMPESERTA for obj in queryset.filter(JNSPEL=JenisPelayananChoices.RAWAT_INAP)]
        list_nmpeserta_sort_freq_rawat_inap = [item for items, c in Counter(NMPESERTA_RI).most_common() for item in
                                               [items] * c]
        list_nmpeserta_no_duplicate_rawat_inap = list(dict.fromkeys(list_nmpeserta_sort_freq_rawat_inap))

        index = random.randrange(len(verifikator))
        for i in range(len(list_nmpeserta_no_duplicate_rawat_inap)):
            queryset.filter(NMPESERTA=list_nmpeserta_no_duplicate_rawat_inap[i],
                            JNSPEL=JenisPelayananChoices.RAWAT_INAP). \
                update(verifikator=verifikator[index], status=StatusDataKlaimChoices.PROSES)
            if index == len(verifikator) - 1:
                index = 0
            else:
                index += 1

        messages.success(request, "Pembagian Ulang Klaim CBG Berhasil")

    context = {
        'register': register,
        'verifikator': verifikator_awal,
        'jumlah_proses': jumlah_proses,
        'jumlah_layak': jumlah_layak,
        'jumlah_pending': jumlah_pending,
        'jumlah_dispute': jumlah_dispute,
        'jumlah_tidak_layak': jumlah_tidak_layak,
        'jumlah_klaim': jumlah_klaim,
        'total_klaim': total_klaim,
        'biaya_proses': biaya_proses,
        'biaya_layak': biaya_layak,
        'biaya_pending': biaya_pending,
        'biaya_dispute': biaya_dispute,
        'biaya_tidak_layak': biaya_tidak_layak,
        'biaya_klaim': biaya_klaim,
    }
    return render(request, 'supervisor/update_pembagian_ulang_verifikasi_cbg_per_verifikator.html', context)


@login_required
@check_device
@permissions(role=['supervisor'])
def api_json_pembagian_data_klaim_cbg(request):
    queryset = DataKlaimCBG.objects.filter(register_klaim__id=67).values('register_klaim__nomor_register_klaim',
                                                                         'bupel', 'faskes__nama',
                                                                         'verifikator__first_name', 'status', 'JNSPEL',
                                                                         'tgl_SLA')
    data = list(queryset)
    return JsonResponse(data, safe=False)


@login_required
@check_device
@permissions(role=['supervisor'])
def pembagian_ulang(request):
    """pembagian ulang

    pembagian ulang register claim ketika ada verifikator gagal verifikasi.
    Siapa verifikator yang dianggap gagal, ditentukan oleh supervisor.
    """
    kantor_cabang = request.user.kantorcabang_set.all().first()

    # pilih hanya verifikator
    group = Group.objects.filter(name='verifikator')
    verifikator = kantor_cabang.user.filter(
        groups__in=group,
        kantorcabang__kode_cabang=kantor_cabang.kode_cabang
    )
    qs_register = RegisterKlaim.objects.filter(
        status=StatusRegisterChoices.VERIFIKASI,
        verifikator__in=verifikator,
        nomor_register_klaim__startswith=kantor_cabang.kode_cabang
    ).order_by('-tgl_aju')
    data_claim = DataKlaimCBG.objects.filter(
        register_klaim__in=qs_register,
        status=StatusDataKlaimChoices.PROSES,
        verifikator__in=verifikator
    )

    # method GET dan modal on
    modal = request.GET.get('modal')

    # list id verifikator yang dipilih yang dianggap gagal verifikasi
    vrf = request.GET.get('vrf')

    # mengubah menjadi list dari str list jika tidak None
    vrf = list(map(int, vrf.split(','))) if vrf else None

    if request.method == 'POST':
        # id verifikator yang akan menerima tugas baru
        vrf_id = request.POST.getlist('verifikator')
        vrf_int = list(map(int, vrf_id))

        # register claim yang akan dipindah ke verifikator baru
        qs_claim = qs_register.filter(
            verifikator_id__in=vrf
        )

        # data claim yang akan dipindah ke verifikator baru
        dt_claim = data_claim.filter(
            register_klaim_id__in=qs_claim.values_list('id', flat=True)
        )

        if len(vrf_int) == qs_claim.count():
            # jumlah verifikator dan jumlah claim sama
            # tugas langsung dibagikan secara langsung
            for claim, verifikator_id in zip(qs_claim, vrf_int):
                claim.verifikator_id = verifikator_id
                claim.save()

                # update dataclaim dengan verifikator baru
                dt_claim.filter(
                    register_klaim=claim
                ).update(
                    verifikator_id=verifikator_id
                )
        else:
            # jumlah verifikator tidak sama dengan jumlah claim
            # bagikan tugas secara merata

            # taks 1 jika jumlah verifikator lebih dari jumlah claim
            # sehingga verifikator diharapkan mendapatkan minimal 1 tugas baru
            task_per_verifikator = 1 if len(vrf_int) > qs_claim.count() \
                else qs_claim.count() // len(vrf_int)

            # sisa tugas(claim) yang tidak habis dibagi
            sisa_task = qs_claim.count() % len(vrf_int)

            start_index = 0
            for verif_id in vrf_int:
                end_index = start_index + task_per_verifikator
                if sisa_task > 0:
                    end_index += 1
                    sisa_task -= 1

                task_for_verifikator = qs_claim[start_index:end_index]
                for claim in task_for_verifikator:
                    # assign verifikator baru
                    claim.verifikator_id = verif_id
                    claim.save()

                    # update dataclaim dengan verifikator baru
                    dt_claim.filter(
                        register_klaim=claim
                    ).update(
                        verifikator_id=verif_id
                    )
                start_index = end_index

            # kalau masih ada sisa task yang belum dibagi
            # bagikan sisanya ke verifikator dari awal lagi
            if sisa_task > 0:
                task_for_verifikator = qs_claim[:sisa_task]
                for i, claim in enumerate(task_for_verifikator):
                    verif_id = vrf_int[i % len(vrf_int)]
                    claim.verifikator_id = verif_id
                    claim.save()
                    dt_claim.filter(
                        register_klaim=claim
                    ).update(
                        verifikator_id=verif_id
                    )

        return redirect(reverse_lazy('supervisor:claim_bagi_ulang'))

    # menampilkan modal
    if modal == '1' and vrf:
        # mengeluarkan id verifikator yang gagal dari list
        verifikator = verifikator.exclude(
            pk__in=vrf
        )
        return render(
            request,
            'supervisor/bagiulang_proses.html',
            {
                'verifikator': verifikator,
                'title': 'Pembagian Tugas Baru'
            }
        )

    context = {}
    context["register_list"] = qs_register
    context['verifikator'] = verifikator

    return render(
        request,
        'supervisor/bagi_ulang_list.html',
        context
    )


@login_required
@check_device
@permissions(role=['supervisor'])
def list_user_verifikator(request):
    verifikator_group = Group.objects.filter(name='verifikator').first()
    user_verifikator = (User.objects.filter(groups=verifikator_group,
                                            kantorcabang__in=request.user.kantorcabang_set.all()).
                        order_by('-is_staff'))

    content = {
        'verifikator': user_verifikator,
    }
    return render(request, 'supervisor/list_user_verifikator.html', content)


@login_required
@check_device
@permissions(role=['supervisor'])
def edit_user_verifikator(request, pk):
    queryset = User.objects.filter(kantorcabang=request.user.kantorcabang_set.all().first())
    instance = queryset.get(id=pk)
    # user = User.objects.get(pk=pk)
    form = IsActiveForm(instance=instance)
    if request.method == 'POST':
        form = IsActiveForm(data=request.POST, instance=instance)
        if not form.has_changed():
            messages.warning(request, f'Status {instance} tidak ada perubahan')
            return redirect('supervisor:list_user_verifikator')
        elif form.is_valid():
            form.save()
            messages.success(request, f'Status {instance} berhasil diubah')
            return redirect('supervisor:list_user_verifikator')
        else:
            form = IsActiveForm()
            messages.warning(request, f'Status {instance} tidak berhasil diubah')

    content = {
        'form': form,
        'instance': instance
    }
    return render(request, 'supervisor/edit_user_verifikator.html', content)


# asal dari verifikator
@login_required
@check_device
@permissions(role=['supervisor'])
def download_data_cbg(request):
    # initial relasi pada kantor cabang
    related_kantor_cabang = request.user.kantorcabang_set.all()
    queryset = DataKlaimCBG.objects.filter(verifikator__kantorcabang__in=related_kantor_cabang)

    # filter
    myFilter = DownloadDataKlaimCBGFilter(request.GET, queryset=queryset, request=request)
    queryset = myFilter.qs

    kantor_cabang = request.user.kantorcabang_set.all().first()

    # fitur download
    download = request.GET.get('download')
    bupel_month = request.GET.get('bupel_month')
    bupel_year = request.GET.get('bupel_year')
    faskes = request.GET.get('faskes')
    nomor_register = request.GET.get('nomor_register_klaim')
    try:
        if nomor_register != '':
            if download == 'download':
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                response['Content-Disposition'] = 'attachment; filename={date}-dataverifikasi.xlsx'.format(
                    date=datetime.datetime.now().strftime('%Y-%m-%d'),
                )
                workbook = Workbook()

                # Get active worksheet/tab
                worksheet = workbook.active
                worksheet.title = 'Data Klaim CBG'

                # Define the titles for columns
                columns = [
                    'namars',
                    'status',
                    'NOREG',
                    'NOSEP',
                    'TGLSEP',
                    'TGLPULANG',
                    'JNSPEL',
                    'NOKARTU',
                    'NMPESERTA',
                    'POLI',
                    'KDINACBG',
                    'BYPENGAJUAN',
                    'verifikator',
                    'jenis_pending',
                    'jenis_dispute',
                    'ket_pending',
                    'ket_jawaban',
                ]
                row_num = 1

                # Assign the titles for each cell of the header
                for col_num, column_title in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title

                # Iterate through all movies
                for queryset in queryset:
                    row_num += 1

                    ket_pending_disput_queryset = ''
                    for x in queryset.ket_pending_dispute.all():
                        ket_pending_disput_queryset += '{0}, '.format(x.ket_pending_dispute)

                    ket_jawaban_pending_queryset = ''
                    for x in queryset.ket_jawaban_pending.all():
                        ket_jawaban_pending_queryset += '{0}, '.format(x.ket_jawaban_pending)

                    # Define the data for each cell in the row
                    row = [
                        queryset.faskes.nama,
                        queryset.status,
                        queryset.register_klaim.nomor_register_klaim,
                        queryset.NOSEP,
                        queryset.TGLSEP,
                        queryset.TGLPULANG,
                        queryset.JNSPEL,
                        queryset.NOKARTU,
                        queryset.NMPESERTA,
                        queryset.POLI,
                        queryset.KDINACBG,
                        queryset.BYPENGAJUAN,
                        queryset.verifikator.username,
                        queryset.jenis_pending,
                        queryset.jenis_dispute,
                        ket_pending_disput_queryset,
                        ket_jawaban_pending_queryset,
                    ]

                    # Assign the data for each cell of the row
                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value

                workbook.save(response)
                return response
        elif bupel_month != '' and bupel_year != '' and faskes != '':
            if download == 'download':
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                response['Content-Disposition'] = 'attachment; filename={date}-dataverifikasi.xlsx'.format(
                    date=datetime.datetime.now().strftime('%Y-%m-%d'),
                )
                workbook = Workbook()

                # Get active worksheet/tab
                worksheet = workbook.active
                worksheet.title = 'Data Klaim CBG'

                # Define the titles for columns
                columns = [
                    'namars',
                    'status',
                    'NOREG',
                    'NOSEP',
                    'TGLSEP',
                    'TGLPULANG',
                    'JNSPEL',
                    'NOKARTU',
                    'NMPESERTA',
                    'POLI',
                    'KDINACBG',
                    'BYPENGAJUAN',
                    'verifikator',
                    'jenis_pending',
                    'jenis_dispute',
                    'ket_pending',
                    'ket_jawaban',
                ]
                row_num = 1

                # Assign the titles for each cell of the header
                for col_num, column_title in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title

                # Iterate through all movies
                for queryset in queryset:
                    row_num += 1

                    ket_pending_disput_queryset = ''
                    for x in queryset.ket_pending_dispute.all():
                        ket_pending_disput_queryset += '{0}, '.format(x.ket_pending_dispute)

                    ket_jawaban_pending_queryset = ''
                    for x in queryset.ket_jawaban_pending.all():
                        ket_jawaban_pending_queryset += '{0}, '.format(x.ket_jawaban_pending)

                    # Define the data for each cell in the row
                    row = [
                        queryset.faskes.nama,
                        queryset.status,
                        queryset.register_klaim.nomor_register_klaim,
                        queryset.NOSEP,
                        queryset.TGLSEP,
                        queryset.TGLPULANG,
                        queryset.JNSPEL,
                        queryset.NOKARTU,
                        queryset.NMPESERTA,
                        queryset.POLI,
                        queryset.KDINACBG,
                        queryset.BYPENGAJUAN,
                        queryset.verifikator.username,
                        queryset.jenis_pending,
                        queryset.jenis_dispute,
                        ket_pending_disput_queryset,
                        ket_jawaban_pending_queryset,
                    ]

                    # Assign the data for each cell of the row
                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value

                workbook.save(response)
                return response
        else:
            messages.warning(request, 'Bulan, Tahun, dan Faskes harus diisi atau Nomor Register Klaim harus diisi!')
    except Exception as e:
        messages.warning(request, "Terjadi Kesalahan Dalam Download Data, dengan Keterangan: " + str(e))

    context = {
        'myFilter': myFilter,
    }
    return render(request, 'verifikator/cbg/download_data_cbg.html', context)


@login_required
@check_device
@permissions(role=['supervisor'])
def list_pengaturan_sla(request):
    queryset = SLA.objects.filter(kantor_cabang=request.user.kantorcabang_set.all().first())

    # # ini harus dihapus
    # query = DataKlaimCBG.objects.filter(register_klaim__nomor_register_klaim='020123090001')
    # query.update(verifikator=None, status='Belum Ver', tgl_SLA=None, JNSPEL=None)

    context = {
        'queryset': queryset
    }

    return render(request, 'supervisor/list_pengaturan_sla.html', context)


@login_required
@check_device
@permissions(role=['supervisor'])
def add_pengaturan_sla(request):
    form = SLACreateForm(request.POST or None)
    if form.is_valid():
        queryset = SLA.objects.filter(kantor_cabang=request.user.kantorcabang_set.all().first(),
                                      jenis_klaim=form.cleaned_data['jenis_klaim']).first()
        if queryset is not None:
            messages.warning(request, f'Data SLA {queryset.jenis_klaim} Sudah Ada')
            return redirect('supervisor:list_pengaturan_sla')
        elif queryset is None:
            data = form.cleaned_data
            data['kantor_cabang'] = request.user.kantorcabang_set.all().first()
            SLA.objects.create(**data)
            messages.success(request, f'Data SLA {data["jenis_klaim"]} Berhasil Dibuat')
            return redirect('supervisor:list_pengaturan_sla')

    context = {
        'form': form
    }

    return render(request, 'supervisor/add_pengaturan_sla.html', context)


@login_required
@check_device
@permissions(role=['supervisor'])
def edit_pengaturan_sla(request, pk):
    instance = SLA.objects.get(id=pk)
    form = SLAUpdateForm(request.POST or None, instance=instance)
    if form.is_valid():
        form.save()
        messages.success(request, f'Data SLA {instance.jenis_klaim.nama} Berhasil Diubah Menjadi Tgl BA Lengkap + {instance.plus_hari_sla} Hari')
        return redirect('supervisor:list_pengaturan_sla')

    context = {
        'form': form,
        'instance': instance
    }

    return render(request, 'supervisor/edit_pengaturan_sla.html', context)


@login_required
@check_device
@permissions(role=['supervisor'])
def download_data_obat(request):
    # initial relasi pada kantor cabang
    related_kantor_cabang = request.user.kantorcabang_set.all()
    queryset = DataKlaimObat.objects.filter(verifikator__kantorcabang__in=related_kantor_cabang)

    # filter
    myFilter = DownloadDataKlaimObatFilter(request.GET, queryset=queryset, request=request)
    queryset = myFilter.qs

    kantor_cabang = request.user.kantorcabang_set.all().first()

    # fitur download
    download = request.GET.get('download')
    bupel_month = request.GET.get('bupel_month')
    bupel_year = request.GET.get('bupel_year')
    faskes = request.GET.get('faskes')
    nomor_register = request.GET.get('nomor_register_klaim')
    try:
        if nomor_register != '':
            if download == 'download':
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                response['Content-Disposition'] = 'attachment; filename={date}-dataverifikasi.xlsx'.format(
                    date=datetime.datetime.now().strftime('%Y-%m-%d'),
                )
                workbook = Workbook()

                # Get active worksheet/tab
                worksheet = workbook.active
                worksheet.title = 'Data Klaim Obat'

                # Define the titles for columns
                columns = [
                    'namars',
                    'status',
                    'NoReg',
                    'NoSEPApotek',
                    'NoSEPAsalResep',
                    'TglResep',
                    'KdJenis',
                    'NoKartu',
                    'NamaPeserta',
                    'ByTagApt',
                    'ByVerApt',
                    'verifikator',
                    'rufil',
                    'jenis_pending',
                    'jenis_dispute',
                    'ket_pending',
                    'ket_jawaban',
                ]
                row_num = 1

                # Assign the titles for each cell of the header
                for col_num, column_title in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title

                # Iterate through all movies
                for queryset in queryset:
                    row_num += 1

                    ket_pending_disput_queryset = ''
                    for x in queryset.ket_pending_dispute.all():
                        ket_pending_disput_queryset += '{0}, '.format(x.ket_pending_dispute)

                    ket_jawaban_pending_queryset = ''
                    for x in queryset.ket_jawaban_pending.all():
                        ket_jawaban_pending_queryset += '{0}, '.format(x.ket_jawaban_pending)

                    # Define the data for each cell in the row
                    row = [
                        queryset.faskes.nama,
                        queryset.status,
                        queryset.register_klaim.nomor_register_klaim,
                        queryset.NoSEPApotek,
                        queryset.NoSEPAsalResep,
                        queryset.TglResep,
                        queryset.KdJenis,
                        queryset.NoKartu,
                        queryset.NamaPeserta,
                        queryset.ByTagApt,
                        queryset.ByVerApt,
                        queryset.verifikator.username,
                        queryset.rufil,
                        queryset.jenis_pending,
                        queryset.jenis_dispute,
                        ket_pending_disput_queryset,
                        ket_jawaban_pending_queryset,
                    ]

                    # Assign the data for each cell of the row
                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value

                workbook.save(response)
                return response
        if bupel_month != '' and bupel_year != '' and faskes != '':
            if download == 'download':
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                response['Content-Disposition'] = 'attachment; filename={date}-dataverifikasi.xlsx'.format(
                    date=datetime.datetime.now().strftime('%Y-%m-%d'),
                )
                workbook = Workbook()

                # Get active worksheet/tab
                worksheet = workbook.active
                worksheet.title = 'Data Klaim Obat'

                # Define the titles for columns
                columns = [
                    'namars',
                    'status',
                    'NoReg',
                    'NoSEPApotek',
                    'NoSEPAsalResep',
                    'TglResep',
                    'KdJenis',
                    'NoKartu',
                    'NamaPeserta',
                    'ByTagApt',
                    'ByVerApt',
                    'verifikator',
                    'rufil',
                    'jenis_pending',
                    'jenis_dispute',
                    'ket_pending',
                    'ket_jawaban',
                ]
                row_num = 1

                # Assign the titles for each cell of the header
                for col_num, column_title in enumerate(columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = column_title

                # Iterate through all movies
                for queryset in queryset:
                    row_num += 1

                    ket_pending_disput_queryset = ''
                    for x in queryset.ket_pending_dispute.all():
                        ket_pending_disput_queryset += '{0}, '.format(x.ket_pending_dispute)

                    ket_jawaban_pending_queryset = ''
                    for x in queryset.ket_jawaban_pending.all():
                        ket_jawaban_pending_queryset += '{0}, '.format(x.ket_jawaban_pending)

                    # Define the data for each cell in the row
                    row = [
                        queryset.faskes.nama,
                        queryset.status,
                        queryset.register_klaim.nomor_register_klaim,
                        queryset.NoSEPApotek,
                        queryset.NoSEPAsalResep,
                        queryset.TglResep,
                        queryset.KdJenis,
                        queryset.NoKartu,
                        queryset.NamaPeserta,
                        queryset.ByTagApt,
                        queryset.ByVerApt,
                        queryset.verifikator.username,
                        queryset.rufil,
                        queryset.jenis_pending,
                        queryset.jenis_dispute,
                        ket_pending_disput_queryset,
                        ket_jawaban_pending_queryset,
                    ]

                    # Assign the data for each cell of the row
                    for col_num, cell_value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = cell_value

                workbook.save(response)
                return response
        else:
            messages.warning(request, 'Bulan, Tahun, dan Faskes harus diisi!')
    except Exception as e:
        messages.warning(request, "Terjadi Kesalahan Dalam Download Data, dengan Keterangan: " + str(e))

    context = {
        'myFilter': myFilter,
    }
    return render(request, 'supervisor/download_data_obat.html', context)


@login_required
@check_device
@permissions(role=['supervisor'])
def pembagian_verifikator_cbg_null(request):
    # Get the list of Kantor Cabang associated with the user
    kantor_cabang_list = request.user.kantorcabang_set.all()
    nomor_register = request.GET.get('nomor_register')
    pembagian_null = request.GET.get('pembagian_null')

    # Queryset of DataKlaimCBG with status BELUM_VER
    queryset = DataKlaimCBG.objects.filter(
        register_klaim__faskes__kantor_cabang__in=kantor_cabang_list,
        register_klaim__nomor_register_klaim=nomor_register,
        status=StatusDataKlaimChoices.BELUM_VER,
    )

    # Get the RegisterKlaim instance
    register = RegisterKlaim.objects.filter(
        faskes__kantor_cabang__in=kantor_cabang_list,
        nomor_register_klaim=nomor_register
    ).first()

    # Get the list of verifikators
    verifikators = list(User.objects.filter(
        kantorcabang__in=kantor_cabang_list,
        groups__name='verifikator',
        is_active=True,
        is_staff=True
    ))

    if not verifikators:
        messages.error(request, "Tidak ada verifikator yang tersedia.")
        return redirect(request.headers.get('Referer'))

    if request.method == 'GET' and pembagian_null == "pembagian_null" and queryset.exists():
        try:
            verifikator_ids = [v.id for v in verifikators]
            num_verifikators = len(verifikator_ids)

            # Update JNSPEL where it's neither 'Rawat Jalan' nor 'Rawat Inap'
            with transaction.atomic():
                queryset_unknown_jnspel = queryset.filter(
                    ~Q(JNSPEL=JenisPelayananChoices.RAWAT_JALAN.value) &
                    ~Q(JNSPEL=JenisPelayananChoices.RAWAT_INAP.value)
                )

                if queryset_unknown_jnspel.exists():
                    queryset_unknown_jnspel.update(
                        JNSPEL=Case(
                            When(KDINACBG__endswith='I', then=Value(JenisPelayananChoices.RAWAT_INAP.value)),
                            When(KDINACBG__endswith='0', then=Value(JenisPelayananChoices.RAWAT_JALAN.value)),
                            default=Value(None),
                            output_field=CharField(),
                        )
                    )

            # Dictionary to track total assignments per verifikator
            verifikator_assignments = {
                vid: {
                    JenisPelayananChoices.RAWAT_JALAN.value: 0,
                    JenisPelayananChoices.RAWAT_INAP.value: 0
                } for vid in verifikator_ids
            }

            # Assign participants per JENIS_PELAYANAN
            def assign_verifikator_per_service_type(jenis_pelayanan, reverse_order=False):
                qs_pelayanan = queryset.filter(JNSPEL=jenis_pelayanan.value)
                if not qs_pelayanan.exists():
                    return

                # Get unique participants
                participants = qs_pelayanan.values_list('NMPESERTA', 'NOKARTU').distinct()
                participant_list = list(participants)

                # Shuffle the participant list
                random.shuffle(participant_list)

                # Determine assignment order
                if reverse_order:
                    verifikator_order = verifikator_ids[::-1]
                else:
                    verifikator_order = verifikator_ids

                # Calculate chunk sizes
                total_participants = len(participant_list)
                base_chunk_size = total_participants // num_verifikators
                remainder = total_participants % num_verifikators

                # Assign participants to verifikators
                chunks = []
                start = 0
                for i in range(num_verifikators):
                    end = start + base_chunk_size + (1 if i < remainder else 0)
                    chunks.append(participant_list[start:end])
                    start = end

                # Map verifikator IDs to participant chunks
                verifikator_participant_map = {
                    verifikator_order[i]: chunks[i] for i in range(num_verifikators)
                }

                # Perform bulk updates
                with transaction.atomic():
                    for verifikator_id, participant_chunk in verifikator_participant_map.items():
                        if not participant_chunk:
                            continue
                        nmpeserta_list = [p[0] for p in participant_chunk]
                        nokartu_list = [p[1] for p in participant_chunk]
                        qs_pelayanan.filter(
                            NMPESERTA__in=nmpeserta_list,
                            NOKARTU__in=nokartu_list,
                        ).update(
                            verifikator_id=verifikator_id,
                            status=StatusDataKlaimChoices.PROSES
                        )
                        # Update assignment counts
                        verifikator_assignments[verifikator_id][jenis_pelayanan.value] += len(participant_chunk)

            # Assign for RAWAT_JALAN
            assign_verifikator_per_service_type(
                JenisPelayananChoices.RAWAT_JALAN,
                reverse_order=False  # Assign in normal order
            )

            # Assign for RAWAT_INAP
            assign_verifikator_per_service_type(
                JenisPelayananChoices.RAWAT_INAP,
                reverse_order=True  # Assign in reverse order
            )

            # Update tgl_SLA for the processed claims
            queryset_proses = DataKlaimCBG.objects.filter(
                register_klaim=register,
                status=StatusDataKlaimChoices.PROSES
            )

            sla = SLA.objects.filter(
                jenis_klaim=register.jenis_klaim,
                kantor_cabang=register.faskes.kantor_cabang
            ).first()

            if sla:
                sla_days = sla.plus_hari_sla
            else:
                sla_days = 6  # Default SLA days

            if register.tgl_ba_lengkap:
                tgl_sla = register.tgl_ba_lengkap + timedelta(days=sla_days)
            elif register.tgl_terima:
                tgl_sla = register.tgl_terima + timedelta(days=15)
            else:
                tgl_sla = None

            if tgl_sla:
                queryset_proses.update(tgl_SLA=tgl_sla)

            messages.success(request, "Pembagian Verifikator Berhasil")

        except Exception as e:
            messages.warning(request, f"Pembagian Verifikator Gagal: {e}")
            return redirect(request.headers.get('Referer'))

    context = {
        'verifikator': verifikators
    }

    return render(request, 'supervisor/pembagian_verifikator_null.html', context)

# def pembagian_verifikator_cbg_null(request):
#     kantor_cabang_list = request.user.kantorcabang_set.all()
#     nomor_register = request.GET.get('nomor_register')
#     pembagian_null = request.GET.get('pembagian_null')
#     queryset = DataKlaimCBG.objects.filter(register_klaim__faskes__kantor_cabang__in=kantor_cabang_list,
#                                            register_klaim__nomor_register_klaim=nomor_register,
#                                            status=StatusDataKlaimChoices.BELUM_VER,
#                                            )
#     register = RegisterKlaim.objects.filter(faskes__kantor_cabang__in=kantor_cabang_list,
#                                             nomor_register_klaim=nomor_register).first()
#
#     if request.method == 'GET' and pembagian_null == "pembagian_null" and queryset.exists():
#         try:
#             verifikator = User.objects.filter(kantorcabang__in=kantor_cabang_list,
#                                               groups__name='verifikator',
#                                               is_active=True,
#                                               is_staff=True)
#
#             NMPESERTA_RJ = [obj.NMPESERTA for obj in queryset.filter(JNSPEL=JenisPelayananChoices.RAWAT_JALAN)]
#             list_nmpeserta_sort_freq_rawat_jalan = [item for items, c in Counter(NMPESERTA_RJ).most_common() for
#                                                     item in
#                                                     [items] * c]
#             list_nmpeserta_no_duplicate_rawat_jalan = list(dict.fromkeys(list_nmpeserta_sort_freq_rawat_jalan))
#             index = random.randrange(len(verifikator))
#
#             with transaction.atomic():
#                 for i in range(len(list_nmpeserta_no_duplicate_rawat_jalan)):
#                     queryset.filter(NMPESERTA=list_nmpeserta_no_duplicate_rawat_jalan[i],
#                                     JNSPEL=JenisPelayananChoices.RAWAT_JALAN). \
#                         update(verifikator=verifikator[index], status=StatusDataKlaimChoices.PROSES)
#                     if index == len(verifikator) - 1:
#                         index = 0
#                     else:
#                         index += 1
#
#             NMPESERTA_RI = [obj.NMPESERTA for obj in queryset.filter(JNSPEL=JenisPelayananChoices.RAWAT_INAP)]
#             list_nmpeserta_sort_freq_rawat_inap = [item for items, c in Counter(NMPESERTA_RI).most_common() for item
#                                                    in
#                                                    [items] * c]
#             list_nmpeserta_no_duplicate_rawat_inap = list(dict.fromkeys(list_nmpeserta_sort_freq_rawat_inap))
#             index = random.randrange(len(verifikator))
#
#             with transaction.atomic():
#                 for i in range(len(list_nmpeserta_no_duplicate_rawat_inap)):
#                     queryset.filter(NMPESERTA=list_nmpeserta_no_duplicate_rawat_inap[i],
#                                     JNSPEL=JenisPelayananChoices.RAWAT_INAP). \
#                         update(verifikator=verifikator[index], status=StatusDataKlaimChoices.PROSES)
#                     if index == len(verifikator) - 1:
#                         index = 0
#                     else:
#                         index += 1
#
#             queryset_proses = DataKlaimCBG.objects.filter(register_klaim=register, status=StatusDataKlaimChoices.PROSES)
#
#             # penentuan SLA
#             sla = SLA.objects.filter(jenis_klaim=register.jenis_klaim,
#                                      kantor_cabang=register.faskes.kantor_cabang).first()
#             if sla:
#                 if register.tgl_ba_lengkap:
#                     queryset_proses.update(tgl_SLA=register.tgl_ba_lengkap + timedelta(days=sla.plus_hari_sla))
#                 elif register.tgl_terima:
#                     queryset_proses.update(tgl_SLA=register.tgl_terima + timedelta(days=15))
#             else:
#                 if register.tgl_ba_lengkap:
#                     queryset_proses.update(tgl_SLA=register.tgl_ba_lengkap + timedelta(days=6))
#                 elif register.tgl_terima:
#                     queryset_proses.update(tgl_SLA=register.tgl_terima + timedelta(days=15))
#
#             messages.success(request, "Pembagian Ulang Verifikator Null Berhasil")
#
#         except Exception as e:
#             messages.warning(request, f"Pembagian Ulang Verifikator Null Gagal: {e}")
#             return redirect(request.headers.get('Referer'))
#
#     context = {
#         'data_klaim': queryset
#     }
#
#     return render(request, 'supervisor/pembagian_verifikator_null.html', context)